import json
import logging
import os
from datetime import datetime
from datetime import timezone, timedelta
IRAQ_TZ = timezone(timedelta(hours=3))  # UTC+3
from typing import Optional
from telegram import (
    Update,
    InlineKeyboardButton, InlineKeyboardMarkup,
    KeyboardButton, ReplyKeyboardMarkup,
)
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ConversationHandler, ContextTypes
)

# ======================= KONFIGURASYON =======================
TOKEN          = "8574020136:AAEaxD4fk9DPQQPsZ0y0lkhdKZXVrONxQJU"
ADMIN_ID       = 7731559022
ANON_GROUP_ID  = None  # Anonim mesaj grubu — Ayarlar'dan set edilir


_default_dir = "/data/bot_data" if os.path.exists("/data") else os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "bot_data")
BASE_DIR  = os.environ.get("DATA_DIR", _default_dir)
FILES_DIR = os.path.join(BASE_DIR, "files")
os.makedirs(BASE_DIR,  exist_ok=True)
os.makedirs(FILES_DIR, exist_ok=True)

DATA_FILE        = os.path.join(BASE_DIR, "bot_data.json")
USERS_FILE       = os.path.join(BASE_DIR, "users.json")
MESSAGES_FILE    = os.path.join(BASE_DIR, "user_messages.json")
SETTINGS_FILE    = os.path.join(BASE_DIR, "settings.json")
ADMINS_FILE      = os.path.join(BASE_DIR, "admins.json")
BLOCKED_FILE     = os.path.join(BASE_DIR, "blocked.json")
VIEW_COUNTS_FILE = os.path.join(BASE_DIR, "view_counts.json")
POLLS_FILE       = os.path.join(BASE_DIR, "polls.json")
CLASS_FILE       = os.path.join(BASE_DIR, "user_classes.json")
FAQ_FILE         = os.path.join(BASE_DIR, "faq.json")
AUTO_RULES_FILE  = os.path.join(BASE_DIR, "auto_rules.json")
FAVORITES_FILE   = os.path.join(BASE_DIR, "favorites.json")
NOTES_FILE       = os.path.join(BASE_DIR, "user_notes.json")
WARNS_FILE       = os.path.join(BASE_DIR, "user_warns.json")
ADMIN_LOG_FILE   = os.path.join(BASE_DIR, "admin_log.json")
SUBS_FILE        = os.path.join(BASE_DIR, "subscriptions.json")
RECENT_FILE      = os.path.join(BASE_DIR, "recently_viewed.json")
FOLDER_DESC_FILE = os.path.join(BASE_DIR, "folder_descs.json")
WHITELIST_FILE   = os.path.join(BASE_DIR, "whitelist.json")
SCHEDULED_FILE   = os.path.join(BASE_DIR, "scheduled.json")
TAGS_FILE        = os.path.join(BASE_DIR, "file_tags.json")
REMINDERS_FILE   = os.path.join(BASE_DIR, "reminders.json")
SEARCH_CACHE_FILE= os.path.join(BASE_DIR, "search_cache.json")
LEADERBOARD_FILE = os.path.join(BASE_DIR, "leaderboard.json")
ACHIEVEMENTS_FILE= os.path.join(BASE_DIR, "achievements.json")
FEEDBACK_FILE    = os.path.join(BASE_DIR, "feedback.json")
PINNED_MSG_FILE  = os.path.join(BASE_DIR, "pinned_msgs.json")
BROADCAST_LOG_FILE=os.path.join(BASE_DIR, "broadcast_log.json")
COUNTDOWN_FILE   = os.path.join(BASE_DIR, "countdowns.json")
GROUP_FILE       = os.path.join(BASE_DIR, "user_groups.json")
ANON_Q_FILE      = os.path.join(BASE_DIR, "anon_questions.json")
USER_NOTES_FILE  = os.path.join(BASE_DIR, "personal_notes.json")
QUIZ_FILE        = os.path.join(BASE_DIR, "quizzes.json")
REPORT_FILE      = os.path.join(BASE_DIR, "file_reports.json")
ADMIN_PERMS_FILE = os.path.join(BASE_DIR, "admin_perms.json")

# ── Arama / AI ayarları ──────────────────────────────────────
SEARCH_TIMEOUT   = 10   # saniye
SEARCH_CACHE_TTL = 3600 # 1 saat cache
MATH_SUBJECTS_AR = {
    "رياضيات","math","حساب","جبر","هندسة","مثلثات","تفاضل","تكامل",
    "معادلة","مشتقة","نهايات","مصفوفة","احتمال","إحصاء",
}
PHYSICS_SUBJECTS_AR = {
    "فيزياء","physics","ميكانيك","كهرباء","مغناطيس","موجات",
    "حرارة","ضوء","ذرة","نووي","قوة","طاقة","سرعة","تسارع",
}
CHEMISTRY_SUBJECTS_AR = {
    "كيمياء","chemistry","عناصر","مركبات","تفاعل","معادلة كيميائية",
    "حمض","قاعدة","اكسيد","كيمياء عضوية",
}
ALL_ACADEMIC_SUBJECTS = MATH_SUBJECTS_AR | PHYSICS_SUBJECTS_AR | CHEMISTRY_SUBJECTS_AR | {
    "برمجة","programming","algorithm","خوارزمية",
    "دوائر","circuits","إلكترونيات","electronics",
    "ميكاترونيك","mechatronics","تحكم","control",
    "مقاومة مواد","strength of materials","ديناميكا","dynamics",
    "احتراق","combustion","ترموديناميك","thermodynamics",
    "ميكانيكا طيران","fluid mechanics","اهتزازات",
}

MAX_WARNS        = 3
RECENT_MAX       = 15
RATE_LIMIT_SEC   = 3   # kullanıcı başına min mesaj aralığı
SPAM_LIMIT       = 10  # 60 saniyede max mesaj sayısı

# Sınıf tanımları (Türkçe ve Arapça)
CLASS_DEFS = {
    "1": {"tr": "1. Sınıf",  "ar": "الأول"},
    "2": {"tr": "2. Sınıf",  "ar": "الثاني"},
    "3": {"tr": "3. Sınıf",  "ar": "الثالث"},
    "4": {"tr": "4. Sınıf",  "ar": "الرابع"},
}

_rate_cache: dict = {}   # {uid: [timestamp, ...]}  — in-memory rate limit
# =============================================================

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO)
logger = logging.getLogger(__name__)

# ================================================================
#  DİL SİSTEMİ
#  Süper Admin (ADMIN_ID) → Türkçe
#  Diğer adminler + kullanıcılar → Arapça
# ================================================================

TR = {
    # Klavye butonları
    "btn_content":   "📁 İçerik",
    "btn_mgmt":      "👥 Yönetim",
    "btn_settings":  "⚙️ Ayarlar",
    "btn_maint":     "🔧 Bakım",
    "btn_search":    "🔍 Ara",
    "btn_help":      "💬 Mesaj Gönder",
    # Genel
    "home":          "🏠 ANA SAYFA",
    "folder_list":   "📁 Klasörler:",
    "file_list":     "📎 Dosyalar:",
    "back":          "◀️ Geri",
    "cancel":        "❌ İptal",
    "close":         "✖️ Kapat",
    "maint_on":      "🔧 Bakım Modu: AÇIK",
    "maint_off":     "✅ Bakım Modu: KAPALI",
    # İçerik
    "content_mgmt":         "📁 İçerik Yönetimi",
    "add_folder":           "➕ Klasör Ekle",
    "add_file":             "📎 Dosya/Medya Ekle",
    "enter_folder_name":    "📁 Yeni klasör adını yazın:",
    "no_folders":           "Klasör yok.",
    "no_files":             "Dosya yok.",
    "add_file_prompt":       "📎 Dosya, resim, video gönderin veya link/metin yazın:\n\n📋 Bu klasördeki mevcut dosyalar:\n{}",
    "add_file_prompt_empty": "📎 Dosya, resim, video gönderin veya link/metin yazın:\n\n📭 Bu klasörde henüz dosya yok.",
    "del_folder":           "🗑 Klasör Sil",
    "rename_folder":        "✏️ Klasör Yeniden Adlandır",
    "del_file":             "🗑 Dosya Sil",
    "rename_file":          "✏️ Dosya Yeniden Adlandır",
    "del_folder_select":    "Silinecek klasörü seçin:",
    "rename_folder_select": "Yeniden adlandırılacak klasörü seçin:",
    "del_file_select":      "Silinecek dosyayı seçin:",
    "rename_file_select":   "Yeniden adlandırılacak dosyayı seçin:",
    # Onay
    "del_folder_confirm": "🗑 *'{}'* klasörünü silmek istediğinizden emin misiniz?\n⚠️ İçindeki tüm dosya ve alt klasörler de silinecek!",
    "del_file_confirm":   "🗑 *'{}'* dosyasını silmek istediğinizden emin misiniz?",
    "confirm_yes":        "✅ Evet, Sil",
    "confirm_no":         "❌ İptal",
    # Klasör/dosya işlemleri
    "folder_exists":  "❌ '{}' zaten var!",
    "folder_empty":   "❌ Ad boş olamaz!",
    "folder_added":   "✅ '{}' klasörü eklendi.",
    "folder_deleted": "✅ '{}' silindi.",
    "folder_renamed": "✅ '{}' → '{}'",
    "new_folder_name": "'{}' için yeni adı yazın:",
    "file_added":        "✅ Eklendi: {}",
    "file_exists_warn":  "⚠️ '{}' adında bir dosya zaten mevcut! Yine de eklendi.",
    "file_deleted":      "✅ '{}' silindi.",
    "file_renamed":      "✅ Yeni ad: '{}'",
    "new_file_name":     "Yeni dosya adını yazın:",
    "unsupported":       "❓ Desteklenmeyen tür.",
    "send_fail":         "❌ Gönderilemedi: {}",
    "file_notfound":     "Dosya bulunamadı.",
    "media_sent":        "✅ {}/{} medya gönderildi.",
    # Yönetim
    "mgmt_panel":    "👥 Yönetim Paneli",
    "stats":         "📊 İstatistik",
    "users":         "👥 Kullanıcılar",
    "add_admin":     "👤 Admin Ekle",
    "del_admin":     "🚫 Admin Çıkar",
    "dm_user":       "💬 Kullanıcıya Mesaj",
    "broadcast":     "📢 Herkese Duyuru",
    "stats_title":   "📊 İSTATİSTİKLER",
    "top_views":     "👁 En Çok Görüntülenen:\n",
    "total_views":   "👁 Toplam Görüntüleme: {}",
    "view_count":    " ({}x)",
    "user_list":     "👥 Kullanıcılar — birini seçin:",
    "no_users":      "Henüz kullanıcı yok.",
    "no_admins":     "Eklenmiş admin yok.",
    "admin_enter_id": "Admin ID veya @kullanıcı adı yazın:",
    "admin_added":   "✅ Admin eklendi: {}",
    "admin_exists":  "❌ Zaten admin.",
    "invalid_id":    "❌ Geçersiz ID.",
    "del_admin_lbl": "Çıkarılacak admini seçin:",
    "admin_removed": "✅ Admin çıkarıldı: {}",
    "dm_select":     "💬 Mesaj göndermek istediğin kullanıcıyı seç:",
    "dm_manual":     "✏️ ID ile yaz",
    "dm_enter_id":   "Kullanıcı ID'sini yaz:",
    "dm_enter_msg":  "💬 {} kişisine gönderilecek mesajı yaz:",
    "dm_sent":       "✅ Mesaj gönderildi → {}",
    "dm_fail":       "❌ Gönderilemedi: {}",
    "dm_no_target":  "❌ Hedef kullanıcı bulunamadı.",
    "broadcast_enter": "📢 Tüm kullanıcılara mesaj yaz:\n💡 Metin, fotoğraf, video veya dosya.",
    "broadcast_done":  "📢 Duyuru tamamlandı!\n✅ Gönderildi: {}\n❌ Başarısız: {}",
    # Anket
    "poll_btn":           "📊 Anket",
    "poll_create":        "➕ Yeni Anket",
    "poll_type_select":   "📊 Anket türünü seç:",
    "poll_type_choice":   "🔘 Çoktan Seçmeli\n(Kullanıcılar şıklar arasından seçer)",
    "poll_type_open":     "✍️ Açık Uçlu\n(Kullanıcılar yazılı cevap verir)",
    "poll_results":       "📈 Sonuçlar",
    "poll_delete":        "🗑 Anketi Sil",
    "poll_panel":         "📊 Anket Yönetimi",
    "poll_enter_q":       "📝 Anket sorusunu yazın:",
    "poll_enter_opts":    "📋 Seçenekleri yazın (her satıra bir seçenek, max 6):\n\nÖrnek:\nEvet\nHayır\nFikirm yok",
    "poll_preview":       "📊 Anket Önizleme:\n\n❓ {q}\n\n{opts}\n\nGönderilsin mi?",
    "poll_send_btn":      "📤 Gönder",
    "poll_cancel_btn":    "❌ İptal",
    "poll_sent":          "📊 Anket gönderildi!\n✅ {s} kullanıcı\n❌ {f} başarısız",
    "poll_no_polls":      "Henüz anket yok.",
    "poll_voted":         "✅ Oyunuz kaydedildi!",
    "poll_already_voted": "❗ Bu ankete zaten oy verdiniz.",
    "poll_closed":        "📊 Bu anket sona erdi.",
    "poll_results_title": "📈 Anket Sonuçları:\n\n❓ {q}\n👥 Toplam oy: {total}\n\n{bars}",
    "poll_no_votes":      "Henüz oy yok.",
    "poll_deleted":       "✅ Anket silindi.",
    "poll_select":        "Seçmek istediğin anketi seç:",
    "poll_comment_prompt":"💬 İstersen bir yorum ekleyebilirsin (isteğe bağlı):\n\nYorum yazmak istemiyorsan /skip yaz.",
    "poll_comment_saved": "💬 Yorumun kaydedildi, teşekkürler!",
    "poll_comment_skip":  "✅ Tamam, yorum olmadan devam edildi.",
    "poll_comments_title":"💬 Kullanıcı Yorumları:",
    "poll_no_comments":   "Henüz yorum yok.",
    # (boş)
    "admin_notif":       "🔔 Kullanıcı talebi\n\n👤 {name}{un}\n🆔 {uid}\n📋 {summary}\n🕐 {time}",
    "new_user_notif":  "🆕 Yeni kullanıcı!\n\n👤 {name}{un}\n🆔 {uid}\n🕐 {time}",
    "reply_btn":     "💬 Yanıtla (ID: {})",
    "dm_prefix":     "📩 Yönetici mesajı:\n\n{}",
    "block_btn":     "🚫 Engelle",
    "unblock_btn":   "✅ Engeli Kaldır",
    "msg_btn":       "💬 Mesaj Gönder",
    "blocked_ok":    "✅ {} engellendi.",
    "unblocked_ok":  "✅ Engel kaldırıldı: {}",
    "send_media_btn": "🖼 Medyaları Gönder ({})",
    "user_last":     "🕐 Son görülme:",
    "user_msgs":     "💬 Mesaj: {} | 📎 Medya: {}",
    # Ayarlar
    "settings_title":     "⚙️ BOT AYARLARI",
    "set_name_btn":       "📝 Bot Adını Değiştir",
    "set_welcome_btn":    "💬 Karşılama Mesajı",
    "set_photo_btn":      "🖼 Bot Fotoğrafı",
    "set_blocked_btn":    "🚫 Engellenenler",
    "btn_mgmt_btn":       "🎛 Buton Yönetimi",
    "set_name_prompt":    "Yeni bot adını yazın (mevcut: {}):",
    "set_welcome_prompt": "Yeni karşılama mesajını yazın:",
    "set_photo_prompt":   "Yeni bot fotoğrafını gönderin:",
    "set_photo_ok":       "✅ Bot fotoğrafı güncellendi.",
    "set_photo_fail":     "❌ Lütfen bir fotoğraf gönderin.",
    "set_name_ok":        "✅ Bot adı: '{}'",
    "set_welcome_ok":     "✅ Karşılama mesajı güncellendi.",
    "blocked_list":       "🚫 Engellenenler ({})",
    "no_blocked":         "Kimse engellenmedi.",
    "maint_toggle":       "Bakım Modu: {}",
    "maint_on_str":       "AÇIK 🔧",
    "maint_off_str":      "KAPALI ✅",
    "btn_mgmt_title":     "🎛 Buton Görünürlük Ayarları",
    # Arama
    "search_prompt":  "🔍 Aramak istediğin kelimeyi yaz:",
    "search_results": "🔍 '{}' için sonuçlar:",
    "search_none":    "❌ '{}' için sonuç bulunamadı.",
    # Mesajlaşma
    "msg_to_admin_prompt": "💬 Mesajınızı yazın, doğrudan yöneticiye iletilecek:",
    "msg_to_admin_sent":   "✅ Mesajınız yöneticiye iletildi.",
    "admin_user_msg":      "💬 YENİ MESAJ\n\n👤 {name}{un}\n🆔 {uid}\n\n📝 {text}\n🕐 {time}",
    # Yardım
    "help_text": (
        "❓ *YARDIM — Süper Admin*\n\n"
        "📁 *İçerik* — Klasör/dosya ekle, sil, düzenle\n"
        "👥 *Yönetim* — Kullanıcılar, adminler, duyuru\n"
        "⚙️ *Ayarlar* — Bot adı, karşılama, fotoğraf\n"
        "🔧 *Bakım* — Bakım modunu aç/kapat\n"
        "🔍 *Ara* — Klasör ve dosyalarda ara\n\n"
        "Eklediğin adminler yalnızca klasör ve dosya ekleyebilir."
    ),
    # ═══ SINIF SİSTEMİ ═══
    "class_select_prompt": (
        "👋 Hoş geldin!\n\n"
        "Lütfen sınıfını seçin. Bu seçim kalıcıdır ve size özel içerik gösterilmesi için kullanılır."
    ),
    "class_selected":   "✅ Sınıfın kaydedildi: {}",
    "class_change_btn": "🎓 Sınıfımı Değiştir",
    "class_change_prompt": "🎓 Yeni sınıfını seçin:",
    "class_changed":    "✅ Sınıf güncellendi: {}",
    "class_label":      "🎓 Sınıf: {}",
    "class_unknown":    "❓ Belirsiz",
    "class_filter_btn": "🎓 Sınıfa Göre Filtrele",
    # ═══ YENİ KULLANICI BİLDİRİMİ ═══
    "new_user_notif":   "🆕 Yeni kullanıcı!\n\n👤 {name}{un}\n🆔 {uid}\n🕐 {time}",
    # ═══ İÇ YAPAY ZEKA ═══
    "ai_btn":           "🤖 Asistan",
    "ai_welcome":       "🤖 Merhaba! Sana nasıl yardımcı olabilirim?",
    "ai_no_answer":     "❓ Bu konuyu anlayamadım. Daha açık bir soru yazar mısın?",
    "ai_faq_btn":       "❓ Sık Sorulan Sorular",
    "faq_add_btn":      "➕ Soru-Cevap Ekle",
    "faq_list_btn":     "📋 Mevcut Sorular",
    "faq_del_btn":      "🗑 Soru Sil",
    "faq_panel":        "❓ FAQ Yönetimi",
    "faq_enter_q":      "❓ Soruyu yazın (veya anahtar kelimeleri):",
    "faq_enter_a":      "💬 Cevabı yazın:",
    "faq_saved":        "✅ Soru-Cevap kaydedildi.",
    "faq_deleted":      "✅ Soru-Cevap silindi.",
    "faq_empty":        "Henüz soru-cevap yok.",
    "faq_list_title":   "📋 Mevcut Sorular ({} adet):",
    "auto_rule_btn":    "⚡ Otomatik Kurallar",
    "rule_add_btn":     "➕ Kural Ekle",
    "rule_list_btn":    "📋 Kurallar",
    "rule_del_btn":     "🗑 Kural Sil",
    "rule_panel":       "⚡ Otomatik Kural Yönetimi",
    "rule_enter_kw":    "🔍 Tetikleyici kelime(ler) yazın (virgülle ayırın):",
    "rule_enter_resp":  "💬 Bu anahtar kelimeye verilecek cevabı yazın:",
    "rule_saved":       "✅ Kural kaydedildi.",
    "rule_deleted":     "✅ Kural silindi.",
    "rule_empty":       "Henüz kural yok.",
    "suggest_btn":      "💡 Benim İçin Öneri",
    "suggest_title":    "💡 Sana Özel Öneriler:",
    "suggest_empty":    "Öneri üretmek için önce birkaç dosya görüntülemelisin.",
    # ═══ FAVORİLER ═══
    "btn_favorites":    "⭐ Favorilerim",
    "fav_added":        "⭐ Favorilere eklendi!",
    "fav_removed":      "💔 Favorilerden kaldırıldı.",
    "fav_list":         "⭐ FAVORİLERİM",
    "fav_empty":        "Henüz favori eklemedin.\n\nDosya görüntülerken ⭐ butonuna bas.",
    "fav_add_btn":      "⭐ Favoriye Ekle",
    "fav_remove_btn":   "💔 Favoriden Çıkar",
    # ═══ SON GÖRÜNTÜLENENLEr ═══
    "btn_recent":       "🕐 Son Görüntülenenler",
    "recent_list":      "🕐 SON GÖRÜNTÜLENENLEr",
    "recent_empty":     "Henüz dosya görüntülemedin.",
    # ═══ UYARI SİSTEMİ ═══
    "warn_btn":             "⚠️ Uyarı Ver",
    "warn_reason_prompt":   "⚠️ Uyarı sebebini yazın:",
    "user_warned":          "⚠️ {} kullanıcısına uyarı verildi. ({}/{})",
    "user_auto_blocked":    "🚫 {} otomatik engellendi! ({} uyarı)",
    "warn_msg_to_user":     "⚠️ BİR UYARI ALDINIZ\n\nSebep: {reason}\n\n🚨 Uyarı: {count}/{max}\n\n{count} üst sınıra ulaşırsa erişiminiz kısıtlanır.",
    "warn_count_label":     "⚠️ Uyarılar: {}/{}",
    "warn_clear_btn":       "🔓 Uyarıları Sıfırla",
    "warns_cleared":        "✅ {} uyarıları sıfırlandı.",
    "no_warns":             "✅ Uyarısı yok.",
    # ═══ KULLANICI NOTU ═══
    "note_btn":         "📝 Not Ekle/Güncelle",
    "note_prompt":      "📝 Bu kullanıcı için not yazın (boş = sil):",
    "note_saved":       "✅ Not kaydedildi.",
    "note_label":       "📝 Admin Notu: ",
    "note_cleared":     "✅ Not silindi.",
    # ═══ YEDEKLEME ═══
    "backup_btn":       "💾 Tam Yedek Al",
    "backup_sending":   "💾 Yedek hazırlanıyor...",
    "backup_done":      "✅ Yedek tamamlandı. {} dosya gönderildi.",
    "export_users_btn": "📤 Kullanıcıları Dışa Aktar",
    "export_done":      "✅ {} kullanıcı dışa aktarıldı.",
    # ═══ HEDEFLİ DUYURU ═══
    "bcast_targeted":   "🎯 Hedefli Duyuru",
    "bcast_all":        "📢 Tüm Kullanıcılar",
    "bcast_class":      "🎓 Sınıfa Göre",
    "bcast_active":     "✅ Aktifler (7 gün)",
    "bcast_new":        "🆕 Yeni Kullanıcılar (7 gün)",
    "bcast_class_select":"🎓 Hangi sınıfa gönderilsin?",
    "bcast_confirm":    "📢 {} kullanıcıya gönderilecek. Devam edilsin mi?",
    "bcast_target_set": "🎯 Hedef: {}",
    # ═══ KLASÖR ABONELİĞİ ═══
    "subscribe_btn":    "🔔 Abone Ol",
    "unsubscribe_btn":  "🔕 Aboneliği Kaldır",
    "subscribed_ok":    "🔔 Abone oldunuz! Yeni dosya eklenince bildirim alacaksınız.",
    "unsubscribed_ok":  "🔕 Abonelik kaldırıldı.",
    "sub_notif":        "🔔 Yeni dosya eklendi!\n📁 {folder}\n📎 {fname}",
    # ═══ KLASÖR AÇIKLAMASI ═══
    "folder_desc_btn":    "📝 Klasör Açıklaması",
    "folder_desc_prompt": "📝 Bu klasör için açıklama yazın (boş bırakırsan silinir):",
    "folder_desc_saved":  "✅ Açıklama kaydedildi.",
    "folder_desc_cleared":"✅ Açıklama silindi.",
    # ═══ DOSYA İŞLEMLERİ ═══
    "pin_file":           "📌 Dosya Sabitle",
    "unpin_file":         "📌 Sabitlemeyi Kaldır",
    "pin_select":         "📌 Sabitlenecek dosyayı seçin:",
    "unpin_select":       "📌 Sabitleme kaldırılacak dosyayı seçin:",
    "file_pinned":        "📌 Sabitlendi: '{}'",
    "file_unpinned":      "📌 Sabitleme kaldırıldı: '{}'",
    "move_file":          "📂 Dosya Taşı",
    "copy_file":          "📋 Dosya Kopyala",
    "move_select_file":   "📂 Taşınacak dosyayı seçin:",
    "copy_select_file":   "📋 Kopyalanacak dosyayı seçin:",
    "move_select_dest":   "📂 Hedef klasörü seçin:\n(Mevcut konum: {})",
    "file_moved":         "✅ '{}' → '{}' klasörüne taşındı.",
    "file_copied":        "✅ '{}' → '{}' klasörüne kopyalandı.",
    "no_dest":            "❌ Hedef klasör bulunamadı.",
    "sort_az":            "🔤 A-Z Sırala",
    "sort_views":         "👁 Görüntülemeye Göre Sırala",
    "sort_date":          "📅 Tarihe Göre Sırala",
    "files_sorted":       "✅ Dosyalar sıralandı.",
    # ═══ ETİKET SİSTEMİ ═══
    "tag_btn":            "🏷 Etiket Ekle",
    "tag_prompt":         "🏷 Etiket(ler) yazın (virgülle ayırın):",
    "tag_saved":          "✅ Etiketler kaydedildi.",
    "tag_search":         "🏷 Etikete Göre Ara",
    "tag_select_prompt":  "🏷 Aranacak etiketi yazın:",
    # ═══ ZAMANLAMA ═══
    "schedule_btn":       "⏰ Zamanlanmış Mesaj",
    "schedule_prompt":    "⏰ Kaç saat sonra gönderilsin? (0.5 - 72)",
    "schedule_msg_prompt":"💬 Gönderilecek mesajı yazın:",
    "schedule_saved":     "✅ Mesaj {} saat sonra gönderilecek.",
    "schedule_list":      "⏰ Zamanlanmış Mesajlar",
    "schedule_empty":     "Zamanlanmış mesaj yok.",
    "schedule_del":       "🗑 İptal Et",
    "schedule_canceled":  "✅ Zamanlanmış mesaj iptal edildi.",
    # ═══ HATIRLATICI ═══
    "reminder_btn":       "🔔 Hatırlatıcı Ekle",
    "reminder_text":      "📝 Hatırlatıcı metnini yazın:",
    "reminder_hour":      "⏰ Kaç saat sonra hatırlatılsın?",
    "reminder_saved":     "✅ Hatırlatıcı kaydedildi.",
    "reminder_msg":       "🔔 HATIRLATICI\n\n{}",
    # ═══ ADMIN LOG ═══
    "admin_log_btn":      "📋 İşlem Günlüğü",
    "admin_log_title":    "📋 SON ADMIN İŞLEMLERİ",
    "admin_log_empty":    "Henüz kayıt yok.",
    # ═══ GİZLİ MOD ═══
    "secret_mode_btn":    "🔒 Gizli Mod",
    "secret_on":          "🔒 Gizli Mod: AÇIK",
    "secret_off":         "🔓 Gizli Mod: KAPALI",
    "secret_add_btn":     "➕ Kullanıcı Ekle",
    "secret_del_btn":     "➖ Kullanıcı Çıkar",
    "secret_enter_id":    "Whitelist'e eklenecek kullanıcı ID'sini yazın:",
    "secret_added":       "✅ {} whitelist'e eklendi.",
    "secret_removed":     "✅ {} whitelist'ten çıkarıldı.",
    "secret_list":        "🔒 Whitelist ({} kullanıcı):",
    # ═══ KULLANICI PROFİL ═══
    "profile_btn":        "👤 Profilim",
    "profile_title":      "👤 PROFİLİM",
    # ═══ SPAM KORUMA ═══
    "spam_warning":       "⚠️ Çok hızlı mesaj gönderiyorsun. Lütfen bekle.",
    # ═══ STARTUP ═══
    "startup_msg":        "✅ Bot yeniden başladı! Tüm servisler aktif.",
    # ═══ WEB ARAMA YAPAY ZEKASI ═══
    "ai_searching":       "🔍 Araştırıyorum...",
    "ai_calculating":     "🧮 Hesaplıyorum...",
    "ai_search_result":   "🔍 *Arama Sonucu:*\n\n{result}\n\n📎 Kaynak: {source}",
    "ai_math_result":     "🧮 *Matematik Çözümü:*\n\n{result}",
    "ai_wiki_result":     "📖 *Vikipedi:*\n\n{result}\n\n🔗 {url}",
    "ai_no_web_result":   "❌ Bu soru için sonuç bulunamadı. Farklı bir ifadeyle tekrar dene.",
    "ai_web_error":       "⚠️ Arama sırasında hata oluştu. Lütfen tekrar dene.",
    "ai_typing":          "🤖 Yazıyor...",
    "ai_source_ddg":      "DuckDuckGo",
    "ai_source_wiki":     "Vikipedi",
    "ai_source_calc":     "Hesap Makinesi",
    "ai_subject_math":    "🧮 Matematik",
    "ai_subject_physics": "⚛️ Fizik",
    "ai_subject_chem":    "🧪 Kimya",
    "ai_subject_prog":    "💻 Programlama",
    "ai_subject_eng":     "⚙️ Mühendislik",
    "ai_detected":        "🎯 Konu: {} — Web'de arıyorum...",
    # ═══ LIDERBOARD / BAŞARI ═══
    "btn_leaderboard":    "🏆 Liderboard",
    "leaderboard_title":  "🏆 EN AKTİF KULLANICILAR",
    "leaderboard_empty":  "Henüz veri yok.",
    "achievement_unlocked":"🏅 Başarı Kazandınız: {}!",
    # ═══ GERİ BİLDİRİM ═══
    "feedback_btn":       "⭐ Geri Bildirim",
    "feedback_prompt":    "⭐ Bu dosya için puan verin (1-5 yıldız):",
    "feedback_saved":     "✅ Geri bildiriminiz kaydedildi!",
    "feedback_stats":     "⭐ Ortalama: {avg:.1f}/5 ({count} oy)",
    # ═══ SABİT MESAJ ═══
    "pin_msg_btn":        "📌 Mesaj Sabitle",
    "pin_msg_prompt":     "📌 Sabitlenecek mesajı yazın:",
    "pin_msg_saved":      "✅ Mesaj sabitlendi.",
    "pinned_msg_label":   "📌 SABİT MESAJ:\n{}",
    # ═══ BROADCAST LOG ═══
    "bcast_history_btn":  "📜 Duyuru Geçmişi",
    "bcast_history_title":"📜 SON DUYURULAR",
    "bcast_history_empty":"Henüz duyuru yok.",
    # ═══ TOPLU MESAJ SİLME ═══
    "clear_chat_btn":     "🧹 Sohbeti Temizle",
    "clear_chat_done":    "✅ Sohbet temizlendi.",
    # ═══ HIZLI CEVAP ═══
    "quick_reply_btn":    "⚡ Hızlı Cevap",
    "quick_reply_saved":  "✅ Hızlı cevap şablonu kaydedildi.",
    # ═══ YENİ BUTONLAR ═══
    "btn_notes":          "📝 Notlarım",
    "btn_reminder":       "⏰ Hatırlatıcım",
    "anon_q_btn":         "❓ Anonim Soru",
    "countdown_btn":      "⏳ Yaklaşan Sınavlar",
    "quiz_btn":           "📝 Mini Test",
    "notes_empty":        "Henüz not yok.",
    "notes_saved":        "✅ Not kaydedildi.",
    "notes_title":        "📝 KİŞİSEL NOTLARIM",
    "notes_prompt":       "📝 Notunuzu yazın:",
    "reminder_list":      "⏰ Hatırlatıcılarım ({} adet):",
    "reminder_none":      "Henüz hatırlatıcı yok.",
    "reminder_add_prompt":"⏰ Hatırlatıcı metni yaz:",
    "reminder_saved":     "✅ {} sonra hatırlatacağım.",
    "reminder_fired":     "🔔 HATIRLATICI\n\n{}",
    "reminder_del":       "✅ Silindi.",
    "anon_q_prompt":      "❓ Sorunuzu yazın (anonim):",
    "anon_q_sent":        "✅ Sorunuz iletildi.",
    "countdown_none":     "Yaklaşan sınav eklenmemiş.",
    "countdown_prompt":   "Sınav adını yazın:",
    "countdown_date":     "Tarih yazın (örn: 20/05/2026):",
    "countdown_saved":    "✅ {} kaydedildi.",
    "quiz_none":          "Aktif test yok.",
    "group_select":       "👥 Grubunu seç (A/B/C → alt grup):",
    "group_selected":     "✅ Grubun kaydedildi: {}",
    "group_change_btn":   "👥 Grubumu Değiştir",
    "group_label":        "👥 Grup: {}",
}

AR = {
    # Klavye butonları
    "btn_content":   "📁 المحتوى",
    "btn_mgmt":      "👥 الإدارة",
    "btn_settings":  "⚙️ الإعدادات",
    "btn_maint":     "🔧 الصيانة",
    "btn_search":    "🔍 بحث",
    "btn_help":      "💬 راسل المسؤول",
    # Genel
    "home":          "🏠 الصفحة الرئيسية",
    "folder_list":   "📁 المجلدات:",
    "file_list":     "📎 الملفات:",
    "back":          "◀️ رجوع",
    "cancel":        "❌ إلغاء",
    "close":         "✖️ إغلاق",
    "maint_on":      "🔧 وضع الصيانة: مفعّل",
    "maint_off":     "✅ وضع الصيانة: معطّل",
    # İçerik
    "content_mgmt":         "📁 إدارة المحتوى",
    "add_folder":           "➕ إضافة مجلد",
    "add_file":             "📎 إضافة ملف/وسائط",
    "enter_folder_name":    "📁 اكتب اسم المجلد الجديد:",
    "no_folders":           "لا توجد مجلدات.",
    "no_files":             "لا توجد ملفات.",
    "add_file_prompt":       "📎 أرسل ملفاً أو صورة أو فيديو، أو اكتب رابطاً/نصاً:\n\n📋 الملفات الموجودة:\n{}",
    "add_file_prompt_empty": "📎 أرسل ملفاً أو صورة أو فيديو، أو اكتب رابطاً/نصاً:\n\n📭 لا توجد ملفات في هذا المجلد بعد.",
    "del_folder":           "🗑 حذف مجلد",
    "rename_folder":        "✏️ إعادة تسمية مجلد",
    "del_file":             "🗑 حذف ملف",
    "rename_file":          "✏️ إعادة تسمية ملف",
    "del_folder_select":    "اختر المجلد للحذف:",
    "rename_folder_select": "اختر المجلد لإعادة التسمية:",
    "del_file_select":      "اختر الملف للحذف:",
    "rename_file_select":   "اختر الملف لإعادة التسمية:",
    # Onay
    "del_folder_confirm": "🗑 هل تريد حذف مجلد *'{}'*؟\n⚠️ سيتم حذف جميع الملفات والمجلدات الفرعية!",
    "del_file_confirm":   "🗑 هل تريد حذف *'{}'*؟",
    "confirm_yes":        "✅ نعم، احذف",
    "confirm_no":         "❌ إلغاء",
    # Klasör/dosya işlemleri
    "folder_exists":  "❌ '{}' موجود بالفعل!",
    "folder_empty":   "❌ لا يمكن أن يكون الاسم فارغاً!",
    "folder_added":   "✅ تمت إضافة المجلد '{}'.",
    "folder_deleted": "✅ تم حذف '{}'.",
    "folder_renamed": "✅ '{}' → '{}'",
    "new_folder_name": "اكتب الاسم الجديد لـ '{}':",
    "file_added":        "✅ تمت الإضافة: {}",
    "file_exists_warn":  "⚠️ يوجد ملف باسم '{}' بالفعل! تمت الإضافة على أي حال.",
    "file_deleted":      "✅ تم حذف '{}'.",
    "file_renamed":      "✅ الاسم الجديد: '{}'",
    "new_file_name":     "اكتب الاسم الجديد للملف:",
    "unsupported":       "❓ نوع غير مدعوم.",
    "send_fail":         "❌ فشل الإرسال: {}",
    "file_notfound":     "الملف غير موجود.",
    "media_sent":        "✅ تم إرسال {}/{} وسائط.",
    # Yönetim
    "mgmt_panel":    "👥 لوحة الإدارة",
    "stats":         "📊 الإحصائيات",
    "users":         "👥 المستخدمون",
    "add_admin":     "👤 إضافة مشرف",
    "del_admin":     "🚫 إزالة مشرف",
    "dm_user":       "💬 رسالة لمستخدم",
    "broadcast":     "📢 رسالة للجميع",
    "stats_title":   "📊 الإحصائيات",
    "top_views":     "👁 الأكثر مشاهدة:\n",
    "total_views":   "👁 إجمالي المشاهدات: {}",
    "view_count":    " ({}x)",
    "user_list":     "👥 المستخدمون — اختر واحداً:",
    "no_users":      "لا يوجد مستخدمون بعد.",
    "no_admins":     "لا يوجد مشرفون.",
    "admin_enter_id": "اكتب ID أو @اسم_المستخدم:",
    "admin_added":   "✅ تمت إضافة المشرف: {}",
    "admin_exists":  "❌ هو مشرف بالفعل.",
    "invalid_id":    "❌ ID غير صالح.",
    "del_admin_lbl": "اختر المشرف للإزالة:",
    "admin_removed": "✅ تمت إزالة المشرف: {}",
    "dm_select":     "💬 اختر المستخدم:",
    "dm_manual":     "✏️ أدخل ID يدوياً",
    "dm_enter_id":   "اكتب ID المستخدم:",
    "dm_enter_msg":  "💬 اكتب الرسالة لـ {}:",
    "dm_sent":       "✅ تم الإرسال → {}",
    "dm_fail":       "❌ فشل الإرسال: {}",
    "dm_no_target":  "❌ المستخدم غير موجود.",
    "broadcast_enter": "📢 اكتب الرسالة للجميع:\n💡 نص أو صورة أو فيديو أو ملف.",
    "broadcast_done":  "📢 اكتمل الإرسال!\n✅ نجح: {}\n❌ فشل: {}",
    # Anket
    "poll_btn":           "📊 استطلاع",
    "poll_create":        "➕ استطلاع جديد",
    "poll_type_select":   "📊 اختر نوع الاستطلاع:",
    "poll_type_choice":   "🔘 اختيار من متعدد\n(المستخدمون يختارون من الخيارات)",
    "poll_type_open":     "✍️ سؤال مفتوح\n(المستخدمون يكتبون إجابتهم)",
    "poll_results":       "📈 النتائج",
    "poll_delete":        "🗑 حذف الاستطلاع",
    "poll_panel":         "📊 إدارة الاستطلاعات",
    "poll_enter_q":       "📝 اكتب سؤال الاستطلاع:",
    "poll_enter_opts":    "📋 اكتب الخيارات (خيار في كل سطر، max 6):\n\nمثال:\nنعم\nلا\nلا أعرف",
    "poll_preview":       "📊 معاينة الاستطلاع:\n\n❓ {q}\n\n{opts}\n\nهل تريد الإرسال؟",
    "poll_send_btn":      "📤 إرسال",
    "poll_cancel_btn":    "❌ إلغاء",
    "poll_sent":          "📊 تم إرسال الاستطلاع!\n✅ {s} مستخدم\n❌ {f} فشل",
    "poll_no_polls":      "لا توجد استطلاعات بعد.",
    "poll_voted":         "✅ تم تسجيل صوتك!",
    "poll_already_voted": "❗ لقد صوّتت بالفعل في هذا الاستطلاع.",
    "poll_closed":        "📊 انتهى هذا الاستطلاع.",
    "poll_results_title": "📈 نتائج الاستطلاع:\n\n❓ {q}\n👥 مجموع الأصوات: {total}\n\n{bars}",
    "poll_no_votes":      "لا توجد أصوات بعد.",
    "poll_deleted":       "✅ تم حذف الاستطلاع.",
    "poll_select":        "اختر الاستطلاع:",
    "poll_comment_prompt":"💬 يمكنك إضافة تعليق (اختياري):\n\nإذا لم تريد التعليق اكتب /skip",
    "poll_comment_saved": "💬 تم حفظ تعليقك، شكراً!",
    "poll_comment_skip":  "✅ تم المتابعة بدون تعليق.",
    "poll_comments_title":"💬 تعليقات المستخدمين:",
    "poll_no_comments":   "لا توجد تعليقات بعد.",
    # (boş)
    "admin_notif":       "🔔 طلب مستخدم\n\n👤 {name}{un}\n🆔 {uid}\n📋 {summary}\n🕐 {time}",
    "new_user_notif":  "🆕 مستخدم جديد!\n\n👤 {name}{un}\n🆔 {uid}\n🕐 {time}",
    "reply_btn":     "💬 رد (ID: {})",
    "dm_prefix":     "📩 رسالة من المسؤول:\n\n{}",
    "block_btn":     "🚫 حظر",
    "unblock_btn":   "✅ إلغاء الحظر",
    "msg_btn":       "💬 إرسال رسالة",
    "blocked_ok":    "✅ تم حظر {}.",
    "unblocked_ok":  "✅ تم إلغاء حظر: {}",
    "send_media_btn": "🖼 إرسال الوسائط ({})",
    "user_last":     "🕐 آخر ظهور:",
    "user_msgs":     "💬 الرسائل: {} | 📎 الوسائط: {}",
    # Ayarlar
    "settings_title":     "⚙️ إعدادات البوت",
    "set_name_btn":       "📝 اسم البوت",
    "set_welcome_btn":    "💬 رسالة الترحيب",
    "set_photo_btn":      "🖼 صورة البوت",
    "set_blocked_btn":    "🚫 المحظورون",
    "btn_mgmt_btn":       "🎛 إدارة الأزرار",
    "set_name_prompt":    "اكتب الاسم الجديد للبوت (الحالي: {}):",
    "set_welcome_prompt": "اكتب رسالة الترحيب الجديدة:",
    "set_photo_prompt":   "أرسل الصورة الجديدة للبوت:",
    "set_photo_ok":       "✅ تم تحديث صورة البوت.",
    "set_photo_fail":     "❌ يرجى إرسال صورة.",
    "set_name_ok":        "✅ اسم البوت: '{}'",
    "set_welcome_ok":     "✅ تم تحديث رسالة الترحيب.",
    "blocked_list":       "🚫 المحظورون ({})",
    "no_blocked":         "لا أحد محظور.",
    "maint_toggle":       "وضع الصيانة: {}",
    "maint_on_str":       "مفعّل 🔧",
    "maint_off_str":      "معطّل ✅",
    "btn_mgmt_title":     "🎛 إعدادات ظهور الأزرار",
    # Arama
    "search_prompt":  "🔍 اكتب كلمة البحث:",
    "search_results": "🔍 نتائج '{}' :",
    "search_none":    "❌ لا توجد نتائج لـ '{}'.",
    # Mesajlaşma
    "msg_to_admin_prompt": "💬 اكتب رسالتك، وستصل مباشرة إلى المسؤول:",
    "msg_to_admin_sent":   "✅ تم إرسال رسالتك إلى المسؤول.",
    "admin_user_msg":      "💬 رسالة جديدة\n\n👤 {name}{un}\n🆔 {uid}\n\n📝 {text}\n🕐 {time}",
    # Yardım
    "help_text": (
        "❓ *المساعدة*\n\n"
        "📁 *المحتوى* — تصفح المجلدات والملفات\n"
        "🔍 *بحث* — ابحث عن ملف أو مجلد\n"
        "💬 *راسل المسؤول* — أرسل طلباً للمسؤول"
    ),
    # ═══ SINIF SİSTEMİ (AR) ═══
    "class_select_prompt": (
        "👋 أهلاً بك!\n\n"
        "يرجى اختيار سنتك الدراسية. سيُستخدم هذا لعرض المحتوى المناسب لك."
    ),
    "class_selected":    "✅ تم تسجيل سنتك الدراسية: {}",
    "class_change_btn":  "🎓 تغيير السنة الدراسية",
    "class_change_prompt":"🎓 اختر سنتك الدراسية الجديدة:",
    "class_changed":     "✅ تم تحديث السنة الدراسية: {}",
    "class_label":       "🎓 السنة: {}",
    "class_unknown":     "❓ غير محدد",
    "class_filter_btn":  "🎓 تصفية حسب السنة",
    # ═══ يدعم بالعربية ═══
    "new_user_notif":    "🆕 مستخدم جديد!\n\n👤 {name}{un}\n🆔 {uid}\n🕐 {time}",
    "ai_btn":            "🤖 المساعد",
    "ai_welcome":        "🤖 مرحباً! كيف أستطيع مساعدتك؟",
    "ai_no_answer":      "❓ لم أفهم سؤالك. هل يمكنك إعادة الصياغة؟",
    "ai_faq_btn":        "❓ الأسئلة الشائعة",
    "faq_add_btn":       "➕ إضافة سؤال وجواب",
    "faq_list_btn":      "📋 الأسئلة الموجودة",
    "faq_del_btn":       "🗑 حذف سؤال",
    "faq_panel":         "❓ إدارة الأسئلة الشائعة",
    "faq_enter_q":       "❓ اكتب السؤال أو الكلمات المفتاحية:",
    "faq_enter_a":       "💬 اكتب الإجابة:",
    "faq_saved":         "✅ تم حفظ السؤال والإجابة.",
    "faq_deleted":       "✅ تم حذف السؤال.",
    "faq_empty":         "لا توجد أسئلة وأجوبة بعد.",
    "faq_list_title":    "📋 الأسئلة الموجودة ({} سؤال):",
    "auto_rule_btn":     "⚡ القواعد التلقائية",
    "rule_add_btn":      "➕ إضافة قاعدة",
    "rule_list_btn":     "📋 القواعد",
    "rule_del_btn":      "🗑 حذف قاعدة",
    "rule_panel":        "⚡ إدارة القواعد التلقائية",
    "rule_enter_kw":     "🔍 اكتب الكلمات المفتاحية (مفصولة بفاصلة):",
    "rule_enter_resp":   "💬 اكتب الرد على هذه الكلمات المفتاحية:",
    "rule_saved":        "✅ تم حفظ القاعدة.",
    "rule_deleted":      "✅ تم حذف القاعدة.",
    "rule_empty":        "لا توجد قواعد بعد.",
    "suggest_btn":       "💡 اقتراحات لي",
    "suggest_title":     "💡 اقتراحات مخصصة لك:",
    "suggest_empty":     "شاهد بعض الملفات أولاً لتحصل على اقتراحات.",
    "btn_favorites":     "⭐ المفضلة",
    "fav_added":         "⭐ تمت الإضافة للمفضلة!",
    "fav_removed":       "💔 تمت الإزالة من المفضلة.",
    "fav_list":          "⭐ قائمة المفضلة",
    "fav_empty":         "لا توجد ملفات مفضلة بعد.\nاضغط ⭐ عند عرض الملفات.",
    "fav_add_btn":       "⭐ إضافة للمفضلة",
    "fav_remove_btn":    "💔 إزالة من المفضلة",
    "btn_recent":        "🕐 المشاهدة مؤخراً",
    "recent_list":       "🕐 المشاهدة مؤخراً",
    "recent_empty":      "لم تشاهد أي ملف بعد.",
    "warn_btn":              "⚠️ تحذير",
    "warn_reason_prompt":    "⚠️ اكتب سبب التحذير:",
    "user_warned":           "⚠️ تم تحذير {}. ({}/{})",
    "user_auto_blocked":     "🚫 تم حظر {} تلقائياً! ({} تحذيرات)",
    "warn_msg_to_user":      "⚠️ تلقّيت تحذيراً\n\nالسبب: {reason}\n\n🚨 التحذيرات: {count}/{max}\n\nعند الوصول للحد الأقصى سيتم تقييد وصولك.",
    "warn_count_label":      "⚠️ التحذيرات: {}/{}",
    "warn_clear_btn":        "🔓 مسح التحذيرات",
    "warns_cleared":         "✅ تم مسح تحذيرات {}.",
    "no_warns":              "✅ لا توجد تحذيرات.",
    "note_btn":          "📝 إضافة/تعديل ملاحظة",
    "note_prompt":       "📝 اكتب ملاحظتك عن هذا المستخدم (فارغ = حذف):",
    "note_saved":        "✅ تم حفظ الملاحظة.",
    "note_label":        "📝 ملاحظة المسؤول: ",
    "note_cleared":      "✅ تم حذف الملاحظة.",
    "backup_btn":        "💾 نسخ احتياطي كامل",
    "backup_sending":    "💾 جارٍ تجهيز النسخ الاحتياطية...",
    "backup_done":       "✅ اكتمل النسخ الاحتياطي. {} ملف.",
    "export_users_btn":  "📤 تصدير المستخدمين",
    "export_done":       "✅ تم تصدير {} مستخدم.",
    "bcast_targeted":    "🎯 إرسال موجّه",
    "bcast_all":         "📢 الجميع",
    "bcast_class":       "🎓 حسب السنة",
    "bcast_active":      "✅ النشطون (7 أيام)",
    "bcast_new":         "🆕 الجدد (7 أيام)",
    "bcast_class_select":"🎓 أي سنة تريد الإرسال إليها؟",
    "bcast_confirm":     "📢 سيتم الإرسال لـ {} مستخدم. هل تريد المتابعة؟",
    "bcast_target_set":  "🎯 الهدف: {}",
    "subscribe_btn":     "🔔 اشترك في المجلد",
    "unsubscribe_btn":   "🔕 إلغاء الاشتراك",
    "subscribed_ok":     "🔔 تم الاشتراك! ستصلك إشعارات عند إضافة ملفات جديدة.",
    "unsubscribed_ok":   "🔕 تم إلغاء الاشتراك.",
    "sub_notif":         "🔔 ملف جديد!\n📁 {folder}\n📎 {fname}",
    "folder_desc_btn":    "📝 وصف المجلد",
    "folder_desc_prompt": "📝 اكتب وصفاً لهذا المجلد (فارغ = حذف):",
    "folder_desc_saved":  "✅ تم حفظ الوصف.",
    "folder_desc_cleared":"✅ تم حذف الوصف.",
    "pin_file":           "📌 تثبيت ملف",
    "unpin_file":         "📌 إلغاء التثبيت",
    "pin_select":         "📌 اختر الملف للتثبيت:",
    "unpin_select":       "📌 اختر الملف لإلغاء التثبيت:",
    "file_pinned":        "📌 تم تثبيت '{}'",
    "file_unpinned":      "📌 تم إلغاء تثبيت '{}'",
    "move_file":          "📂 نقل ملف",
    "copy_file":          "📋 نسخ ملف",
    "move_select_file":   "📂 اختر الملف للنقل:",
    "copy_select_file":   "📋 اختر الملف للنسخ:",
    "move_select_dest":   "📂 اختر المجلد الهدف:\n(الموقع الحالي: {})",
    "file_moved":         "✅ تم نقل '{}' إلى '{}'.",
    "file_copied":        "✅ تم نسخ '{}' إلى '{}'.",
    "no_dest":            "❌ المجلد الهدف غير موجود.",
    "sort_az":            "🔤 ترتيب أبجدي",
    "sort_views":         "👁 حسب المشاهدات",
    "sort_date":          "📅 حسب التاريخ",
    "files_sorted":       "✅ تم ترتيب الملفات.",
    "tag_btn":            "🏷 إضافة وسوم",
    "tag_prompt":         "🏷 اكتب الوسوم (مفصولة بفاصلة):",
    "tag_saved":          "✅ تم حفظ الوسوم.",
    "tag_search":         "🏷 البحث بالوسم",
    "tag_select_prompt":  "🏷 اكتب الوسم للبحث:",
    "schedule_btn":       "⏰ رسالة مجدولة",
    "schedule_prompt":    "⏰ كم ساعة؟ (0.5 - 72)",
    "schedule_msg_prompt":"💬 اكتب الرسالة المجدولة:",
    "schedule_saved":     "✅ ستُرسل الرسالة بعد {} ساعة.",
    "schedule_list":      "⏰ الرسائل المجدولة",
    "schedule_empty":     "لا توجد رسائل مجدولة.",
    "schedule_del":       "🗑 إلغاء",
    "schedule_canceled":  "✅ تم إلغاء الرسالة المجدولة.",
    "reminder_btn":       "🔔 إضافة تذكير",
    "reminder_text":      "📝 اكتب نص التذكير:",
    "reminder_hour":      "⏰ بعد كم ساعة؟",
    "reminder_saved":     "✅ تم حفظ التذكير.",
    "reminder_msg":       "🔔 تذكير\n\n{}",
    "admin_log_btn":      "📋 سجل الإجراءات",
    "admin_log_title":    "📋 آخر إجراءات المسؤول",
    "admin_log_empty":    "لا توجد سجلات بعد.",
    "secret_mode_btn":    "🔒 الوضع السري",
    "secret_on":          "🔒 الوضع السري: مفعّل",
    "secret_off":         "🔓 الوضع السري: معطّل",
    "secret_add_btn":     "➕ إضافة مستخدم",
    "secret_del_btn":     "➖ إزالة مستخدم",
    "secret_enter_id":    "اكتب ID المستخدم للإضافة:",
    "secret_added":       "✅ تمت إضافة {} للقائمة البيضاء.",
    "secret_removed":     "✅ تمت إزالة {} من القائمة البيضاء.",
    "secret_list":        "🔒 القائمة البيضاء ({} مستخدم):",
    "profile_btn":        "👤 ملفي الشخصي",
    "profile_title":      "👤 ملفي الشخصي",
    "spam_warning":       "⚠️ أنت تُرسل بسرعة كبيرة. الرجاء الانتظار.",
    "startup_msg":        "✅ البوت يعمل مجدداً! جميع الخدمات متاحة.",
    # ═══ WEB ARAMA YAPAY ZEKASI (AR) ═══
    "ai_searching":       "🔍 جارٍ البحث...",
    "ai_calculating":     "🧮 جارٍ الحساب...",
    "ai_search_result":   "🔍 *نتيجة البحث:*\n\n{result}\n\n📎 المصدر: {source}",
    "ai_math_result":     "🧮 *الحل الرياضي:*\n\n{result}",
    "ai_wiki_result":     "📖 *ويكيبيديا:*\n\n{result}\n\n🔗 {url}",
    "ai_no_web_result":   "❌ لم أجد نتيجة لهذا السؤال. جرب صياغة مختلفة.",
    "ai_web_error":       "⚠️ حدث خطأ أثناء البحث. حاول مرة أخرى.",
    "ai_typing":          "🤖 يكتب...",
    "ai_source_ddg":      "DuckDuckGo",
    "ai_source_wiki":     "ويكيبيديا",
    "ai_source_calc":     "الآلة الحاسبة",
    "ai_subject_math":    "🧮 رياضيات",
    "ai_subject_physics": "⚛️ فيزياء",
    "ai_subject_chem":    "🧪 كيمياء",
    "ai_subject_prog":    "💻 برمجة",
    "ai_subject_eng":     "⚙️ هندسة",
    "ai_detected":        "🎯 الموضوع: {} — جارٍ البحث على الويب...",
    "btn_leaderboard":    "🏆 المتصدرون",
    "leaderboard_title":  "🏆 أكثر المستخدمين نشاطاً",
    "leaderboard_empty":  "لا توجد بيانات بعد.",
    "achievement_unlocked":"🏅 حصلت على إنجاز: {}!",
    "feedback_btn":       "⭐ تقييم",
    "feedback_prompt":    "⭐ قيّم هذا الملف (1-5 نجوم):",
    "feedback_saved":     "✅ تم حفظ تقييمك!",
    "feedback_stats":     "⭐ المتوسط: {avg:.1f}/5 ({count} تقييم)",
    "pin_msg_btn":        "📌 تثبيت رسالة",
    "pin_msg_prompt":     "📌 اكتب الرسالة المراد تثبيتها:",
    "pin_msg_saved":      "✅ تم تثبيت الرسالة.",
    "pinned_msg_label":   "📌 الرسالة المثبتة:\n{}",
    "bcast_history_btn":  "📜 سجل الإعلانات",
    "bcast_history_title":"📜 آخر الإعلانات",
    "bcast_history_empty":"لا توجد إعلانات بعد.",
    "clear_chat_btn":     "🧹 مسح المحادثة",
    "clear_chat_done":    "✅ تم مسح المحادثة.",
    "quick_reply_btn":    "⚡ رد سريع",
    "quick_reply_saved":  "✅ تم حفظ قالب الرد السريع.",
    # ═══ أزرار جديدة ═══
    "btn_notes":          "📝 ملاحظاتي",
    "btn_reminder":       "⏰ تذكيراتي",
    "anon_q_btn":         "❓ سؤال مجهول",
    "countdown_btn":      "⏳ الامتحانات القادمة",
    "quiz_btn":           "📝 اختبار قصير",
    "notes_empty":        "لا توجد ملاحظات بعد.",
    "notes_saved":        "✅ تم حفظ الملاحظة.",
    "notes_title":        "📝 ملاحظاتي الشخصية",
    "notes_prompt":       "📝 اكتب ملاحظتك:",
    "reminder_list":      "⏰ تذكيراتي ({} تذكير):",
    "reminder_none":      "لا توجد تذكيرات.",
    "reminder_add_prompt":"⏰ اكتب نص التذكير:",
    "reminder_saved":     "✅ سأذكرك بعد {}.",
    "reminder_fired":     "🔔 تذكير\n\n{}",
    "reminder_del":       "✅ تم الحذف.",
    "anon_q_prompt":      "❓ اكتب سؤالك (مجهول الهوية):",
    "anon_q_sent":        "✅ تم إرسال سؤالك.",
    "countdown_none":     "لا توجد امتحانات مضافة.",
    "countdown_prompt":   "اكتب اسم الامتحان:",
    "countdown_date":     "اكتب التاريخ (مثال: 20/05/2026):",
    "countdown_saved":    "✅ تم حفظ {}.",
    "quiz_none":          "لا يوجد اختبار نشط.",
    "group_select":       "👥 اختر مجموعتك (A/B/C → مجموعة فرعية):",
    "group_selected":     "✅ تم حفظ مجموعتك: {}",
    "group_change_btn":   "👥 تغيير المجموعة",
    "group_label":        "👥 المجموعة: {}",
}

DEFAULT_WELCOME_AR = (
    "🎓 أهلاً وسهلاً في بوت مهندسي المستقبل! 🚀\n\n"
    "هذا البوت رفيقك الدراسي على طريق التفوّق.\n"
    "ستجد هنا كل ما تحتاجه:\n\n"
    "📁 ملفات المواد والمحاضرات\n"
    "🖼 الصور والمخططات الهندسية\n"
    "📚 المراجع والكتب التقنية\n\n"
    "﴿ وَقُل رَّبِّ زِدْنِي عِلْمًا ﴾ 📖\n\n"
    "استخدم الأزرار أدناه للبدء 👇\n\n"
    "─────────────────────────\n"
    "🤖 هذا البوت من تصميم وتطوير\n"
    "المهندس عبدالرحمن أردال"
)

# ================================================================
#  YARDIMCI FONKSİYONLAR
# ================================================================

def is_main_admin(uid): return int(uid) == ADMIN_ID

def L(uid, key, *args):
    lang = TR if is_main_admin(uid) else AR
    val  = lang.get(key, TR.get(key, key))
    return val.format(*args) if args else val

def load_json(path, default):
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"load_json {path}: {e}")
    return default

def save_json(path, data):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"save_json {path}: {e}")

# ================================================================
#  VARSAYILAN KLASÖR YAPISI — Bot her başladığında eksikse oluşturur
# ================================================================


def _make_group_folders():
    """Her sınıf için A, B, C grubu ve her grubun 3 alt grubu."""
    groups = {}
    for g in ("A", "B", "C"):
        sub_groups = {}
        for i in range(1, 4):
            sub_groups[f"{g}{i}"] = {"folders": {}, "files": []}
        groups[g] = {"folders": sub_groups, "files": []}
    return groups

def _make_folder(*subs):
    """İçi boş klasör şablonu."""
    return {"folders": {s: {"folders": {}, "files": []} for s in subs}, "files": []}

DEFAULT_CONTENT = {
    "folders": {
        "الاول": {
            "folders": {
                "المجموعات": {"folders": _make_group_folders(), "files": []},
                "مواد كورس الأول": _make_folder(
                    "رياضيات", "اسس", "فيزياء", "انجليزي", "ورش", "حقوق", "حاسوب"
                ),
                "مواد كورس الثاني": _make_folder(
                    "اسس", "رياضيات", "فيزياء", "برمجة", "رسم", "عربي", "رقميه"
                ),
                "جدول": {"folders": {}, "files": []},
            },
            "files": [],
        },
        "الثاني": {
            "folders": {
                "المجموعات": {"folders": _make_group_folders(), "files": []},
                "مواد كورس الأول": _make_folder(
                    "رياضيات", "فيزياء", "اسس", "انجليزي", "ورش"
                ),
                "مواد كورس الثاني": _make_folder(
                    "رياضيات", "فيزياء", "ميكانيك", "برمجة", "رسم"
                ),
                "جدول": {"folders": {}, "files": []},
            },
            "files": [],
        },
        "الثالث": {
            "folders": {
                "المجموعات": {"folders": _make_group_folders(), "files": []},
                "مواد كورس الأول": _make_folder(
                    "رياضيات", "فيزياء", "اسس", "انجليزي"
                ),
                "مواد كورس الثاني": _make_folder(
                    "رياضيات", "فيزياء", "مواد متخصصة", "مشروع"
                ),
                "جدول": {"folders": {}, "files": []},
            },
            "files": [],
        },
        "الرابع": {
            "folders": {
                "المجموعات": {"folders": _make_group_folders(), "files": []},
                "مواد كورس الأول": _make_folder(
                    "مواد متخصصة", "مشروع تخرج", "انجليزي"
                ),
                "مواد كورس الثاني": _make_folder(
                    "مواد متخصصة", "مشروع تخرج", "تدريب عملي"
                ),
                "جدول": {"folders": {}, "files": []},
            },
            "files": [],
        },
    },
    "files": [],
}

def load_content():
    data = load_json(DATA_FILE, None)
    if data is None:
        # İlk çalışma — varsayılan yapıyı oluştur ve kaydet
        save_json(DATA_FILE, DEFAULT_CONTENT)
        logger.info("✅ Varsayılan klasör yapısı oluşturuldu.")
        return DEFAULT_CONTENT
    # Eksik ana klasörleri tamamla (güncelleme sonrası koruma)
    changed = False
    for name, val in DEFAULT_CONTENT["folders"].items():
        if name not in data.setdefault("folders", {}):
            data["folders"][name] = val
            changed = True
    # طلاب مؤقتون kaldır
    if TEMP_FOLDER_NAME in data.get("folders", {}):
        del data["folders"][TEMP_FOLDER_NAME]
        changed = True
    if changed:
        save_json(DATA_FILE, data)
        logger.info("Eksik klasorler tamamlandi.")
    return data

def save_content(d):     save_json(DATA_FILE, d)
def load_users():        return load_json(USERS_FILE,       {})
def save_users(d):       save_json(USERS_FILE, d)
def load_messages():     return load_json(MESSAGES_FILE,    {})
def save_messages(d):    save_json(MESSAGES_FILE, d)
def load_admins():       return load_json(ADMINS_FILE,      [])
def save_admins(d):      save_json(ADMINS_FILE, d)
def load_blocked():      return load_json(BLOCKED_FILE,     [])
def save_blocked(d):     save_json(BLOCKED_FILE, d)
def load_view_counts():  return load_json(VIEW_COUNTS_FILE, {})
def save_view_counts(d): save_json(VIEW_COUNTS_FILE, d)
def load_polls():        return load_json(POLLS_FILE, {})
def save_polls(d):       save_json(POLLS_FILE, d)

# ── Yeni veri dosyaları ───────────────────────────────────────
def load_classes():       return load_json(CLASS_FILE, {})
def save_classes(d):      save_json(CLASS_FILE, d)
def load_faq():           return load_json(FAQ_FILE, [])
def save_faq(d):          save_json(FAQ_FILE, d)
def load_auto_rules():    return load_json(AUTO_RULES_FILE, [])
def save_auto_rules(d):   save_json(AUTO_RULES_FILE, d)
def load_favorites():     return load_json(FAVORITES_FILE, {})
def save_favorites(d):    save_json(FAVORITES_FILE, d)
def load_notes():         return load_json(NOTES_FILE, {})
def save_notes(d):        save_json(NOTES_FILE, d)
def load_warns():         return load_json(WARNS_FILE, {})
def save_warns(d):        save_json(WARNS_FILE, d)
def load_admin_log():     return load_json(ADMIN_LOG_FILE, [])
def save_admin_log(d):    save_json(ADMIN_LOG_FILE, d)
def load_subs():          return load_json(SUBS_FILE, {})
def save_subs(d):         save_json(SUBS_FILE, d)
def load_recent():        return load_json(RECENT_FILE, {})
def save_recent(d):       save_json(RECENT_FILE, d)
def load_folder_descs():  return load_json(FOLDER_DESC_FILE, {})
def save_folder_descs(d): save_json(FOLDER_DESC_FILE, d)
def load_whitelist():     return load_json(WHITELIST_FILE, {"enabled": False, "users": []})
def save_whitelist(d):    save_json(WHITELIST_FILE, d)
def load_scheduled():     return load_json(SCHEDULED_FILE, [])
def save_scheduled(d):    save_json(SCHEDULED_FILE, d)
def load_tags():          return load_json(TAGS_FILE, {})
def save_tags(d):         save_json(TAGS_FILE, d)
def load_reminders():     return load_json(REMINDERS_FILE, [])
def save_reminders(d):    save_json(REMINDERS_FILE, d)
def load_search_cache():  return load_json(SEARCH_CACHE_FILE, {})
def save_search_cache(d): save_json(SEARCH_CACHE_FILE, d)
def load_leaderboard():   return load_json(LEADERBOARD_FILE, {})
def save_leaderboard(d):  save_json(LEADERBOARD_FILE, d)
def load_achievements():  return load_json(ACHIEVEMENTS_FILE, {})
def save_achievements(d): save_json(ACHIEVEMENTS_FILE, d)
def load_feedback():      return load_json(FEEDBACK_FILE, {})
def save_feedback(d):     save_json(FEEDBACK_FILE, d)
def load_pinned_msgs():   return load_json(PINNED_MSG_FILE, {})
def save_pinned_msgs(d):  save_json(PINNED_MSG_FILE, d)
def load_bcast_log():     return load_json(BROADCAST_LOG_FILE, [])
def save_bcast_log(d):    save_json(BROADCAST_LOG_FILE, d[-100:])

CLASS_NAMES_FILE = os.path.join(BASE_DIR, "class_names.json")

def load_class_names():    return load_json(CLASS_NAMES_FILE, {})
def save_class_names(d):   save_json(CLASS_NAMES_FILE, d)

def get_class_display_name(cls_id: str) -> str:
    """Sınıfın özel adını döndür, yoksa varsayılan ar adını kullan."""
    names = load_class_names()
    return names.get(str(cls_id), CLASS_DEFS.get(str(cls_id), {}).get("ar", cls_id))

def load_countdowns():    return load_json(COUNTDOWN_FILE, [])
def load_groups():        return load_json(GROUP_FILE, {})
def save_groups(d):       save_json(GROUP_FILE, d)

SHIFT_FILE = os.path.join(BASE_DIR, "user_shifts.json")
def load_shifts():     return load_json(SHIFT_FILE, {})
def save_shifts(d):    save_json(SHIFT_FILE, d)
def get_user_shift(uid: str) -> str:
    return load_shifts().get(str(uid), "")
def set_user_shift(uid: str, shift: str):
    d = load_shifts(); d[str(uid)] = shift; save_shifts(d)

TEMP_FOLDER_NAME = "\u0637\u0644\u0627\u0628 \u0645\u0624\u0642\u062a\u0648\u0646"  # Gecici Ogrenciler
def save_countdowns(d):   save_json(COUNTDOWN_FILE, d)
def load_anon_q():        return load_json(ANON_Q_FILE, [])
def save_anon_q(d):       save_json(ANON_Q_FILE, d)
def load_personal_notes():return load_json(USER_NOTES_FILE, {})
def save_personal_notes(d):save_json(USER_NOTES_FILE, d)
def load_quizzes():       return load_json(QUIZ_FILE, [])
def save_quizzes(d):      save_json(QUIZ_FILE, d)
def load_reports():       return load_json(REPORT_FILE, [])
def load_admin_perms():   return load_json(ADMIN_PERMS_FILE, {})
def save_admin_perms(d):  save_json(ADMIN_PERMS_FILE, d)
def save_reports(d):      save_json(REPORT_FILE, d)

def get_user_notes(uid: str) -> list:
    """Kullanıcının tüm notlarını döndür (liste formatı)."""
    data = load_personal_notes()
    raw = data.get(str(uid), [])
    # Eski format (string) → liste formatına çevir
    if isinstance(raw, str):
        return [{"type":"text","content":raw,"id":"0","time":datetime.now(IRAQ_TZ).strftime("%H:%M")}] if raw else []
    return raw


NOTE_SUBJECTS = [
    "رياضيات", "فيزياء", "اسس", "برمجة", "انجليزي",
    "ميكانيك", "رسم", "كيمياء", "حقوق", "عربي",
    "رقميه", "مواد متخصصة", "مشروع", "أخرى",
]


# ================================================================
#  LABORATUVAR SİSTEMİ
#  Her haftayı bir gruba atayan sistem (B1/B2/B3 dönüşümlü)
# ================================================================
LAB_FILE = os.path.join(BASE_DIR, "lab_schedule.json")


# ── Grup Yapılandırma ──────────────────────────────────────────
CLASS_GROUPS_FILE = os.path.join(BASE_DIR, "class_groups.json")

def load_class_groups() -> dict:
    """Sınıf grup yapılandırması. {cls: {grp: [sub1, sub2, ...]}}"""
    default = {
        "1": {"A": ["A1","A2","A3"], "B": ["B1","B2","B3"], "C": ["C1","C2","C3"]},
        "2": {"A": ["A1","A2","A3"], "B": ["B1","B2","B3"], "C": ["C1","C2","C3"]},
        "3": {"A": ["A1","A2","A3"], "B": ["B1","B2","B3"], "C": ["C1","C2","C3"]},
        "4": {"A": ["A1","A2","A3"], "B": ["B1","B2","B3"], "C": ["C1","C2","C3"]},
    }
    d = load_json(CLASS_GROUPS_FILE, {})
    return d if d else default

def save_class_groups(d): save_json(CLASS_GROUPS_FILE, d)

def get_class_groups(cls_id: str) -> dict:
    """Bir sınıfın gruplarını döndür. {grp: [subs]}"""
    return load_class_groups().get(str(cls_id), {})

# ── Sınıf Kanalları ───────────────────────────────────────────
CLASS_CHANNELS_FILE = os.path.join(BASE_DIR, "class_channels.json")

def load_class_channels() -> dict:
    """Sınıf kanal/bot linkleri. {cls: {"link": ..., "name": ...}}"""
    return load_json(CLASS_CHANNELS_FILE, {})

def save_class_channels(d): save_json(CLASS_CHANNELS_FILE, d)

def load_lab_schedule():  return load_json(LAB_FILE, [])
def save_lab_schedule(d): save_json(LAB_FILE, d)

def get_current_lab_week():
    """Bu haftanın laboratuvar grubunu döndür."""
    from datetime import date
    schedule = load_lab_schedule()
    if not schedule: return None
    today = date.today()
    # Pazartesi başlangıcını bul
    week_start = today - __import__("datetime").timedelta(days=today.weekday())
    week_str = week_start.strftime("%Y-%m-%d")
    for entry in schedule:
        if entry.get("week") == week_str:
            return entry
    return None

def add_lab_week(week_start_str: str, group: str, note: str = ""):
    """Tarihe grup+lab ata. Aynı tarih-grup varsa güncelle, yoksa ekle."""
    schedule = load_lab_schedule()
    # Aynı tarih + aynı grup varsa güncelle
    for entry in schedule:
        if entry.get("week") == week_start_str and entry.get("group") == group:
            entry["note"] = note
            save_lab_schedule(schedule)
            return
    # Yeni ekle — Aynı tarihte farklı gruplar olabilir
    schedule.append({"week": week_start_str, "group": group, "note": note,
                     "added": datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M")})
    schedule = sorted(schedule, key=lambda x: (x["week"], x.get("group","")))
    save_lab_schedule(schedule)

def get_upcoming_lab_weeks(n=4):
    """Yaklaşan n hafta laboratuvar programını döndür."""
    from datetime import date, timedelta
    schedule = load_lab_schedule()
    today    = date.today()
    result   = []
    for entry in schedule:
        try:
            wd = __import__("datetime").date.fromisoformat(entry["week"])
            if wd >= today - timedelta(days=6):
                result.append(entry)
        except: pass
    return result[:n]

def add_user_note(uid: str, note_type: str, content: str, file_id: str = None, subject: str = "أخرى"):
    """Not ekle (type: text/photo/voice/document)."""
    d = load_personal_notes()
    notes = get_user_notes(uid)
    import time
    note = {
        "type": note_type,
        "content": content[:500],
        "id": str(int(time.time())),
        "time": datetime.now(IRAQ_TZ).strftime("%d/%m %H:%M"),
        "subject": subject,
    }
    if file_id:
        note["file_id"] = file_id
    notes.append(note)
    d[str(uid)] = notes[-20:]  # max 20 not
    save_personal_notes(d)

def delete_user_note(uid: str, note_id: str):
    d = load_personal_notes()
    notes = get_user_notes(uid)
    notes = [n for n in notes if n.get("id") != note_id]
    d[str(uid)] = notes
    save_personal_notes(d)

# Eski compat
def get_personal_note(uid: str) -> str:
    notes = get_user_notes(uid)
    return notes[0]["content"] if notes else ""

def set_personal_note(uid: str, text: str):
    add_user_note(uid, "text", text)

def add_reminder(uid: str, text: str, fire_ts: float):
    rems = load_reminders()
    rems.append({"uid": str(uid), "text": text, "fire_ts": fire_ts,
                 "created": datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M")})
    save_reminders(rems)

def get_user_reminders(uid: str) -> list:
    return [r for r in load_reminders() if str(r.get("uid")) == str(uid)]

def delete_reminder(uid: str, idx: int):
    rems = load_reminders()
    user_rems = [(i, r) for i, r in enumerate(rems) if str(r.get("uid")) == str(uid)]
    if 0 <= idx < len(user_rems):
        rems.pop(user_rems[idx][0])
    save_reminders(rems)

def parse_time_input(text: str) -> int:
    """
    Doğal dil zaman ifadesini dakikaya çevirir.
    Örnekler:
      '2 saat'       → 120
      '30 dakika'    → 30
      '1 gün'        → 1440
      '1.5 saat'     → 90
      '2s', '30d'    → kısa format da desteklenir
    """
    import re
    text = text.strip().lower()
    total = 0

    # Türkçe + Arapça zaman kelimeleri
    patterns = [
        (r'(\d+(?:[.,]\d+)?)\s*(?:saat|ساعة|ساعات|hour|hours|s\b)',    60),
        (r'(\d+(?:[.,]\d+)?)\s*(?:gün|يوم|أيام|day|days|g\b)',         1440),
        (r'(\d+(?:[.,]\d+)?)\s*(?:dakika|دقيقة|دقائق|minute|minutes|min|dk|d\b)', 1),
        (r'(\d+(?:[.,]\d+)?)\s*(?:hafta|أسبوع|week|weeks)',             10080),
    ]
    for pattern, multiplier in patterns:
        for m in re.finditer(pattern, text):
            val = float(m.group(1).replace(',', '.'))
            total += int(val * multiplier)

    return total if total > 0 else 0

def add_countdown(name: str, date_str: str, cls: str = "", shift: str = "", group: str = "") -> bool:
    """date_str: 'DD/MM/YYYY'. shift: sabahi/masaiy. group: A/B/C veya boş=hepsi"""
    try:
        from datetime import datetime as dt
        target = dt.strptime(date_str.strip(), "%d/%m/%Y")
        cd = load_countdowns()
        cd.append({"name": name, "date": date_str, "ts": target.timestamp(),
                   "cls": cls, "shift": shift, "group": group,
                   "created": datetime.now(IRAQ_TZ).strftime("%Y-%m-%d")})
        save_countdowns(cd)
        return True
    except: return False

def get_countdowns(cls: str = "", shift: str = "", uid: str = "") -> list:
    """Gelecekteki geri sayımları döndür. Ek sınıf kayıtlarını da dahil eder."""
    import time
    now = time.time()
    cd  = load_countdowns()
    # Kullanıcının ek sınıflarını al
    extra_cls = []
    if uid:
        extra_data = load_json(os.path.join(BASE_DIR, "user_extra_cls.json"), {})
        extra_cls  = extra_data.get(str(uid), [])
    # Ana sınıf + ek sınıflar
    all_cls = []
    if cls: all_cls.append(cls)
    for ec in extra_cls:
        if ec not in all_cls: all_cls.append(ec)
    result = []
    for c in cd:
        if c.get("ts", 0) > now:
            cd_cls   = c.get("cls", "")
            cd_shift = c.get("shift", "")
            # Sınıf — ana veya ek sınıftan biri eşleşmeli
            if all_cls and cd_cls and cd_cls not in all_cls:
                continue
            # Grup — A/B/C filtresi
            cd_grp = c.get("group","")
            if cd_grp and uid:
                u_grp = load_groups().get(str(uid),"")
                if not u_grp.startswith(cd_grp):
                    continue
            # Shift
            if cd_shift and shift:
                u_norm = "sabahi" if shift in ("sabahi","sabah") else "masaiy"
                c_norm = "sabahi" if cd_shift in ("sabahi","sabah") else "masaiy"
                if u_norm != c_norm:
                    continue
            days_left = int((c["ts"] - now) / 86400)
            result.append({**c, "days_left": days_left})
    return sorted(result, key=lambda x: x["ts"])

# ================================================================
#  WEB ARAMA & YAPAY ZEKA MOTORU
#  Dışarıya bağlı değil — DuckDuckGo + Wikipedia + Python hesap
# ================================================================

import re as _re
import asyncio
import html as _html

# ── Başarı Sistemi ─────────────────────────────────────────────
ACHIEVEMENT_DEFS = {
    "first_view":   {"name": "أول ملف 👁", "desc": "شاهدت أول ملف", "threshold": 1,  "metric": "views"},
    "ten_views":    {"name": "10 ملفات 📚", "desc": "شاهدت 10 ملفات", "threshold": 10, "metric": "views"},
    "fifty_views":  {"name": "50 ملف 🌟",   "desc": "شاهدت 50 ملف",  "threshold": 50, "metric": "views"},
    "first_fav":    {"name": "مفضلة ⭐",    "desc": "أضفت أول مفضلة", "threshold": 1,  "metric": "favs"},
    "explorer":     {"name": "مستكشف 🗺",   "desc": "فتحت 5 مجلدات", "threshold": 5,  "metric": "folders"},
}

def check_achievements(uid: str, context_data: dict):
    """Yeni başarılar kazanıldıysa listelerini döndür."""
    achs = load_achievements()
    user_achs = achs.get(str(uid), [])
    new_unlocked = []
    for key, defn in ACHIEVEMENT_DEFS.items():
        if key in user_achs: continue
        metric = defn["metric"]
        value  = context_data.get(metric, 0)
        if value >= defn["threshold"]:
            user_achs.append(key)
            new_unlocked.append(defn["name"])
    if new_unlocked:
        achs[str(uid)] = user_achs
        save_achievements(achs)
    return new_unlocked

# ── Liderboard ────────────────────────────────────────────────
def update_leaderboard(uid: str, points: int = 1):
    lb = load_leaderboard()
    uid = str(uid)
    lb[uid] = lb.get(uid, 0) + points
    save_leaderboard(lb)

def get_leaderboard(top_n: int = 10) -> list:
    lb    = load_leaderboard()
    users = load_users()
    result = []
    for uid, pts in sorted(lb.items(), key=lambda x: x[1], reverse=True)[:top_n]:
        u    = users.get(uid, {})
        name = u.get("full_name") or u.get("first_name") or f"ID:{uid}"
        result.append((name, pts))
    return result

# ── Dosya Geri Bildirimi ──────────────────────────────────────
def add_feedback(file_key: str, uid: str, stars: int):
    fb  = load_feedback()
    key = file_key[:100]
    if key not in fb:
        fb[key] = {}
    fb[key][str(uid)] = stars
    save_feedback(fb)

def get_feedback_stats(file_key: str) -> dict:
    fb    = load_feedback()
    votes = fb.get(file_key[:100], {})
    if not votes: return {"avg": 0, "count": 0}
    avg = sum(votes.values()) / len(votes)
    return {"avg": avg, "count": len(votes)}

# ── Sabit Mesaj ───────────────────────────────────────────────
def get_pinned_msg() -> str:
    msgs = load_pinned_msgs()
    return msgs.get("global", "")

def set_pinned_msg(text: str):
    msgs = load_pinned_msgs()
    if text.strip():
        msgs["global"] = text.strip()
    else:
        msgs.pop("global", None)
    save_pinned_msgs(msgs)

# ── Sınıf sistemi ─────────────────────────────────────────────

# Varsayılan admin yetkileri
DEFAULT_ADMIN_PERMS = {
    "can_warn":        True,   # Uyarı (süper admin onayı gerekli)
    "can_block":       True,   # Engelleme (süper admin onayı gerekli)
    "can_broadcast":   True,   # Duyuru
    "can_reply":       True,   # Kullanıcıya cevap
    "can_add_folder":  True,   # Klasör ekle
    "can_del_folder":  False,  # Klasör sil — kapalı
    "can_add_file":    True,   # Dosya ekle
    "can_del_file":    False,  # Dosya sil — kapalı
    "can_rename_file": False,  # Yeniden adlandır — kapalı
    "can_poll":        True,   # Anket kur
    "can_quiz":        False,  # Quiz — kapalı
    "can_countdown":   True,   # Sınav ekle
    "can_view_users":  False,  # Kullanıcı görme — kapalı
    "cls":             None,   # None = tüm sınıflar
}

def get_admin_perm(admin_uid: str, perm: str) -> bool:
    """Admin'in belirli bir yetkisi var mı?"""
    if is_main_admin(admin_uid): return True
    perms = load_admin_perms()
    admin_perms = perms.get(str(admin_uid), DEFAULT_ADMIN_PERMS)
    return admin_perms.get(perm, DEFAULT_ADMIN_PERMS.get(perm, False))

def get_admin_cls(admin_uid: str):
    """Admin hangi sınıfı yönetiyor? None=hepsi."""
    if is_main_admin(admin_uid): return None
    perms = load_admin_perms()
    return perms.get(str(admin_uid), {}).get("cls", None)

def get_admin_grp(admin_uid: str):
    """Admin hangi grubu yönetiyor? None=hepsi (A/B/C)."""
    if is_main_admin(admin_uid): return None
    perms = load_admin_perms()
    return perms.get(str(admin_uid), {}).get("grp", None)

def get_admin_subgrp(admin_uid: str):
    """Admin hangi alt grubu yönetiyor? None=hepsi (A1/A2/A3 vb.)."""
    if is_main_admin(admin_uid): return None
    perms = load_admin_perms()
    return perms.get(str(admin_uid), {}).get("subgrp", None)

def admin_can_manage_user(admin_uid: str, user_uid: str) -> bool:
    """Admin bu kullanıcıyı yönetebilir mi? (sınıf+grup+subgrp kontrolü)"""
    if is_main_admin(admin_uid): return True
    adm_cls    = get_admin_cls(admin_uid)
    adm_grp    = get_admin_grp(admin_uid)
    adm_subgrp = get_admin_subgrp(admin_uid)
    usr_cls    = get_user_class(user_uid)
    usr_grp    = get_user_group(user_uid)  # A1/A2/B3 vb.
    # Sınıf kontrolü
    if adm_cls and usr_cls != adm_cls:
        return False
    # Grup kontrolü (A/B/C)
    if adm_grp and (not usr_grp or not usr_grp.startswith(adm_grp)):
        return False
    # Alt grup kontrolü (A1/A2/A3)
    if adm_subgrp and usr_grp != adm_subgrp:
        return False
    return True

def set_admin_perm(admin_uid: str, perm: str, value):
    perms = load_admin_perms()
    if str(admin_uid) not in perms:
        perms[str(admin_uid)] = dict(DEFAULT_ADMIN_PERMS)
    perms[str(admin_uid)][perm] = value
    save_admin_perms(perms)


def build_target_keyboard(prefix: str, step: str, val: str = "") -> list:
    """
    Adım adım hedefleme klavyesi.
    step: "cls" | "shift" | "grp"
    prefix: callback_data başı (örn: "poll|tgt" veya "bcast|tgt")
    val: önceki seçimler ("|" ile ayrılmış)
    """
    sep = "|" if val else ""
    base = f"{prefix}|{val}{sep}" if val else f"{prefix}|"

    if step == "cls":
        kb = [
            [InlineKeyboardButton("🌐 الجميع", callback_data=f"{base}cls_all|shift")],
        ]
        for cls_id, cls_def in CLASS_DEFS.items():
            lbl = get_class_display_name(cls_id)
            kb.append([InlineKeyboardButton(lbl, callback_data=f"{base}cls_{cls_id}|shift")])
        return kb

    if step == "shift":
        kb = [
            [InlineKeyboardButton("🌐 الكل",    callback_data=f"{base}sft_all|grp"),
             InlineKeyboardButton("☀️ صباحي",  callback_data=f"{base}sft_sabahi|grp"),
             InlineKeyboardButton("🌙 مسائي",  callback_data=f"{base}sft_masaiy|grp")],
        ]
        return kb

    if step == "grp":
        # Grup seçimi — seçince direkt done (alt grup yok)
        kb = [
            [InlineKeyboardButton("🌐 الكل", callback_data=f"{base}grp_all|done"),
             InlineKeyboardButton("A",        callback_data=f"{base}grp_A|done"),
             InlineKeyboardButton("B",        callback_data=f"{base}grp_B|done"),
             InlineKeyboardButton("C",        callback_data=f"{base}grp_C|done")],
        ]
        return kb

    return []

def parse_target_val(val: str) -> dict:
    """
    'cls_1|sft_sabahi|grp_A|sub_A2' → {cls:"1", shift:"sabahi", grp:"A", subgrp:"A2"}
    """
    result = {"cls": None, "shift": None, "grp": None, "subgrp": None}
    for part in val.split("|"):
        part = part.strip()
        if part.startswith("cls_") and part != "cls_all":
            result["cls"] = part[4:]
        elif part.startswith("sft_") and part != "sft_all":
            result["shift"] = part[4:]
        elif part.startswith("grp_") and part != "grp_all":
            result["grp"] = part[4:]
        elif part.startswith("sub_") and part != "sub_all":
            result["subgrp"] = part[4:]
    return result

def get_target_users(target_val: str) -> list:
    """Hedef kullanıcı listesini döndür."""
    t     = parse_target_val(target_val)
    users = load_users()
    shfts = load_shifts()
    grps  = load_groups()
    result = []
    for uid_ in users:
        if int(uid_) == ADMIN_ID or is_blocked(uid_): continue
        if t["cls"] and get_user_class(uid_) != t["cls"]: continue
        if t["shift"]:
            u_shift = shfts.get(uid_, "")
            # sabah/gece eski değerleri de eşleştir
            if t["shift"] == "sabahi" and u_shift not in ("sabahi","sabah"): continue
            if t["shift"] == "masaiy" and u_shift not in ("masaiy","gece"): continue
        if t["grp"]:
            u_grp = grps.get(uid_, "")
            if not u_grp.startswith(t["grp"]): continue
        if t["subgrp"] and grps.get(uid_,"") != t["subgrp"]: continue
        result.append(uid_)
    return result

def target_label(target_val: str) -> str:
    """Hedef açıklaması."""
    t = parse_target_val(target_val)
    parts = []
    if t["cls"]: parts.append(get_class_display_name(t["cls"]))
    if t["shift"]: parts.append("صباحي" if t["shift"]=="sabahi" else "مسائي")
    if t["grp"]: parts.append(f"Grup {t['grp']}")
    if t["subgrp"]: parts.append(t["subgrp"])
    return " / ".join(parts) if parts else "الجميع"

def get_user_group(uid: str) -> str:
    """A1/A2/A3/B1/B2/B3/C1/C2/C3 formatında grup döndür."""
    return load_groups().get(str(uid), "")

def set_user_group(uid: str, group: str):
    d = load_groups(); d[str(uid)] = group; save_groups(d)

def users_by_group(cls: str, group: str) -> list:
    """Belirli sınıf ve gruptaki kullanıcıları döndür."""
    cls_users = users_by_class(cls)
    groups    = load_groups()
    return [u for u in cls_users if groups.get(u) == group]

def get_user_class(uid: str) -> str:
    """Kullanıcının seçtiği sınıfı döndürür. None = seçmemiş."""
    return load_classes().get(str(uid))

def set_user_class(uid: str, cls: str):
    classes = load_classes()
    classes[str(uid)] = cls
    save_classes(classes)

def class_label(uid: str) -> str:
    """Kullanıcının sınıfını görüntülenebilir metin olarak döndürür."""
    cls = get_user_class(str(uid))
    if not cls: return L(uid, "class_unknown")
    return get_class_display_name(cls)

def users_by_class(cls: str) -> list:
    """Belirli bir sınıftaki kullanıcı ID listesi."""
    classes = load_classes()
    return [uid for uid, c in classes.items() if c == cls]

# ── Spam / Rate-limit ─────────────────────────────────────────
def check_rate_limit(uid: str) -> bool:
    """True = geçebilir, False = spam."""
    import time
    now  = time.time()
    lst  = _rate_cache.get(uid, [])
    lst  = [t for t in lst if now - t < 60]   # son 60 saniye
    if len(lst) >= SPAM_LIMIT:
        _rate_cache[uid] = lst
        return False
    lst.append(now)
    _rate_cache[uid] = lst
    return True

# ── İç Yapay Zeka (FAQ + Kural motoru) ───────────────────────
# ── Admin İşlem Günlüğü ───────────────────────────────────────
def log_admin_action(admin_uid: str, action: str, detail: str = ""):
    log = load_admin_log()
    log.append({
        "uid":    str(admin_uid),
        "action": action,
        "detail": str(detail)[:80],
        "time":   datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M")
    })
    save_admin_log(log[-500:])

# ── Uyarı Sistemi ─────────────────────────────────────────────
def get_warns(uid: str) -> list:
    return load_warns().get(str(uid), [])

def add_warn(uid: str, reason: str, by_uid: str) -> int:
    warns = load_warns()
    uid   = str(uid)
    warns.setdefault(uid, []).append({
        "reason": reason, "by": by_uid,
        "time": datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M")
    })
    save_warns(warns)
    return len(warns[uid])

def clear_warns(uid: str):
    warns = load_warns()
    warns.pop(str(uid), None)
    save_warns(warns)

# ── Kullanıcı Notu ────────────────────────────────────────────
def get_note(uid: str) -> str:
    return load_notes().get(str(uid), "")

def set_note(uid: str, note: str):
    notes = load_notes()
    if note.strip():
        notes[str(uid)] = note.strip()
    else:
        notes.pop(str(uid), None)
    save_notes(notes)

# ── Favoriler ─────────────────────────────────────────────────
def get_favorites(uid: str) -> list:
    return load_favorites().get(str(uid), [])

def toggle_favorite(uid: str, f: dict) -> bool:
    """True=eklendi, False=çıkarıldı."""
    favs = load_favorites()
    uid  = str(uid)
    key  = f.get("file_id") or f.get("caption","?")
    lst  = favs.get(uid, [])
    match = [x for x in lst if (x.get("file_id") or x.get("caption","?")) == key]
    if match:
        for x in match: lst.remove(x)
        favs[uid] = lst
        save_favorites(favs)
        return False
    lst.append(f)
    favs[uid] = lst[-50:]
    save_favorites(favs)
    return True

def is_in_favorites(uid: str, f: dict) -> bool:
    key  = f.get("file_id") or f.get("caption","?")
    return any((x.get("file_id") or x.get("caption","?")) == key for x in get_favorites(uid))

# ── Son Görüntülenenler ───────────────────────────────────────
def add_recently_viewed(uid: str, f: dict):
    recent = load_recent()
    uid    = str(uid)
    lst    = recent.get(uid, [])
    key    = f.get("file_id") or f.get("caption","?")
    lst    = [x for x in lst if (x["file"].get("file_id") or x["file"].get("caption","?")) != key]
    lst.append({"file": f, "time": datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M")})
    recent[uid] = lst[-RECENT_MAX:]
    save_recent(recent)

def get_recently_viewed(uid: str) -> list:
    return load_recent().get(str(uid), [])

# ── Klasör Aboneliği ──────────────────────────────────────────
def path_key(path: list) -> str:
    return "~~".join(path) if path else "__root__"

def is_subscribed(uid: str, path: list) -> bool:
    return path_key(path) in load_subs().get(str(uid), [])

def toggle_sub(uid: str, path: list) -> bool:
    subs = load_subs(); uid = str(uid); key = path_key(path)
    lst  = subs.setdefault(uid, [])
    if key in lst: lst.remove(key); subs[uid]=lst; save_subs(subs); return False
    lst.append(key); subs[uid]=lst; save_subs(subs); return True

def get_folder_subscribers(path: list) -> list:
    key = path_key(path)
    return [uid for uid, lst in load_subs().items() if key in lst]

# ── Klasör Açıklaması ─────────────────────────────────────────
def get_folder_desc(path: list) -> str:
    return load_folder_descs().get(path_key(path), "")

def set_folder_desc(path: list, desc: str):
    descs = load_folder_descs(); key = path_key(path)
    if desc.strip(): descs[key] = desc.strip()
    else: descs.pop(key, None)
    save_folder_descs(descs)

# ── Whitelist / Gizli Mod ─────────────────────────────────────

def all_folder_paths(node=None, prefix=None) -> list:
    if node is None:   node = load_content()
    if prefix is None: prefix = []
    result = []
    for name, sub in node.get("folders", {}).items():
        p = prefix + [name]
        result.append(p)
        result.extend(all_folder_paths(sub, p))
    return result

# ── Etkin Kullanıcı Filtresi ──────────────────────────────────
def active_users(days: int = 7) -> list:
    from datetime import timedelta
    cutoff = (datetime.now(IRAQ_TZ) - timedelta(days=days)).strftime("%Y-%m-%d")
    users  = load_users()
    return [uid for uid, u in users.items()
            if u.get("last_seen","") >= cutoff and int(uid) != ADMIN_ID]

def new_users(days: int = 7) -> list:
    from datetime import timedelta
    msgs   = load_messages()
    cutoff = (datetime.now(IRAQ_TZ) - timedelta(days=days)).strftime("%Y-%m-%d")
    users  = load_users()
    new    = []
    for uid, u in users.items():
        if int(uid) == ADMIN_ID: continue
        first_seen = None
        for m in msgs.get(uid, []):
            if m.get("type") == "command" and "/start" in m.get("content",""):
                first_seen = m.get("time","")[:10]; break
        if first_seen and first_seen >= cutoff:
            new.append(uid)
    return new

# ── Dosya görüntüleme sayacı ──────────────────────────────────
def get_file_view_count(f: dict) -> int:
    key = f.get("file_id") or f.get("caption","unknown")
    vc  = load_view_counts()
    v   = vc.get(key)
    return v.get("count", 0) if isinstance(v, dict) else 0

def make_poll_id():
    return datetime.now(IRAQ_TZ).strftime("%Y%m%d%H%M%S")

def poll_bar(count, total):
    """Görsel oy barı."""
    if total == 0: pct = 0
    else: pct = count / total * 100
    filled = int(pct / 10)
    bar = "█" * filled + "░" * (10 - filled)
    return f"{bar}  {pct:.0f}%  ({count} oy)"

def build_poll_results_text(poll, uid, show_voters=False):
    """
    Anket sonuçlarını güzel formatta oluşturur.
    show_voters=True → kimin ne seçtiği de gösterilir (sadece admin için)
    """
    votes  = poll.get("votes", {})   # {uid: "seçenek"}
    opts   = poll.get("options", [])
    total  = len(votes)
    counts = {o: 0 for o in opts}
    voters = {o: [] for o in opts}   # {seçenek: [kullanıcı adları]}

    users_data = load_users()
    for voter_uid, chosen in votes.items():
        if chosen in counts:
            counts[chosen] += 1
            u    = users_data.get(str(voter_uid), {})
            name = u.get("full_name") or u.get("first_name") or f"ID:{voter_uid}"
            un   = f" @{u['username']}" if u.get("username") else ""
            voters[chosen].append(f"{name}{un}")

    lines = [
        f"📊 {'نتائج الاستطلاع' if not is_main_admin(uid) else 'Anket Sonuçları'}",
        f"",
        f"❓ {poll['question']}",
        f"",
        f"👥 {'مجموع الأصوات' if not is_main_admin(uid) else 'Toplam Oy'}: {total}",
        "",
    ]

    for o in opts:
        lines.append(f"\n🔹 {o}")
        lines.append(f"   {poll_bar(counts[o], total)}")
        if show_voters and voters[o]:
            for v in voters[o][:10]:  # max 10 isim göster
                lines.append(f"   ├ 👤 {v}")
            if len(voters[o]) > 10:
                lines.append(f"   └ ... +{len(voters[o])-10} kişi daha")

    lines.append("")
    active_str = "✅ Aktif" if poll.get("active") else "🔴 Kapalı"
    if is_main_admin(uid):
        lines.append(active_str)

    return "\n".join(lines)

def load_settings():
    default = {
        "maintenance":       False,
        "maintenance_text":  "🔧 البوت قيد التحديث، يرجى المحاولة لاحقاً...",
        "bot_name":          "بوت مهندسي المستقبل",
        "welcome_msg":       DEFAULT_WELCOME_AR,
        "bot_photo_id":      None,
        "exam_remind_days":  1,   # Sınavdan kaç gün önce bildirim
        "lab_remind_days":   1,   # Lab'dan kaç gün önce bildirim
        "anon_group_id":     None, # Anonim mesaj gönderilecek grup ID
        "lab_remind_hour":   20,  # Laboratuvar bildiriminin saati (20:00)
        "user_buttons": {
            "btn_search": True,
            "btn_help":   True,
        },
    }
    data = load_json(SETTINGS_FILE, default)
    for k, v in default.items():
        data.setdefault(k, v)
    return data

def save_settings(d): save_json(SETTINGS_FILE, d)

def is_admin(uid):
    return int(uid) == ADMIN_ID or int(uid) in [int(x) for x in load_admins()]

def is_blocked(uid):
    return int(uid) in [int(x) for x in load_blocked()]

def register_user(user) -> bool:
    """Kullanıcıyı kaydeder. Yeni kullanıcıysa True döner."""
    users = load_users()
    uid   = str(user.id)
    is_new = uid not in users
    users[uid] = {
        "id":         user.id,
        "first_name": user.first_name or "",
        "last_name":  user.last_name  or "",
        "username":   user.username   or "",
        "full_name":  user.full_name  or "",
        "last_seen":  datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M:%S"),
    }
    save_users(users)
    return is_new

def log_user_message(user, msg_type: str, content: str, file_id: str = None):
    uid = str(user.id)
    if is_main_admin(uid): return  # sadece süper admin hariç, diğer adminler kaydedilir
    msgs = load_messages()
    entry = {"time": datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M:%S"),
             "type": msg_type, "content": str(content)[:300]}
    if file_id: entry["file_id"] = file_id
    msgs.setdefault(uid, []).append(entry)
    msgs[uid] = msgs[uid][-2000:]
    save_messages(msgs)

def increment_view(f: dict):
    """Dosya görüntüleme sayacını artırır."""
    key  = f.get("file_id") or f.get("caption","unknown")
    name = f.get("caption", f.get("name","?"))
    vc   = load_view_counts()
    if key not in vc:
        vc[key] = {"count": 0, "name": name}
    vc[key]["count"] += 1
    vc[key]["name"]   = name   # isim güncellenmiş olabilir
    save_view_counts(vc)

def get_folder(content, path):
    cur = content
    for name in path:
        cur = cur.setdefault("folders", {}).setdefault(name, {"folders": {}, "files": []})
    return cur

def folder_file_names(folder: dict) -> list:
    return [f.get("caption", f.get("name", "")) for f in folder.get("files", [])]

def folder_item_count(folder: dict) -> str:
    """Klasör içindeki eleman sayısını kısa metin olarak döner."""
    n_files   = len(folder.get("files",   []))
    n_folders = len(folder.get("folders", {}))
    parts = []
    if n_folders: parts.append(f"{n_folders}📁")
    if n_files:   parts.append(f"{n_files}📎")
    return " ".join(parts) if parts else ""

async def download_file(context, file_id: str, filename: str) -> Optional[str]:
    """Dosyayı diske indir. Volume yoksa veya 20MB+ ise atla, file_id yeterli."""
    try:
        tg_file = await context.bot.get_file(file_id)
        # Telegram 20MB üzeri dosyalara get_file verir ama download bazen başarısız olur
        safe = "".join(c for c in filename if c.isalnum() or c in "._- ").strip() or \
               f"file_{datetime.now(IRAQ_TZ).strftime('%Y%m%d%H%M%S')}"
        dest = os.path.join(FILES_DIR, safe)
        base, ext = os.path.splitext(dest)
        n = 1
        while os.path.exists(dest):
            dest = f"{base}_{n}{ext}"; n += 1
        await tg_file.download_to_drive(dest)
        logger.info(f"✅ İndirildi: {dest}")
        return dest
    except Exception as e:
        logger.warning(f"⚠️ İndirme atlandı ({filename}): {e} — file_id kullanılacak")
        return None

# ================================================================
#  STATES
# ================================================================
(WAIT_FOLDER, WAIT_FILE,
 WAIT_ADMIN_ID, WAIT_RENAME_FOLDER, WAIT_RENAME_FILE,
 WAIT_BOT_NAME, WAIT_WELCOME, WAIT_PHOTO,
 WAIT_DM_USER_ID, WAIT_DM_MSG, WAIT_BROADCAST_MSG,
 WAIT_SEARCH, WAIT_USER_MSG,
 WAIT_POLL_QUESTION, WAIT_POLL_OPTIONS, WAIT_POLL_COMMENT,
 WAIT_WARN_REASON, WAIT_USER_NOTE, WAIT_FOLDER_DESC,
 WAIT_FAQ_Q, WAIT_FAQ_A,
 WAIT_RULE_KW, WAIT_RULE_RESP,
 WAIT_SCHEDULE_HOURS, WAIT_SCHEDULE_MSG,
 WAIT_TAG, WAIT_SECRET_ID) = range(27)

ALL_BTNS = {
    TR["btn_content"],   AR["btn_content"],
    TR["btn_mgmt"],
    TR["btn_settings"],
    TR["btn_maint"],     AR["btn_maint"],
    TR["btn_search"],    AR["btn_search"],
    TR["btn_help"],      AR["btn_help"],
    TR["btn_favorites"], AR["btn_favorites"],
    TR["btn_recent"],    AR["btn_recent"],
    TR["profile_btn"],   AR["profile_btn"],
    TR["btn_leaderboard"],AR["btn_leaderboard"],
    TR["btn_notes"],     AR["btn_notes"],
    TR["btn_reminder"],  AR["btn_reminder"],
    TR["countdown_btn"], AR["countdown_btn"],
    TR["quiz_btn"],      AR["quiz_btn"],
    "\U0001f52c \u0627\u0644\u0645\u062e\u062a\u0628\u0631",
    "🔬 المختبر",
}

# ================================================================
#  ARAMA
# ================================================================

def search_content(query: str, node=None, path=None, results=None):
    if node    is None: node    = load_content()
    if path    is None: path    = []
    if results is None: results = []
    q = query.lower()
    for idx, f in enumerate(node.get("files", [])):
        name = f.get("caption", f.get("name", ""))
        if q in name.lower():
            results.append({"path": path[:], "name": name, "idx": idx,
                            "type": f.get("type",""), "is_folder": False})
    for fname, folder in node.get("folders", {}).items():
        if q in fname.lower():
            results.append({"path": path[:], "name": fname, "idx": None,
                            "type": "folder", "is_folder": True})
        search_content(query, folder, path + [fname], results)
    return results

async def do_search(message, uid: str, query: str):
    # Kullanıcı: kendi sınıfında ara
    if not is_admin(uid):
        cls    = get_user_class(uid)
        cls_ar = CLASS_DEFS.get(cls, {}).get("ar", "") if cls else ""
        def _nar(s): return s.replace("أ","ا").replace("إ","ا").replace("آ","ا")
        all_c  = load_content()
        if cls_ar:
            node = next((v for k,v in all_c.get("folders",{}).items()
                         if _nar(k) == _nar(cls_ar)), all_c)
        else:
            node = all_c
        results = search_content(query, node=node)
    else:
        results = search_content(query)

    if not results:
        await message.reply_text(L(uid, "search_none").format(query))
        return

    kb = []
    for r in results[:12]:
        path_str = " › ".join(r["path"]) if r["path"] else ""
        if r["is_folder"]:
            fp_enc = "~~".join(r["path"] + [r["name"]])
            label  = f"📁 {r['name'][:30]}"
            if path_str: label += f"  ({path_str[:18]})"
            kb.append([InlineKeyboardButton(label + " →",
                callback_data=f"goto_folder|{fp_enc}")])
        else:
            path_encoded = "~~".join(r["path"]) if r["path"] else ""
            label = f"📎 {r['name'][:38]}"
            kb.append([InlineKeyboardButton(label,
                callback_data=f"srch|{path_encoded}|{r['idx']}")])

    if not kb:
        await message.reply_text(L(uid, "search_none").format(query))
        return
    kb.append([InlineKeyboardButton(L(uid, "back"), callback_data="nav|root")])
    await message.reply_text(
        L(uid, "search_results").format(query),
        reply_markup=InlineKeyboardMarkup(kb))

# ================================================================
#  KLAVYELER
# ================================================================

def reply_kb(uid):
    uid = str(uid)
    if is_main_admin(uid):
        return ReplyKeyboardMarkup([
            [KeyboardButton(TR["btn_content"]),  KeyboardButton(TR["btn_mgmt"])],
            [KeyboardButton(TR["btn_settings"]), KeyboardButton(TR["btn_maint"])],
            [KeyboardButton(TR["btn_search"]),   KeyboardButton(TR["btn_help"])],
        ], resize_keyboard=True)
    elif is_admin(uid):
        s  = load_settings()
        ub = s.get("user_buttons", {})
        rows = []
        # Satır 1: Ara + Mesaj Gönder (kullanıcı gibi)
        row1 = []
        if ub.get("btn_search",   True): row1.append(KeyboardButton(AR["btn_search"]))
        if ub.get("btn_help",     True): row1.append(KeyboardButton(AR["btn_help"]))
        if row1: rows.append(row1)
        # Satır 2: Profil
        rows.append([KeyboardButton(AR["profile_btn"])])
        # Satır 3: Favoriler + Son
        row3 = []
        if ub.get("btn_favorites",True): row3.append(KeyboardButton(AR["btn_favorites"]))
        if ub.get("btn_recent",   True): row3.append(KeyboardButton(AR["btn_recent"]))
        if row3: rows.append(row3)
        # Satır 4: Notlar + Hatırlatıcı
        row4 = []
        if ub.get("btn_notes",    True): row4.append(KeyboardButton(AR["btn_notes"]))
        if ub.get("btn_reminder", True): row4.append(KeyboardButton(AR["btn_reminder"]))
        if row4: rows.append(row4)
        # Satır 5: Sınavlar + Lab
        row5 = []
        if ub.get("countdown_btn",True): row5.append(KeyboardButton(AR["countdown_btn"]))
        row5.append(KeyboardButton("🔬 المختبر"))
        if row5: rows.append(row5)
        # Satır 6: Liderboard + Quiz
        row6 = []
        if ub.get("btn_leaderboard",True): row6.append(KeyboardButton(AR["btn_leaderboard"]))
        if ub.get("quiz_btn",     True): row6.append(KeyboardButton(AR["quiz_btn"]))
        if row6: rows.append(row6)
        # Admin özel: İçerik + Bakım
        rows.append([KeyboardButton(AR["btn_content"]), KeyboardButton(AR["btn_maint"])])
        return ReplyKeyboardMarkup(rows, resize_keyboard=True)
    else:
        s  = load_settings()
        ub = s.get("user_buttons", {})
        rows = []
        # Satır 1: Ara + Mesaj Gönder
        row1 = []
        if ub.get("btn_search",   True): row1.append(KeyboardButton(AR["btn_search"]))
        if ub.get("btn_help",     True): row1.append(KeyboardButton(AR["btn_help"]))
        if row1: rows.append(row1)
        # Satır 2: AI Asistan + Profilim
        row2 = []
        row2.append(KeyboardButton(AR["profile_btn"]))
        if row2: rows.append(row2)
        # Satır 3: Favoriler + Son Görüntülenen
        row3 = []
        if ub.get("btn_favorites",True): row3.append(KeyboardButton(AR["btn_favorites"]))
        if ub.get("btn_recent",   True): row3.append(KeyboardButton(AR["btn_recent"]))
        if row3: rows.append(row3)
        # Satır 4: Notlarım + Hatırlatıcım
        row4 = []
        if ub.get("btn_notes",    True): row4.append(KeyboardButton(AR["btn_notes"]))
        if ub.get("btn_reminder", True): row4.append(KeyboardButton(AR["btn_reminder"]))
        if row4: rows.append(row4)
        # Satır 5: Anonim Soru + Geri Sayım
        row5 = []
        if ub.get("countdown_btn",True): row5.append(KeyboardButton(AR["countdown_btn"]))
        row5.append(KeyboardButton("🔬 المختبر"))
        if row5: rows.append(row5)
        # Satır 6: Liderboard + Test
        row6 = []
        if ub.get("btn_leaderboard", True): row6.append(KeyboardButton(AR["btn_leaderboard"]))
        if ub.get("quiz_btn",     True): row6.append(KeyboardButton(AR["quiz_btn"]))
        if row6: rows.append(row6)
        if not rows: rows = [[]]
        return ReplyKeyboardMarkup(rows, resize_keyboard=True)

def folder_text(folder, path, uid):
    uid    = str(uid)
    header = "📂 " + " › ".join(path) if path else L(uid, "home")
    lines  = [header, ""]
    # Klasör açıklaması
    desc = get_folder_desc(path)
    if desc:
        lines.append(f"💬 {desc}")
        lines.append("")

    folds  = folder.get("folders", {})
    files  = folder.get("files",   [])

    # Kullanıcılar tüm sınıfların dosyalarını görebilir
    # Kendi sınıfı önde göster
    if not path and not is_admin(uid):
        cls    = get_user_class(uid)
        cls_ar = CLASS_DEFS.get(cls, {}).get("ar", "") if cls else ""
        if cls_ar:
            def _nar_ft(s): return s.replace("أ","ا").replace("إ","ا").replace("آ","ا")
            cls_ar_nft = _nar_ft(cls_ar)
            # Kendi sınıfı en üste al
            own   = {k: v for k, v in folds.items() if _nar_ft(k) == cls_ar_nft}
            other = {k: v for k, v in folds.items() if _nar_ft(k) != cls_ar_nft}
            folds = {**own, **other}

    if folds:
        lines.append(L(uid, "folder_list"))
        for name, sub in folds.items():
            cnt = folder_item_count(sub)
            lines.append(f"  • {name}" + (f"  ({cnt})" if cnt else ""))
    if files:
        lines.append(L(uid, "file_list"))
        pinned = [f for f in files if f.get("pinned")]
        normal = [f for f in files if not f.get("pinned")]
        for f in pinned + normal:
            pin = "📌" if f.get("pinned") else "  "
            lines.append(f" {pin} {f.get('caption', f.get('name','?'))}")
    if is_admin(uid):
        s = load_settings()
        lines.append("")
        lines.append(L(uid, "maint_on") if s["maintenance"] else L(uid, "maint_off"))
    return "\n".join(lines)

def folder_kb(path, folder, uid, page=0):
    uid         = str(uid)
    kb          = []
    raw_folders = folder.get("folders", {})
    # Kullanıcı tüm sınıfları görür, kendi sınıfı önde
    if not path and not is_admin(uid):
        cls    = get_user_class(uid)
        cls_ar = CLASS_DEFS.get(cls, {}).get("ar", "") if cls else ""
        if cls_ar:
            def _nar_kb(s): return s.replace("أ","ا").replace("إ","ا").replace("آ","ا")
            cls_ar_nkb = _nar_kb(cls_ar)
            own_f  = {k: v for k, v in raw_folders.items() if _nar_kb(k) == cls_ar_nkb}
            rest_f = {k: v for k, v in raw_folders.items() if _nar_kb(k) != cls_ar_nkb}
            raw_folders = {**own_f, **rest_f}
    all_folders = list(raw_folders.items())
    PAGE_SIZE   = 20
    f_start     = page * PAGE_SIZE
    f_end       = f_start + PAGE_SIZE
    page_folders = all_folders[f_start:f_end]

    for name, sub in page_folders:
        cnt   = folder_item_count(sub)
        label = f"📁 {name}" + (f" ({cnt})" if cnt else "")
        kb.append([InlineKeyboardButton(label, callback_data=f"open|{name}")])

    # Dosyalar
    files = folder.get("files", [])
    for idx, f in enumerate(files[:12]):
        cap = f.get("caption", f.get("name", "?"))
        kb.append([InlineKeyboardButton(f"📎 {cap}", callback_data=f"getfile|{idx}")])

    # Sayfa navigasyonu
    total_pages = max(1, -(-len(all_folders) // PAGE_SIZE))
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("◀️", callback_data=f"page|{page-1}"))
    if total_pages > 1:
        nav.append(InlineKeyboardButton(f"{page+1}/{total_pages}", callback_data="noop"))
    if f_end < len(all_folders):
        nav.append(InlineKeyboardButton("▶️", callback_data=f"page|{page+1}"))
    if nav:
        kb.append(nav)

    if path:
        kb.append([InlineKeyboardButton(L(uid, "back"), callback_data="nav|back")])
    return InlineKeyboardMarkup(kb)

async def show_folder(query, context, path, note=""):
    content = load_content()
    folder  = get_folder(content, path)
    uid     = str(query.from_user.id)
    page    = context.user_data.get("page", 0)
    text    = folder_text(folder, path, uid)
    if note: text += f"\n\n{note}"
    # Klasör gezintisini logla
    if not is_admin(uid) and path:
        folder_name = " / ".join(path)
        log_user_message(query.from_user, "folder", folder_name)
    try:
        await query.edit_message_text(text, reply_markup=folder_kb(path, folder, uid, page))
    except Exception:
        pass

async def show_folder_new(message, uid, path=None, page=0):
    path    = path or []
    content = load_content()
    folder  = get_folder(content, path)
    uid     = str(uid)
    await message.reply_text(
        folder_text(folder, path, uid),
        reply_markup=folder_kb(path, folder, uid, page))

async def _delete_last_inline(context, chat_id):
    msg_id = context.user_data.get("last_inline_msg")
    if msg_id:
        try: await context.bot.delete_message(chat_id=chat_id, message_id=msg_id)
        except: pass
        context.user_data.pop("last_inline_msg", None)

# ================================================================
#  DOSYA GÖNDERME
# ================================================================

async def _send_file(message, f: dict, uid: str):
    await _send_file_get_msg(message, f, uid)

async def _send_file_get_msg(message, f: dict, uid: str):
    """Dosyayı gönderir ve gönderilen mesajı döndürür."""
    ftype = f.get("type",""); fid = f.get("file_id","")
    cap   = f.get("caption",""); url = f.get("url","")
    sent  = None
    try:
        if   ftype == "photo":     sent = await message.reply_photo(fid, caption=cap)
        elif ftype == "video":     sent = await message.reply_video(fid, caption=cap)
        elif ftype == "animation": sent = await message.reply_animation(fid, caption=cap)
        elif ftype == "document":  sent = await message.reply_document(fid, caption=cap)
        elif ftype == "audio":     sent = await message.reply_audio(fid, caption=cap)
        elif ftype == "voice":     sent = await message.reply_voice(fid, caption=cap)
        elif ftype == "link":      sent = await message.reply_text(f"🔗 {cap}\n{url}" if cap != url else f"🔗 {url}")
        elif ftype == "text":      sent = await message.reply_text(cap)
        else: sent = await message.reply_text(L(uid, "unsupported"))
        increment_view(f)
    except Exception as e:
        await message.reply_text(L(uid, "send_fail").format(e))
    return sent

# ================================================================
#  /start   /menu
# ================================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user; uid = str(user.id)

    is_new = register_user(user)
    log_user_message(user, "command", "/start")

    if is_blocked(uid) and not is_admin(uid): return

    s = load_settings()
    context.user_data["path"] = []
    context.user_data["mode"] = "browse"
    context.user_data.pop("action", None)

    # /start mesajını sil
    try: await update.message.delete()
    except: pass

    # Bakım modu kontrolü (kullanıcılar için)
    if not is_admin(uid) and s["maintenance"]:
        await update.message.chat.send_message(
            s.get("maintenance_text", "🔧"),
            reply_markup=reply_kb(uid))
        return

    # Klavyeyi gönder
    await update.message.chat.send_message("👋", reply_markup=reply_kb(uid))

    # Yeni kullanıcı bildirimi (admin'e)
    if is_new and not is_admin(uid):
        u  = load_users().get(uid, {})
        un = f" @{u['username']}" if u.get("username") else ""
        try:
            await context.bot.send_message(
                ADMIN_ID,
                TR["new_user_notif"].format(
                    name=u.get("full_name") or u.get("first_name") or f"ID:{uid}",
                    un=un, uid=uid,
                    time=datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M")),
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton(TR["msg_btn"], callback_data=f"dm_quick|{uid}")
                ]])
            )
        except Exception as e:
            logger.warning(f"Yeni kullanıcı bildirimi gönderilemedi: {e}")

    if is_admin(uid):
        # Admin — normal klasör göster
        await show_folder_new(update.message, uid)
        return

    # ── KULLANICI AKIŞI ──────────────────────────────────────
    cls = get_user_class(uid)

    if not cls:
        # Sınıf seçmemiş — karşılama + sınıf seçimi göster
        if s.get("bot_photo_id"):
            try:
                await update.message.chat.send_photo(
                    s["bot_photo_id"],
                    caption=s.get("welcome_msg") or "")
            except: pass
        elif s.get("welcome_msg"):
            # Karşılama mesajını sabit tut (silinmez)
            await update.message.chat.send_message(s["welcome_msg"])

        kb = [
            [InlineKeyboardButton(CLASS_DEFS["1"]["ar"], callback_data="class_pick|1"),
             InlineKeyboardButton(CLASS_DEFS["2"]["ar"], callback_data="class_pick|2")],
            [InlineKeyboardButton(CLASS_DEFS["3"]["ar"], callback_data="class_pick|3"),
             InlineKeyboardButton(CLASS_DEFS["4"]["ar"], callback_data="class_pick|4")],
        ]
        await update.message.chat.send_message(
            L(uid, "class_select_prompt"),
            reply_markup=InlineKeyboardMarkup(kb))
    else:
        # Sınıf seçmiş — ana sayfayı göster (tüm sınıflar, kendi sınıfı önde)
        context.user_data["path"] = []
        await show_folder_new(update.message, uid)

async def menu_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user; uid = str(user.id)
    register_user(user)
    context.user_data["mode"]  = "browse"
    context.user_data.pop("action", None)

    if is_admin(uid):
        context.user_data["path"] = []
        await show_folder_new(update.message, uid)
        return

    # Kullanıcılar — kendi sınıf klasörüne git
    # Kullanıcılar — ana sayfa (tüm sınıflar, kendi sınıfı önde)
    context.user_data["path"] = []
    await show_folder_new(update.message, uid)
# ================================================================
#  REPLY KLAVYE HANDLER
# ================================================================

async def handle_reply_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user; uid = str(user.id)
    text = (update.message.text or "").strip()
    register_user(user)
    await _delete_last_inline(context, update.effective_chat.id)
    try: await update.message.delete()
    except: pass

    # Engellenmiş kullanıcı — hiçbir şey yapamaz
    if is_blocked(uid) and not is_admin(uid):
        await update.message.reply_text("🚫 تم حظرك من استخدام البوت.")
        return

    # Kullanıcı/alt-admin buton seçimini kaydet
    if not is_main_admin(uid) and not is_blocked(uid):
        log_user_message(user, "button", text)

    # ── Arama (herkes) ───────────────────────────────
    if text in (TR["btn_search"], AR["btn_search"]):
        context.user_data["action"] = "search"
        await update.message.reply_text(L(uid, "search_prompt"))
        return


    # ── Profil (kullanıcılar) ────────────────────────
    if text in (TR["profile_btn"], AR["profile_btn"]):
        if not is_admin(uid):
            u    = load_users().get(uid, {})
            name = u.get("full_name") or u.get("first_name") or "—"
            cls  = class_label(uid)
            favs = len(get_favorites(uid))
            rcnt = len(get_recently_viewed(uid))
            warns_list = get_warns(uid)
            grp   = get_user_group(uid)
            shift = get_user_shift(uid)
            shift_lbl = "صباحي" if shift in ("sabahi","sabah") else ("مسائي" if shift in ("masaiy","gece") else "—")
            txt = (
                f"👤 *{name}*\n"
                f"🆔 {uid}\n"
                f"🎓 {L(uid,'class_label').format(cls)}\n"
                f"⏰ الفترة: {shift_lbl}\n"
                f"👥 المجموعة: {grp or '—'}\n"
                f"⭐ {L(uid,'fav_list')}: {favs}\n"
                f"🕐 {L(uid,'recent_list')}: {rcnt}\n"
                f"⚠️ {L(uid,'warn_count_label').format(len(warns_list), MAX_WARNS)}"
            )
            kb = [
                [InlineKeyboardButton(L(uid,"class_change_btn"), callback_data="class_change")],
                [InlineKeyboardButton(L(uid,"group_change_btn"), callback_data="group_pick_start")],
            ]
            await update.message.reply_text(txt, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(kb))
        return

    # ── Favoriler ────────────────────────────────────
    if text in (TR["btn_favorites"], AR["btn_favorites"]):
        if not is_admin(uid):
            favs = get_favorites(uid)
            if not favs:
                await update.message.reply_text(L(uid,"fav_empty")); return
            kb = []
            for idx, f in enumerate(favs[:15]):
                cap = f.get("caption", f.get("name","?"))[:35]
                kb.append([InlineKeyboardButton(f"⭐ {cap}", callback_data=f"fav|open|{idx}")])
            kb.append([InlineKeyboardButton(L(uid,"back"),  callback_data="nav|root")])
            await update.message.reply_text(L(uid,"fav_list"), reply_markup=InlineKeyboardMarkup(kb))
        return

    # ── Son Görüntülenenler ──────────────────────────
    if text in (TR["btn_recent"], AR["btn_recent"]):
        if not is_admin(uid):
            recent = get_recently_viewed(uid)
            if not recent:
                await update.message.reply_text(L(uid,"recent_empty")); return
            kb = []
            for entry in reversed(recent[-10:]):
                f   = entry["file"]
                cap = f.get("caption", f.get("name","?"))[:35]
                t   = entry.get("time","")[-5:]
                key = f.get("file_id") or f.get("caption","?")
                kb.append([InlineKeyboardButton(f"🕐 {t} — {cap}", callback_data=f"recent|open|{key}")])
            kb.append([InlineKeyboardButton(L(uid,"back"),  callback_data="nav|root")])
            await update.message.reply_text(L(uid,"recent_list"), reply_markup=InlineKeyboardMarkup(kb))
        return

    # ── Liderboard (herkes) ──────────────────────────
    if text in (TR["btn_leaderboard"], AR["btn_leaderboard"]):
        lb    = get_leaderboard(10)
        lines = [L(uid,"leaderboard_title"), ""]
        medals = ["🥇","🥈","🥉","4️⃣","5️⃣","6️⃣","7️⃣","8️⃣","9️⃣","🔟"]
        for i,(name,pts) in enumerate(lb):
            medal = medals[i] if i < len(medals) else f"{i+1}."
            lines.append(f"{medal} {name[:25]}  —  {pts} pts")
        if not lb: lines.append(L(uid,"leaderboard_empty"))
        own_pts = load_leaderboard().get(uid, 0)
        lines.append(f"\n\n👤 {L(uid,'profile_title')}: {own_pts} pts")
        await update.message.reply_text("\n".join(lines))
        return

    # ── Notlarım ─────────────────────────────────────
    if text in (TR["btn_notes"], AR["btn_notes"]):
        if not is_admin(uid):
            notes = get_user_notes(uid)
            kb = []
            # Son 15 not — her not buton olarak
            for n in notes[-15:]:
                icon = {"text":"✍️","photo":"🖼","video":"🎥","voice":"🎙","audio":"🎵","document":"📄"}.get(n.get("type","text"),"📌")
                subj = n.get("subject","") or ""
                cap  = n.get("content","")[:25] or f"[{n.get('type','medya')}]"
                lbl  = f"{icon} {subj+': ' if subj else ''}{cap}"
                kb.append([
                    InlineKeyboardButton(lbl[:45], callback_data=f"notes|view|{n['id']}"),
                    InlineKeyboardButton("🗑", callback_data=f"notes|del|{n['id']}"),
                ])
            kb.append([InlineKeyboardButton("➕ إضافة ملاحظة", callback_data="notes|new")])
            kb.append([InlineKeyboardButton(L(uid,"back"), callback_data="nav|root")])
            cnt_txt = f"({len(notes)} ملاحظة)" if notes else "لا توجد ملاحظات بعد"
            sent = await update.message.reply_text(
                f"📝 ملاحظاتي\n\n{cnt_txt}",
                reply_markup=InlineKeyboardMarkup(kb))
            context.user_data["last_inline_msg"] = sent.message_id
        return

    # ── Hatırlatıcım ─────────────────────────────────
    if text in (TR["btn_reminder"], AR["btn_reminder"]):
        if not is_admin(uid):
            rems = get_user_reminders(uid)
            if not rems:
                kb = [
                    [InlineKeyboardButton("➕ " + ("Ekle" if is_main_admin(uid) else "إضافة"),
                                          callback_data="reminder|add")],
                    [InlineKeyboardButton(L(uid,"back"), callback_data="nav|root")],
                ]
                await update.message.reply_text(L(uid,"reminder_none"), reply_markup=InlineKeyboardMarkup(kb))
            else:
                import time
                now = time.time()
                lines = [L(uid,"reminder_list").format(len(rems))]
                kb = []
                for i, r in enumerate(rems[:10]):
                    left = int((r["fire_ts"] - now) / 60)
                    left_str = f"{left}dk" if left < 60 else f"{left//60}sa"
                    lines.append(f"\n🔔 {r['text'][:40]}\n   ⏰ {left_str} kaldı")
                    kb.append([InlineKeyboardButton(f"🗑 {r['text'][:25]}", callback_data=f"reminder|del|{i}")])
                kb.append([InlineKeyboardButton("➕ Yeni", callback_data="reminder|add")])
                kb.append([InlineKeyboardButton(L(uid,"back"), callback_data="nav|root")])
                sent = await update.message.reply_text(
                    "\n".join(lines), reply_markup=InlineKeyboardMarkup(kb))
                context.user_data["last_inline_msg"] = sent.message_id
        return

    # ── Anonim Soru ──────────────────────────────────

    # ── Geri Sayım / Yaklaşan Sınavlar ──────────────
    if text in (TR.get("btn_anon","✉️ Anonim Mesaj"), AR.get("btn_anon","✉️ رسالة مجهولة")) and not is_admin(uid):
        s = load_settings()
        if not s.get("anon_group_id"):
            await update.message.reply_text(L(uid,"anon_disabled"))
            return
        context.user_data["action"] = "anon_msg"
        kb = [[InlineKeyboardButton("◀️ إلغاء", callback_data="nav|root")]]
        sent = await update.message.reply_text(
            "✉️ اكتب رسالتك المجهولة:\n(ستصل إلى المجموعة بدون اسمك)",
            reply_markup=InlineKeyboardMarkup(kb))
        context.user_data["last_inline_msg"] = sent.message_id
        return WAIT_FOLDER

    if text == "🔬 المختبر" and not is_admin(uid):
        from datetime import datetime as _dt
        ARABIC_DAYS = {0:"الاثنين",1:"الثلاثاء",2:"الأربعاء",3:"الخميس",4:"الجمعة",5:"السبت",6:"الأحد"}
        user_grp = load_groups().get(uid, "")
        schedule = load_lab_schedule()
        today_ts = _dt.now().date()

        my_labs = []
        for e in schedule:
            try:
                e_dt = _dt.strptime(e["week"], "%Y-%m-%d").date()
                if e_dt >= today_ts and e.get("group","") == user_grp:
                    day_name = ARABIC_DAYS.get(e_dt.weekday(), "")
                    date_disp = f"{day_name} {e_dt.strftime('%d/%m/%Y')}"
                    note = e.get("note","")
                    my_labs.append((date_disp, note))
            except: pass

        if not my_labs:
            await update.message.reply_text(
                f"🔬 المختبر\n\nلا توجد مواعيد مختبر لمجموعتك ({user_grp or '؟'}) قريباً.")
            return

        lines = [f"🔬 مواعيد مختبر مجموعة {user_grp}\n"]
        for date_disp, note in my_labs[:8]:
            nt = f"\n     📝 {note}" if note else ""
            lines.append(f"  📅 {date_disp}{nt}")

        await update.message.reply_text("\n".join(lines))
        return

    if text in (TR["countdown_btn"], AR["countdown_btn"]):
        cls   = get_user_class(uid) if not is_admin(uid) else ""
        shift = get_user_shift(uid)  if not is_admin(uid) else ""
        cds   = get_countdowns(cls, shift, uid=uid)
        if not cds:
            await update.message.reply_text(L(uid,"countdown_none"))
        else:
            lines = [f"⏳ {'Yaklaşan Sınavlar' if is_main_admin(uid) else 'الامتحانات القادمة'}",
                     ""]
            remind_kb = []
            for idx_c, cd in enumerate(cds[:8]):
                days = cd["days_left"]
                if days == 0:
                    when = "🔴 اليوم!" if not is_main_admin(uid) else "🔴 Bugün!"
                elif days == 1:
                    when = "🟠 غداً" if not is_main_admin(uid) else "🟠 Yarın"
                elif days <= 7:
                    when = f"🟡 {days} {'أيام' if not is_main_admin(uid) else 'gün'}"
                else:
                    when = f"🟢 {days} {'يوم' if not is_main_admin(uid) else 'gün'}"
                lines.append(f"\n📅 {cd['name']}\n   {when} — {cd['date']}")
                if is_admin(uid):
                    remind_kb.append([InlineKeyboardButton(
                        f"📢 تذكير: {cd['name'][:20]}",
                        callback_data=f"cnt|remind|{idx_c}")])
            kb_final = remind_kb if remind_kb else None
            await update.message.reply_text(
                "\n".join(lines),
                reply_markup=InlineKeyboardMarkup(remind_kb) if remind_kb else None)
        return

    # ── Mini Test (Quiz) ─────────────────────────────
    if text in (TR["quiz_btn"], AR["quiz_btn"]):
        if not is_admin(uid):
            quizzes = [q for q in load_quizzes() if q.get("active")]
            cls = get_user_class(uid)
            quizzes = [q for q in quizzes if not q.get("cls") or q.get("cls") == cls]
            if not quizzes:
                await update.message.reply_text(L(uid,"quiz_none"))
            else:
                q = quizzes[0]
                opts = q.get("options", [])
                kb = [[InlineKeyboardButton(f"{'ABCD'[i]}. {o}", callback_data=f"quiz|ans|{q['id']}|{i}")]
                      for i, o in enumerate(opts)]
                kb.append([InlineKeyboardButton(L(uid,"back"), callback_data="nav|root")])
                await update.message.reply_text(
                    f"📝 {q.get('question','?')}\n",
                    reply_markup=InlineKeyboardMarkup(kb))
        return

    # ── Mesaj Gönder / Yardım (herkes) ──────────────
    if text in (TR["btn_help"], AR["btn_help"]):
        if is_admin(uid) and not is_main_admin(uid):
            # Alt admin → süper admin'e mesaj gönder
            u_adm = load_users().get(uid, {})
            name_adm = u_adm.get("full_name") or u_adm.get("first_name") or f"ID:{uid}"
            adm_cls  = get_admin_cls(uid) or "الكل"
            adm_grp  = get_admin_grp(uid) or "الكل"
            context.user_data["action"] = "admin_msg_to_super"
            context.user_data["admin_identity"] = f"👮 مسؤول: {name_adm}\n🎓 مسؤول عن: {adm_cls} / {adm_grp}"
            kb = [[InlineKeyboardButton("◀️ إلغاء", callback_data="nav|root")]]
            await update.message.reply_text(
                "📨 اكتب رسالتك للمسؤول الرئيسي:",
                reply_markup=InlineKeyboardMarkup(kb))
            return
        if is_main_admin(uid):
            await update.message.reply_text(L(uid, "help_text"), parse_mode="Markdown")
        else:
            cls_id = get_user_class(uid)
            # Sınıf admini var mı?
            cls_admin_exists = False
            if cls_id:
                for adm_id_str, adm_p in load_admin_perms().items():
                    if adm_p.get("cls") == cls_id:
                        cls_admin_exists = True
                        break
            # Her zaman seçim göster — sınıf admini yoksa o butona bastıkta bilgi ver
            kb = [
                [InlineKeyboardButton(
                    "👤 مسؤول صفي" if cls_admin_exists else "👤 مسؤول صفي (غير متاح)",
                    callback_data="msg_target|class_admin")],
                [InlineKeyboardButton(
                    "🔑 المسؤول الرئيسي",
                    callback_data="msg_target|main_admin")],
                [InlineKeyboardButton(L(uid,"back"), callback_data="nav|root")],
            ]
            sent = await update.message.reply_text(
                "📨 إلى من تريد إرسال رسالتك؟",
                reply_markup=InlineKeyboardMarkup(kb))
            context.user_data["last_inline_msg"] = sent.message_id
            context.user_data["cls_admin_exists"] = cls_admin_exists
        return

    # ── Sadece adminler buradan devam ────────────────
    # ── İçerik ──────────────────────────────────────
    if text in (TR["btn_content"], AR["btn_content"]):
        if is_main_admin(uid):
            kb = [
                [InlineKeyboardButton(L(uid,"add_folder"),    callback_data="cnt|add_folder"),
                 InlineKeyboardButton(L(uid,"del_folder"),    callback_data="cnt|del_folder")],
                [InlineKeyboardButton(L(uid,"rename_folder"), callback_data="cnt|rename_folder"),
                 InlineKeyboardButton(L(uid,"add_file"),      callback_data="cnt|add_file")],
                [InlineKeyboardButton(L(uid,"del_file"),      callback_data="cnt|del_file"),
                 InlineKeyboardButton(L(uid,"rename_file"),   callback_data="cnt|rename_file")],
                [InlineKeyboardButton(L(uid,"pin_file"),      callback_data="extra|pin"),
                 InlineKeyboardButton(L(uid,"move_file"),     callback_data="extra|move")],
                [InlineKeyboardButton(L(uid,"copy_file"),     callback_data="extra|copy"),
                 InlineKeyboardButton(L(uid,"sort_az"),       callback_data="extra|sort_az")],
                [InlineKeyboardButton(L(uid,"sort_views"),    callback_data="extra|sort_views"),
                 InlineKeyboardButton(L(uid,"folder_desc_btn"),callback_data="extra|folder_desc")],
                [InlineKeyboardButton("📅 Sinav Ekle",       callback_data="cnt|add_countdown")],
                [InlineKeyboardButton(L(uid,"back"),          callback_data="nav|root")],
            ]
            sent = await update.message.reply_text(L(uid,"content_mgmt"), reply_markup=InlineKeyboardMarkup(kb))
            context.user_data["last_inline_msg"] = sent.message_id
        elif is_admin(uid):
            kb = []
            row_add = []
            if get_admin_perm(uid, "can_add_folder"):
                row_add.append(InlineKeyboardButton("📁 إضافة مجلد", callback_data="cnt|add_folder"))
            if get_admin_perm(uid, "can_del_folder"):
                row_add.append(InlineKeyboardButton("🗑 حذف مجلد", callback_data="cnt|del_folder"))
            if row_add: kb.append(row_add)

            row_file = []
            if get_admin_perm(uid, "can_add_file"):
                row_file.append(InlineKeyboardButton("📎 إضافة ملف", callback_data="cnt|add_file"))
            if get_admin_perm(uid, "can_del_file"):
                row_file.append(InlineKeyboardButton("🗑 حذف ملف", callback_data="cnt|del_file"))
            if get_admin_perm(uid, "can_rename_file"):
                row_file.append(InlineKeyboardButton("✏️ إعادة تسمية", callback_data="cnt|rename_file"))
            if row_file: kb.append(row_file)

            if get_admin_perm(uid, "can_countdown"):
                kb.append([InlineKeyboardButton("📅 إضافة امتحان", callback_data="cnt|add_countdown")])
                kb.append([InlineKeyboardButton("🔬 إضافة موعد مختبر", callback_data="lab|add_new")])
            if get_admin_perm(uid, "can_poll"):
                kb.append([InlineKeyboardButton("📊 إنشاء استطلاع", callback_data="poll|create")])
            if get_admin_perm(uid, "can_broadcast"):
                kb.append([InlineKeyboardButton("📢 إرسال إعلان", callback_data="admin_bcast|panel")])
            if not kb:
                await update.message.reply_text("لا توجد صلاحيات محددة لك حتى الآن.")
                return
            if get_admin_perm(uid, "can_view_users"):
                kb.append([InlineKeyboardButton("👥 طلاب صفي", callback_data="mgmt|my_users")])
            kb.append([InlineKeyboardButton(L(uid,"back"), callback_data="nav|root")])
            sent = await update.message.reply_text("📋 لوحة المحتوى:", reply_markup=InlineKeyboardMarkup(kb))
            context.user_data["last_inline_msg"] = sent.message_id
        else:
            # Normal kullanıcı — klasörlere git
            context.user_data["path"] = []
            await show_folder_new(update.message, uid)
        return

    # ── Sadece adminler buradan devam ────────────────
    if not is_admin(uid): return

    # ── Bakım (tüm adminler) ─────────────────────────
    if text in (TR["btn_maint"], AR["btn_maint"]):
        s = load_settings(); s["maintenance"] = not s["maintenance"]; save_settings(s)
        durum = L(uid,"maint_on_str") if s["maintenance"] else L(uid,"maint_off_str")
        await update.message.reply_text(L(uid,"maint_toggle").format(durum))
        return

    # ── Sadece süper admin ───────────────────────────
    if not is_main_admin(uid): return

    if text == TR["btn_mgmt"]:
        kb = [
            [InlineKeyboardButton(TR["stats"],            callback_data="mgmt|stats"),
             InlineKeyboardButton(TR["users"],            callback_data="mgmt|users")],
            [InlineKeyboardButton("🔬 Lab Programi",     callback_data="mgmt|lab")],
            [InlineKeyboardButton("📋 Admin Log İndir",   callback_data="mgmt|admin_log_export")],
            [InlineKeyboardButton("⏰ Bildirim Ayarları",    callback_data="set|remind_cfg")],
            [InlineKeyboardButton("✉️ Anonim Mesaj Grubu",   callback_data="set|anon_group")],
            [InlineKeyboardButton(TR["add_admin"],        callback_data="mgmt|add_admin"),
             InlineKeyboardButton(TR["del_admin"],        callback_data="mgmt|del_admin")],
            [InlineKeyboardButton(TR["dm_user"],          callback_data="mgmt|dm_user"),
             InlineKeyboardButton(TR["broadcast"],        callback_data="mgmt|broadcast")],
            [InlineKeyboardButton(TR["bcast_targeted"],   callback_data="bcast|panel"),
             InlineKeyboardButton("🎓 Sınıf İstat.",     callback_data="mgmt|class_stats")],
            [InlineKeyboardButton(TR["poll_btn"],         callback_data="mgmt|poll"),
             InlineKeyboardButton(TR["admin_log_btn"],    callback_data="admin|log")],
            [InlineKeyboardButton(TR["backup_btn"],       callback_data="backup|do"),
             InlineKeyboardButton(TR["export_users_btn"], callback_data="export|users")],
            [InlineKeyboardButton(TR["bcast_history_btn"],callback_data="bcast|history"),
             InlineKeyboardButton("🏆 Liderboard",       callback_data="misc|leaderboard")],
            [InlineKeyboardButton("⏳ Sınav Ekle",       callback_data="countdown|add"),
],
            [InlineKeyboardButton("📝 Quiz Oluştur",     callback_data="admin|quiz_create"),
             InlineKeyboardButton("📊 Sınıf Analizi",   callback_data="admin|class_analysis")],
            [InlineKeyboardButton("◀️ Geri",              callback_data="nav|root")],
        ]
        sent = await update.message.reply_text(TR["mgmt_panel"], reply_markup=InlineKeyboardMarkup(kb))
        context.user_data["last_inline_msg"] = sent.message_id
        return

    if text == TR["btn_settings"]:
        s   = load_settings()
        txt = (
            f"{TR['settings_title']}\n\n"
            f"📝 Ad: {s.get('bot_name','—')}\n"
            f"💬 Karşılama: {str(s.get('welcome_msg','—'))[:60]}...\n"
            f"📄 Açıklama: {str(s.get('bot_description','—'))[:40]}...\n"
            f"🖼 Fotoğraf: {'✅' if s.get('bot_photo_id') else '❌'}\n"
            f"🔧 Bakım metni: {str(s.get('maintenance_text','—'))[:40]}\n"
            f"🚫 Engellenenler: {len(load_blocked())} kişi"
        )
        kb = [
            [InlineKeyboardButton("📝 Bot Adı",          callback_data="set|name"),
             InlineKeyboardButton("💬 Karşılama",         callback_data="set|welcome")],
            [InlineKeyboardButton("🖼 Bot Fotoğrafı",    callback_data="set|photo"),
             InlineKeyboardButton("📄 Bot Açıklaması",   callback_data="set|set_description")],
            [InlineKeyboardButton("🔧 Bakım Metni",      callback_data="set|maint_text"),
             InlineKeyboardButton(TR["set_blocked_btn"], callback_data="set|blocked")],
            [InlineKeyboardButton(TR["btn_mgmt_btn"],    callback_data="set|btn_mgmt"),
             InlineKeyboardButton("🎓 Sınıf İsimleri",   callback_data="set|class_names")],
            [InlineKeyboardButton("👥 Grup Yapılandırma", callback_data="set|class_groups"),
             InlineKeyboardButton("👮 Admin Yetkileri",   callback_data="set|admin_perms")],
            [InlineKeyboardButton("⏰ Bildirim Ayarları", callback_data="set|remind_cfg")],
            [InlineKeyboardButton("❓ FAQ Yönetimi",     callback_data="faqmgmt|panel"),
             InlineKeyboardButton("⚡ Otomatik Kurallar",callback_data="rulemgmt|panel")],
            [InlineKeyboardButton(TR["pin_msg_btn"],      callback_data="set|pin_msg"),
             InlineKeyboardButton(TR["bcast_history_btn"],callback_data="bcast|history")],
            [InlineKeyboardButton("◀️ Geri",             callback_data="nav|root")],
        ]
        sent = await update.message.reply_text(txt, reply_markup=InlineKeyboardMarkup(kb))
        context.user_data["last_inline_msg"] = sent.message_id
        return

# ================================================================
#  CALLBACK HANDLER
# ================================================================

async def callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    user = query.from_user; uid = str(user.id); adm = is_admin(uid); cb = query.data

    if is_blocked(uid) and not adm: return ConversationHandler.END
    if not adm:
        s = load_settings()
        if s["maintenance"]:
            await query.answer(s.get("maintenance_text","🔧"), show_alert=True)
            return ConversationHandler.END

    content = load_content()
    path    = context.user_data.get("path", [])

    # ── Sınıf Seçimi ─────────────────────────────────
    if cb.startswith("class_pick|"):
        cls = cb.split("|")[1]
        if cls in CLASS_DEFS:
            set_user_class(uid, cls)
            cls_ar    = CLASS_DEFS[cls]["ar"]
            cls_tr    = CLASS_DEFS[cls]["tr"]
            cls_name  = cls_ar if not is_main_admin(uid) else cls_tr
            # Sınıfa ait klasörü bul ve oraya git
            content2  = load_content()
            def _nar3(s): return s.replace("أ","ا").replace("إ","ا").replace("آ","ا")
            cls_ar_n3 = _nar3(cls_ar)
            cls_folder = None
            for fname in content2.get("folders", {}):
                if _nar3(fname) == cls_ar_n3:
                    cls_folder = fname
                    break
            confirm_text = L(uid, "class_selected").format(cls_name)
            # Sınıf seçildikten sonra grup sor
            kb_shift = [
                [InlineKeyboardButton("صباحي (نهار)", callback_data="shift_pick|sabahi"),
                 InlineKeyboardButton("مسائي (مساء)", callback_data="shift_pick|masaiy")],
            ]
            await query.edit_message_text(
                confirm_text + "\n\nVaridiyanizi secin:",
                reply_markup=InlineKeyboardMarkup(kb_shift))
            if cls_folder:
                context.user_data["path"] = [cls_folder]
        return ConversationHandler.END

    if cb.startswith("msg_target|") and not is_admin(uid):
        target_type = cb.split("|")[1]
        # Sınıf admini yoksa bilgi ver
        if target_type == "class_admin" and not context.user_data.get("cls_admin_exists", True):
            await query.answer("لا يوجد مسؤول لصفك حالياً، رسالتك ستصل للمسؤول الرئيسي.", show_alert=True)
            target_type = "main_admin"
        context.user_data["msg_to"] = target_type
        context.user_data["action"] = "user_msg_to_admin"
        prompt = "💬 اكتب رسالتك لمسؤول صفك:" if target_type == "class_admin" else "💬 اكتب رسالتك للمسؤول الرئيسي:"
        await query.edit_message_text(
            prompt,
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ إلغاء", callback_data="nav|root")
            ]]))
        return WAIT_FOLDER

    if cb == "group_pick_start" and not is_admin(uid):
        kb = [
            [InlineKeyboardButton("صباحي (نهار)", callback_data="shift_pick|sabahi"),
             InlineKeyboardButton("مسائي (مساء)", callback_data="shift_pick|masaiy")],
            [InlineKeyboardButton(L(uid,"back"), callback_data="nav|root")],
        ]
        await query.edit_message_text(
            "Vardiyanizi secin:" if is_main_admin(uid) else "اختر الفترة الدراسية:",
            reply_markup=InlineKeyboardMarkup(kb))
        return ConversationHandler.END

    if cb.startswith("shift_pick|") and not is_admin(uid):
        shift = cb.split("|")[1]
        set_user_shift(uid, shift)
        shift_lbl = "صباحي" if shift in ("sabahi","sabahi") else "مسائي"
        kb_grp = [
            [InlineKeyboardButton("A", callback_data="group_pick|A"),
             InlineKeyboardButton("B", callback_data="group_pick|B"),
             InlineKeyboardButton("C", callback_data="group_pick|C")],
        ]
        await query.edit_message_text(
            f"{shift_lbl} - " + L(uid,"group_select"),
            reply_markup=InlineKeyboardMarkup(kb_grp))
        return ConversationHandler.END

    if cb.startswith("group_pick|") and not is_admin(uid):
        group_main = cb.split("|")[1]  # A, B, or C
        # Alt grup seç
        kb_sub = [
            [InlineKeyboardButton(f"{group_main}1", callback_data=f"group_sub|{group_main}1"),
             InlineKeyboardButton(f"{group_main}2", callback_data=f"group_sub|{group_main}2"),
             InlineKeyboardButton(f"{group_main}3", callback_data=f"group_sub|{group_main}3")],
        ]
        await query.edit_message_text(
            L(uid,"group_select") + f"\n\n→ {group_main}",
            reply_markup=InlineKeyboardMarkup(kb_sub))
        return ConversationHandler.END

    if cb.startswith("group_sub|") and not is_admin(uid):
        group_full = cb.split("|")[1]   # A1, A2 ... C3
        set_user_group(uid, group_full)
        cls = get_user_class(uid)
        cls_ar = CLASS_DEFS.get(cls, {}).get("ar", "") if cls else ""
        # İlgili klasöre git
        def _nar_g(s): return s.replace("أ","ا").replace("إ","ا").replace("آ","ا")
        cls_ar_ng = _nar_g(cls_ar)
        content2 = load_content()
        cls_folder = next((f for f in content2.get("folders",{}) if _nar_g(f)==cls_ar_ng), None)
        path_new = [cls_folder] if cls_folder else []
        context.user_data["path"] = path_new
        await query.edit_message_text(L(uid,"group_selected").format(group_full))
        if path_new:
            await show_folder_new(query.message, uid, path=path_new)
        else:
            await show_folder_new(query.message, uid)
        return ConversationHandler.END

    if cb == "class_change":
        kb = [
            [InlineKeyboardButton(CLASS_DEFS["1"]["ar"], callback_data="class_pick|1"),
             InlineKeyboardButton(CLASS_DEFS["2"]["ar"], callback_data="class_pick|2")],
            [InlineKeyboardButton(CLASS_DEFS["3"]["ar"], callback_data="class_pick|3"),
             InlineKeyboardButton(CLASS_DEFS["4"]["ar"], callback_data="class_pick|4")],
            [InlineKeyboardButton(L(uid,"cancel"), callback_data="close")],
        ]
        await query.edit_message_text(L(uid,"class_change_prompt"), reply_markup=InlineKeyboardMarkup(kb))
        return ConversationHandler.END

    # ── Favoriler ─────────────────────────────────────
    if cb.startswith("fav|"):
        action = cb.split("|")[1]
        if action == "open":
            idx  = int(cb.split("|")[2])
            favs = get_favorites(uid)
            if idx < len(favs):
                f = favs[idx]
                await _send_file(query.message, f, uid)
                log_user_message(user, "file_view", f.get("caption",""))
        elif action == "toggle":
            key  = cb.split("|")[2]
            # key = file_id or caption, find in recently viewed or content
            recent = get_recently_viewed(uid)
            for entry in recent:
                f = entry["file"]
                fk = f.get("file_id") or f.get("caption","?")
                if fk == key:
                    added = toggle_favorite(uid, f)
                    await query.answer(L(uid,"fav_added") if added else L(uid,"fav_removed"), show_alert=True)
                    return ConversationHandler.END
            await query.answer("❓", show_alert=True)
        return ConversationHandler.END

    # ── Son Görüntülenenler ───────────────────────────
    if cb.startswith("recent|"):
        action = cb.split("|")[1]
        if action == "open":
            key    = cb.split("|",2)[2]
            recent = get_recently_viewed(uid)
            for entry in recent:
                f = entry["file"]
                fk = f.get("file_id") or f.get("caption","?")
                if fk == key:
                    await _send_file(query.message, f, uid)
                    log_user_message(user, "file_view", f.get("caption",""))
                    return ConversationHandler.END
            await query.answer(L(uid,"file_notfound"), show_alert=True)
        return ConversationHandler.END

    # ── Klasör Aboneliği ──────────────────────────────
    if cb == "sub|toggle":
        added = toggle_sub(uid, path)
        await query.answer(L(uid,"subscribed_ok") if added else L(uid,"unsubscribed_ok"), show_alert=True)
        await show_folder(query, context, path)
        return ConversationHandler.END

    # ── İç Yapay Zeka / FAQ ───────────────────────────
    if cb.startswith("faqmgmt|") and is_main_admin(uid):
        action = cb.split("|")[1]
        if action == "panel":
            faqs = load_faq()
            kb = [
                [InlineKeyboardButton(TR["faq_add_btn"],  callback_data="faqmgmt|add"),
                 InlineKeyboardButton(TR["faq_list_btn"], callback_data="faqmgmt|list")],
                [InlineKeyboardButton(TR["faq_del_btn"],  callback_data="faqmgmt|del")],
                [InlineKeyboardButton("◀️ Geri",          callback_data="nav|root")],
            ]
            await query.edit_message_text(
                f"{TR['faq_panel']}\n\n📋 Toplam: {len(faqs)} soru",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END
        if action == "add":
            context.user_data["action"] = "faq_q"
            await query.edit_message_text(TR["faq_enter_q"],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_FAQ_Q
        if action == "list":
            faqs = load_faq()
            if not faqs:
                await query.answer(TR["faq_empty"], show_alert=True); return ConversationHandler.END
            lines = [f"{TR['faq_list_title'].format(len(faqs))}"]
            for i, item in enumerate(faqs):
                q = ", ".join(item.get("keywords",[]))[:50]
                a = item.get("answer","")[:60]
                lines.append(f"\n{i+1}. ❓ {q}\n   💬 {a}")
            await query.edit_message_text(
                "\n".join(lines)[:4000],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")]]))
            return ConversationHandler.END
        if action == "del":
            faqs = load_faq()
            if not faqs:
                await query.answer(TR["faq_empty"], show_alert=True); return ConversationHandler.END
            kb = [[InlineKeyboardButton(f"🗑 {item.get('keywords',['?'])[0][:40]}",
                                        callback_data=f"faqmgmt|del_confirm|{i}")]
                  for i, item in enumerate(faqs)]
            kb.append([InlineKeyboardButton(TR["cancel"], callback_data="close")])
            await query.edit_message_text("Silinecek soruyu seçin:", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END
        if action == "del_confirm":
            idx  = int(cb.split("|")[2])
            faqs = load_faq()
            if 0 <= idx < len(faqs):
                faqs.pop(idx); save_faq(faqs)
            await query.answer(TR["faq_deleted"], show_alert=True)
            try: await query.delete_message()
            except: pass
            return ConversationHandler.END

    # ── Otomatik Kural Yönetimi (admin) ──────────────
    if cb.startswith("rulemgmt|") and is_main_admin(uid):
        action = cb.split("|")[1]
        if action == "panel":
            rules = load_auto_rules()
            kb = [
                [InlineKeyboardButton(TR["rule_add_btn"],  callback_data="rulemgmt|add"),
                 InlineKeyboardButton(TR["rule_list_btn"], callback_data="rulemgmt|list")],
                [InlineKeyboardButton(TR["rule_del_btn"],  callback_data="rulemgmt|del")],
                [InlineKeyboardButton("◀️ Geri",           callback_data="nav|root")],
            ]
            await query.edit_message_text(
                f"{TR['rule_panel']}\n\n⚡ Toplam: {len(rules)} kural",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END
        if action == "add":
            context.user_data["action"] = "rule_kw"
            await query.edit_message_text(TR["rule_enter_kw"],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_RULE_KW
        if action == "list":
            rules = load_auto_rules()
            if not rules:
                await query.answer(TR["rule_empty"], show_alert=True); return ConversationHandler.END
            lines = [f"⚡ Kurallar ({len(rules)} adet):"]
            for i, r in enumerate(rules):
                kws  = ", ".join(r.get("keywords",[]))[:40]
                resp = r.get("response","")[:50]
                lines.append(f"\n{i+1}. 🔍 {kws}\n   💬 {resp}")
            await query.edit_message_text(
                "\n".join(lines)[:4000],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")]]))
            return ConversationHandler.END
        if action == "del":
            rules = load_auto_rules()
            if not rules:
                await query.answer(TR["rule_empty"], show_alert=True); return ConversationHandler.END
            kb = [[InlineKeyboardButton(f"🗑 {', '.join(r.get('keywords',[]))[:40]}",
                                        callback_data=f"rulemgmt|del_confirm|{i}")]
                  for i, r in enumerate(rules)]
            kb.append([InlineKeyboardButton(TR["cancel"], callback_data="close")])
            await query.edit_message_text("Silinecek kuralı seçin:", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END
        if action == "del_confirm":
            idx   = int(cb.split("|")[2])
            rules = load_auto_rules()
            if 0 <= idx < len(rules):
                rules.pop(idx); save_auto_rules(rules)
            await query.answer(TR["rule_deleted"], show_alert=True)
            try: await query.delete_message()
            except: pass
            return ConversationHandler.END

    # ── Yedekleme (admin) ────────────────────────────
    if cb == "backup|do" and is_main_admin(uid):
        await query.answer(TR["backup_sending"], show_alert=False)
        files_to_send = [
            (DATA_FILE,     "bot_data.json"),
            (USERS_FILE,    "users.json"),
            (MESSAGES_FILE, "user_messages.json"),
            (SETTINGS_FILE, "settings.json"),
            (POLLS_FILE,    "polls.json"),
            (CLASS_FILE,    "user_classes.json"),
            (FAQ_FILE,      "faq.json"),
            (VIEW_COUNTS_FILE, "view_counts.json"),
        ]
        sent_count = 0
        for fpath, fname in files_to_send:
            if os.path.exists(fpath):
                try:
                    with open(fpath, "rb") as f:
                        await context.bot.send_document(
                            int(uid), f, filename=fname,
                            caption=f"💾 {fname}")
                    sent_count += 1
                except Exception as e:
                    logger.warning(f"Yedek gönderilemedi {fname}: {e}")
        await context.bot.send_message(int(uid), TR["backup_done"].format(sent_count))
        log_admin_action(uid, "BACKUP", f"{sent_count} dosya yedeklendi")
        return ConversationHandler.END

    # ── Kullanıcı dışa aktar (admin) ────────────────
    if cb == "export|users" and is_main_admin(uid):
        users   = load_users()
        classes = load_classes()
        lines = ["ID,Ad,Kullanıcı Adı,Sınıf,Son Görülme"]
        for uid_, u in users.items():
            if int(uid_) == ADMIN_ID: continue
            cls  = CLASS_DEFS.get(classes.get(uid_,""), {}).get("tr","—")
            name = u.get("full_name") or u.get("first_name","—")
            un   = u.get("username","—")
            last = u.get("last_seen","—")
            lines.append(f"{uid_},{name},{un},{cls},{last}")
        csv_text = "\n".join(lines).encode("utf-8-sig")
        import io
        await context.bot.send_document(
            int(uid),
            io.BytesIO(csv_text),
            filename="kullanici_listesi.csv",
            caption=TR["export_done"].format(len(lines)-1))
        log_admin_action(uid, "EXPORT_USERS", f"{len(lines)-1} kullanıcı")
        return ConversationHandler.END

    # ── Hedefli Duyuru ────────────────────────────────
    if cb.startswith("bcast|") and is_main_admin(uid):
        action = cb.split("|")[1]
        if action == "panel":
            kb = [
                [InlineKeyboardButton(TR["bcast_all"],    callback_data="bcast|target|all"),
                 InlineKeyboardButton(TR["bcast_active"], callback_data="bcast|target|active")],
                [InlineKeyboardButton(TR["bcast_new"],    callback_data="bcast|target|new")],
                [InlineKeyboardButton("🎯 " + TR["bcast_targeted"], callback_data="bcast|tgt_start")],
                [InlineKeyboardButton("◀️ Geri",          callback_data="nav|root")],
            ]
            await query.edit_message_text(TR["bcast_targeted"], reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "tgt_start":
            # Adım 1: Sınıf seç
            kb = build_target_keyboard("bcast|tgt", "cls")
            kb.append([InlineKeyboardButton("◀️", callback_data="close")])
            await query.edit_message_text("📢 السنة الدراسية:", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "tgt":
            # Adım adım hedefleme
            parts_b = cb.split("|")
            step_b  = parts_b[-1]
            val_b   = "|".join(parts_b[2:-1]) if len(parts_b) > 3 else (parts_b[2] if len(parts_b) > 2 else "")
            if step_b == "done":
                target_uids = get_target_users(val_b)
                lbl = target_label(val_b)
                context.user_data["bcast_targets"] = target_uids
                context.user_data["bcast_label"]   = lbl
                context.user_data["action"]        = "broadcast_targeted"
                await query.edit_message_text(
                    TR["bcast_confirm"].format(len(target_uids)) + f"\n\n🎯 {lbl}",
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("✅ Evet, Gönder", callback_data="bcast|confirm"),
                        InlineKeyboardButton(TR["cancel"],      callback_data="close"),
                    ]]))
            else:
                kb = build_target_keyboard("bcast|tgt", step_b, val_b)
                kb.append([InlineKeyboardButton("◀️", callback_data="close")])
                labels_b = {"cls":"السنة","shift":"الفترة","grp":"المجموعة","subgrp":"الفرعية"}
                await query.edit_message_text(
                    f"📢 {labels_b.get(step_b, step_b)}:",
                    reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "class_select":
            kb = [[InlineKeyboardButton(CLASS_DEFS[c]["tr"], callback_data=f"bcast|target|class_{c}")]
                  for c in CLASS_DEFS]
            kb.append([InlineKeyboardButton(TR["cancel"], callback_data="close")])
            await query.edit_message_text(TR["bcast_class_select"], reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END
        if action == "target":
            t = cb.split("|")[2]
            if t == "all":
                targets = [uid_ for uid_ in load_users() if int(uid_) != ADMIN_ID and not is_blocked(uid_)]
                label   = TR["bcast_all"]
            elif t == "active":
                targets = [uid_ for uid_ in active_users(7) if not is_blocked(uid_)]
                label   = TR["bcast_active"]
            elif t == "new":
                targets = [uid_ for uid_ in new_users(7) if not is_blocked(uid_)]
                label   = TR["bcast_new"]
            elif t.startswith("class_"):
                cls     = t.split("_")[1]
                targets = [uid_ for uid_ in users_by_class(cls) if not is_blocked(uid_)]
                label   = CLASS_DEFS.get(cls,{}).get("tr","?")
            else:
                targets = []
                label   = "?"
            context.user_data["bcast_targets"] = targets
            context.user_data["bcast_label"]   = label
            context.user_data["action"]        = "broadcast_targeted"
            await query.edit_message_text(
                TR["bcast_confirm"].format(len(targets)) + f"\n\n{TR['bcast_target_set'].format(label)}",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("✅ Evet, Gönder", callback_data="bcast|confirm"),
                     InlineKeyboardButton(TR["cancel"],      callback_data="close")],
                ]))
            return ConversationHandler.END
        if action == "confirm":
            targets = context.user_data.get("bcast_targets", [])
            label   = context.user_data.get("bcast_label","")
            context.user_data.pop("bcast_targets", None)
            context.user_data.pop("bcast_label",   None)
            context.user_data["action"] = "broadcast"
            await query.edit_message_text(
                TR["broadcast_enter"],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            context.user_data["broadcast_targets"] = targets
            return WAIT_BROADCAST_MSG

    # ── Admin İşlem Günlüğü (admin) ──────────────────
    if cb == "admin|log" and is_main_admin(uid):
        log  = load_admin_log()
        if not log:
            await query.answer(TR["admin_log_empty"], show_alert=True); return ConversationHandler.END
        lines = [TR["admin_log_title"]]
        for entry in reversed(log[-30:]):
            lines.append(
                f"🕐 {entry['time']}\n"
                f"  ⚙️ {entry['action']}"
                + (f": {entry['detail']}" if entry.get('detail') else "")
            )
        await query.edit_message_text(
            "\n".join(lines)[:4000],
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")]]))
        return ConversationHandler.END


    # ── Admin Yetki Yönetimi ──────────────────────────
    if cb.startswith("adminperm|") and is_main_admin(uid):
        parts_p = cb.split("|")
        act_p   = parts_p[1]
        adm_id  = parts_p[2] if len(parts_p) > 2 else ""

        if act_p == "show":
            u = load_users().get(str(adm_id), {})
            name = u.get("full_name") or u.get("first_name") or f"ID:{adm_id}"
            perms = load_admin_perms().get(str(adm_id), DEFAULT_ADMIN_PERMS)
            perm_labels = [
                ("can_add_folder",  "📁 Klasör Ekle"),
                ("can_del_folder",  "🗑 Klasör Sil"),
                ("can_add_file",    "📎 Dosya Ekle"),
                ("can_del_file",    "🗑 Dosya Sil"),
                ("can_rename_file", "✏️ Yeniden Adlandır"),
                ("can_countdown",   "📅 Sınav Ekle"),
                ("can_poll",        "📊 Anket Kur"),
                ("can_quiz",        "📝 Quiz"),
                ("can_broadcast",   "📢 Duyuru Gönder"),
                ("can_reply",       "💬 Kullanıcıya Cevap"),
                ("can_warn",        "⚠️ Uyarı Ver"),
                ("can_block",       "🚫 Engelle"),
                ("can_view_users",  "👥 Kullanıcı Görme"),
            ]
            kb = []
            for pkey, plabel in perm_labels:
                val  = perms.get(pkey, DEFAULT_ADMIN_PERMS.get(pkey, False))
                icon = "✅" if val else "❌"
                kb.append([InlineKeyboardButton(
                    f"{icon} {plabel}",
                    callback_data=f"adminperm|toggle|{adm_id}|{pkey}")])
            # Sınıf kısıtı
            cls_val = perms.get("cls", None)
            grp_val = perms.get("grp", None)
            sub_val = perms.get("subgrp", None)
            cls_lbl = cls_val or "Hepsi"
            grp_lbl = grp_val or "Hepsi"
            sub_lbl = sub_val or "Hepsi"
            kb.append([
                InlineKeyboardButton(f"Sinif: {cls_lbl}", callback_data=f"adminperm|cls|{adm_id}"),
                InlineKeyboardButton(f"Grup: {grp_lbl}",  callback_data=f"adminperm|grp|{adm_id}"),
            ])
            if grp_val:
                kb.append([InlineKeyboardButton(
                    f"Alt Grup: {sub_lbl}",
                    callback_data=f"adminperm|subgrp|{adm_id}|{grp_val}")])
            kb.append([InlineKeyboardButton("Geri", callback_data="set|admin_perms")])
            await query.edit_message_text(
                f"👮 {name}\n\nYetkilerini düzenle:",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if act_p == "toggle":
            pkey = parts_p[3] if len(parts_p) > 3 else ""
            perms = load_admin_perms()
            adm_perms = perms.get(str(adm_id), dict(DEFAULT_ADMIN_PERMS))
            adm_perms[pkey] = not adm_perms.get(pkey, DEFAULT_ADMIN_PERMS.get(pkey, False))
            perms[str(adm_id)] = adm_perms
            save_admin_perms(perms)
            # Ekranı güncelle
            cb_new = f"adminperm|show|{adm_id}"
            query.data = cb_new
            return await callback(update, context)

        if act_p == "cls":
            kb = [
                [InlineKeyboardButton("Hepsi", callback_data=f"adminperm|setcls|{adm_id}|all")],
                [InlineKeyboardButton("1.Sinif", callback_data=f"adminperm|setcls|{adm_id}|1"),
                 InlineKeyboardButton("2.Sinif", callback_data=f"adminperm|setcls|{adm_id}|2"),
                 InlineKeyboardButton("3.Sinif", callback_data=f"adminperm|setcls|{adm_id}|3"),
                 InlineKeyboardButton("4.Sinif", callback_data=f"adminperm|setcls|{adm_id}|4")],
            ]
            await query.edit_message_text("Sinif kisiti:", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if act_p == "setcls":
            cls_val = None if parts_p[3] == "all" else parts_p[3]
            set_admin_perm(adm_id, "cls", cls_val)
            await query.answer("Kaydedildi", show_alert=True)
            query.data = f"adminperm|show|{adm_id}"
            return await callback(update, context)

        if act_p == "grp":
            # Grup kısıtı — A/B/C + None
            kb = [
                [InlineKeyboardButton("Hepsi", callback_data=f"adminperm|setgrp|{adm_id}|all")],
                [InlineKeyboardButton("A",     callback_data=f"adminperm|setgrp|{adm_id}|A"),
                 InlineKeyboardButton("B",     callback_data=f"adminperm|setgrp|{adm_id}|B"),
                 InlineKeyboardButton("C",     callback_data=f"adminperm|setgrp|{adm_id}|C")],
            ]
            await query.edit_message_text("Grup kisiti:", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if act_p == "setgrp":
            grp_val = None if parts_p[3] == "all" else parts_p[3]
            set_admin_perm(adm_id, "grp", grp_val)
            await query.answer("Kaydedildi", show_alert=True)
            query.data = f"adminperm|show|{adm_id}"
            return await callback(update, context)

        if act_p == "subgrp":
            # Alt grup kısıtı — A1/A2/A3 vb.
            grp_main = parts_p[3] if len(parts_p) > 3 else "A"
            kb = [
                [InlineKeyboardButton("Hepsi", callback_data=f"adminperm|setsubgrp|{adm_id}|all")],
                [InlineKeyboardButton(f"{grp_main}1", callback_data=f"adminperm|setsubgrp|{adm_id}|{grp_main}1"),
                 InlineKeyboardButton(f"{grp_main}2", callback_data=f"adminperm|setsubgrp|{adm_id}|{grp_main}2"),
                 InlineKeyboardButton(f"{grp_main}3", callback_data=f"adminperm|setsubgrp|{adm_id}|{grp_main}3")],
            ]
            await query.edit_message_text(f"Alt grup ({grp_main}):", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if act_p == "setsubgrp":
            subgrp_val = None if parts_p[3] == "all" else parts_p[3]
            set_admin_perm(adm_id, "subgrp", subgrp_val)
            await query.answer("Kaydedildi", show_alert=True)
            query.data = f"adminperm|show|{adm_id}"
            return await callback(update, context)

        return ConversationHandler.END


    # ── Quiz Oluştur (admin) ──────────────────────────
    if cb == "admin|quiz_create" and is_main_admin(uid):
        context.user_data["action"] = "quiz_question"
        await query.edit_message_text(
            "📝 Quiz sorusunu yazın:\n\n"
            "Format:\n"
            "Soru metni\n"
            "A. Seçenek 1\n"
            "B. Seçenek 2\n"
            "C. Seçenek 3\n"
            "D. Seçenek 4\n"
            "CEVAP: A",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ İptal", callback_data="nav|root")
            ]]))
        return WAIT_FOLDER

    # ── Sınıf Analizi (admin) ─────────────────────────
    if cb == "admin|class_analysis" and is_main_admin(uid):
        users_d = load_users()
        msgs_d  = load_messages()
        lines   = ["📊 SINIF ANALİZİ", "═"*24]
        for cls_id, cls_def in CLASS_DEFS.items():
            cls_users = users_by_class(cls_id)
            cls_msgs  = sum(len(msgs_d.get(u,"")) for u in cls_users)
            active_7  = sum(1 for u in cls_users
                           if u in msgs_d and msgs_d[u] and
                           msgs_d[u][-1].get("time","") >= (datetime.now(IRAQ_TZ).replace(
                               hour=0,minute=0,second=0).strftime("%Y-%m-%d")))
            lines.append(
                f"\n🎓 {cls_def['ar']}\n"
                f"  👥 {len(cls_users)} öğrenci\n"
                f"  💬 {cls_msgs} kayıt\n"
                f"  🟢 {active_7} bugün aktif"
            )
        await query.edit_message_text(
            "\n".join(lines)[:4000],
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Geri", callback_data="nav|root")
            ]]))
        return ConversationHandler.END

    # ── Sınıf İstatistikleri (admin) ─────────────────
    if cb == "mgmt|class_stats" and is_main_admin(uid):
        classes = load_classes()
        counts  = {c: 0 for c in CLASS_DEFS}
        none_c  = 0
        for cls in classes.values():
            if cls in counts: counts[cls] += 1
            else: none_c += 1
        lines = ["🎓 SINIF İSTATİSTİKLERİ", ""]
        for c, cnt in counts.items():
            label = CLASS_DEFS[c]["tr"]
            bar   = "█" * min(cnt, 20)
            lines.append(f"{label}: {cnt} kişi  {bar}")
        lines.append(f"❓ Seçmemiş: {none_c} kişi")
        await query.edit_message_text(
            "\n".join(lines),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")]]))
        return ConversationHandler.END

    # ── İçerik işlemleri — pin/move/copy/sort/folder_desc ──
    if cb.startswith("extra|") and is_main_admin(uid):
        action = cb.split("|")[1]
        folder = get_folder(content, path)

        if action == "sort_az":
            files = folder.get("files", [])
            files.sort(key=lambda f: f.get("caption", f.get("name","")).lower())
            save_content(content)
            await show_folder(query, context, path, note=L(uid,"files_sorted"))
            log_admin_action(uid, "SORT_AZ", "›".join(path))
            return ConversationHandler.END

        if action == "sort_views":
            files = folder.get("files", [])
            vc    = load_view_counts()
            files.sort(key=lambda f: vc.get(f.get("file_id") or f.get("caption",""),{}).get("count",0), reverse=True)
            save_content(content)
            await show_folder(query, context, path, note=L(uid,"files_sorted"))
            return ConversationHandler.END

        if action == "folder_desc":
            context.user_data["action"] = "folder_desc"
            desc = get_folder_desc(path)
            prompt = L(uid,"folder_desc_prompt")
            if desc:
                prompt += f"\n\nMevcut: {desc[:100]}"
            await query.edit_message_text(prompt,
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")]]))
            return WAIT_FOLDER_DESC

        if action == "pin":
            files = folder.get("files", [])
            if not files:
                await query.answer(L(uid,"no_files"), show_alert=True); return ConversationHandler.END
            kb = [[InlineKeyboardButton(f"📌 {f.get('caption','?')[:35]}", callback_data=f"extra|do_pin|{i}")]
                  for i,f in enumerate(files)]
            kb.append([InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")])
            await query.edit_message_text(L(uid,"pin_select"), reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "do_pin":
            idx   = int(cb.split("|")[2])
            files = folder.get("files",[])
            if 0 <= idx < len(files):
                files[idx]["pinned"] = not files[idx].get("pinned", False)
                save_content(content)
                fname  = files[idx].get("caption","?")
                is_pin = files[idx]["pinned"]
                await show_folder(query, context, path,
                    note=L(uid,"file_pinned" if is_pin else "file_unpinned").format(fname))
                log_admin_action(uid, "PIN_FILE" if is_pin else "UNPIN_FILE", fname)
            return ConversationHandler.END

        if action == "move":
            files = folder.get("files",[])
            if not files:
                await query.answer(L(uid,"no_files"), show_alert=True); return ConversationHandler.END
            kb = [[InlineKeyboardButton(f"📂 {f.get('caption','?')[:35]}", callback_data=f"extra|do_move_sel|{i}")]
                  for i,f in enumerate(files)]
            kb.append([InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")])
            await query.edit_message_text(L(uid,"move_select_file"), reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "do_move_sel":
            context.user_data["move_idx"] = int(cb.split("|")[2])
            # Hedef klasörleri listele
            all_paths = all_folder_paths()
            kb = []
            for p in all_paths[:20]:
                if p == path: continue
                label = " › ".join(p)
                penc  = "~~".join(p)
                kb.append([InlineKeyboardButton(f"📁 {label}", callback_data=f"extra|do_move_dest|{penc}")])
            kb.append([InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")])
            await query.edit_message_text(
                L(uid,"move_select_dest").format(" › ".join(path) or "🏠"),
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "do_move_dest":
            dest_enc = cb.split("|",2)[2]
            dest_path = dest_enc.split("~~") if dest_enc else []
            src_idx   = context.user_data.pop("move_idx", -1)
            src_files = folder.get("files",[])
            if 0 <= src_idx < len(src_files):
                f     = src_files.pop(src_idx)
                dest_folder = get_folder(content, dest_path)
                dest_folder.setdefault("files",[]).append(f)
                save_content(content)
                fname = f.get("caption","?")
                dest_name = " › ".join(dest_path) or "🏠"
                await show_folder(query, context, path, note=L(uid,"file_moved").format(fname, dest_name))
                log_admin_action(uid, "MOVE_FILE", f"{fname} → {dest_name}")
                # Abonelere bildir
                for sub_uid in get_folder_subscribers(dest_path):
                    try:
                        await context.bot.send_message(
                            int(sub_uid),
                            L(sub_uid,"sub_notif").format(folder=dest_name, fname=fname))
                    except: pass
            return ConversationHandler.END

        if action == "copy":
            files = folder.get("files",[])
            if not files:
                await query.answer(L(uid,"no_files"), show_alert=True); return ConversationHandler.END
            kb = [[InlineKeyboardButton(f"📋 {f.get('caption','?')[:35]}", callback_data=f"extra|do_copy_sel|{i}")]
                  for i,f in enumerate(files)]
            kb.append([InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")])
            await query.edit_message_text(L(uid,"copy_select_file"), reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "do_copy_sel":
            context.user_data["copy_idx"] = int(cb.split("|")[2])
            all_paths = all_folder_paths()
            kb = []
            for p in all_paths[:20]:
                label = " › ".join(p)
                penc  = "~~".join(p)
                kb.append([InlineKeyboardButton(f"📁 {label}", callback_data=f"extra|do_copy_dest|{penc}")])
            kb.append([InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")])
            await query.edit_message_text(
                L(uid,"move_select_dest").format(" › ".join(path) or "🏠"),
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "do_copy_dest":
            import copy as _copy
            dest_enc  = cb.split("|",2)[2]
            dest_path = dest_enc.split("~~") if dest_enc else []
            src_idx   = context.user_data.pop("copy_idx", -1)
            src_files = folder.get("files",[])
            if 0 <= src_idx < len(src_files):
                f    = _copy.deepcopy(src_files[src_idx])
                dest_folder = get_folder(content, dest_path)
                dest_folder.setdefault("files",[]).append(f)
                save_content(content)
                fname = f.get("caption","?")
                dest_name = " › ".join(dest_path) or "🏠"
                await show_folder(query, context, path, note=L(uid,"file_copied").format(fname, dest_name))
                log_admin_action(uid, "COPY_FILE", f"{fname} → {dest_name}")
            return ConversationHandler.END

    # ── Alt Admin Duyuru ─────────────────────────────
    if cb.startswith("admin_bcast|") and is_admin(uid) and not is_main_admin(uid):
        if not get_admin_perm(uid, "can_broadcast"):
            await query.answer("ليس لديك صلاحية الإعلانات", show_alert=True); return ConversationHandler.END
        ab_parts  = cb.split("|")
        ab_action = ab_parts[1]
        adm_cls   = get_admin_cls(uid)   # None=hepsi
        adm_grp   = get_admin_grp(uid)   # None=hepsi

        if ab_action == "panel":
            # Adım 1: sınıf seç (sabit sınıf varsa atla)
            if adm_cls:
                context.user_data["ab_cls"] = adm_cls
                # Vardiya seç
                kb = [
                    [InlineKeyboardButton("☀️ صباحي", callback_data="admin_bcast|sft|sabahi"),
                     InlineKeyboardButton("🌙 مسائي",  callback_data="admin_bcast|sft|masaiy"),
                     InlineKeyboardButton("📋 الكل",   callback_data="admin_bcast|sft|all")],
                    [InlineKeyboardButton("◀️ إلغاء",  callback_data="nav|root")],
                ]
                await query.edit_message_text("📢 الفترة الدراسية:", reply_markup=InlineKeyboardMarkup(kb))
            else:
                # Sınıf yok — sınıf seç
                kb = []
                for cls_id, cls_def in CLASS_DEFS.items():
                    kb.append([InlineKeyboardButton(cls_def["ar"], callback_data=f"admin_bcast|cls|{cls_id}")])
                kb.append([InlineKeyboardButton("📋 الكل", callback_data="admin_bcast|cls|all")])
                kb.append([InlineKeyboardButton("◀️ إلغاء", callback_data="nav|root")])
                await query.edit_message_text("📢 السنة الدراسية:", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if ab_action == "cls":
            cls_val = ab_parts[2] if len(ab_parts) > 2 else "all"
            context.user_data["ab_cls"] = "" if cls_val == "all" else cls_val
            # Vardiya seç
            kb = [
                [InlineKeyboardButton("☀️ صباحي", callback_data="admin_bcast|sft|sabahi"),
                 InlineKeyboardButton("🌙 مسائي",  callback_data="admin_bcast|sft|masaiy"),
                 InlineKeyboardButton("📋 الكل",   callback_data="admin_bcast|sft|all")],
                [InlineKeyboardButton("◀️ رجوع",   callback_data="admin_bcast|panel")],
            ]
            await query.edit_message_text("📢 الفترة الدراسية:", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if ab_action == "sft":
            sft_val = ab_parts[2] if len(ab_parts) > 2 else "all"
            context.user_data["ab_sft"] = "" if sft_val == "all" else sft_val
            cls_fixed = adm_cls or context.user_data.get("ab_cls","")
            if adm_grp:
                # Grup sabit — direkt yaz
                context.user_data["ab_grp"] = adm_grp
                context.user_data["action"] = "admin_bcast_msg"
                await query.edit_message_text("📢 اكتب نص الإعلان:",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ إلغاء","nav|root")]]))
                return WAIT_FOLDER
            # Grup seç
            grp_keys = list(get_class_groups(cls_fixed).keys()) if cls_fixed else ["A","B","C"]
            kb = []
            row = [InlineKeyboardButton("📋 الكل", callback_data="admin_bcast|grp|all")]
            for g in grp_keys:
                row.append(InlineKeyboardButton(g, callback_data=f"admin_bcast|grp|{g}"))
            kb.append(row)
            kb.append([InlineKeyboardButton("◀️ رجوع", callback_data="admin_bcast|panel")])
            await query.edit_message_text("📢 المجموعة:", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if ab_action == "grp":
            grp_val = ab_parts[2] if len(ab_parts) > 2 else "all"
            context.user_data["ab_grp"] = "" if grp_val == "all" else grp_val
            context.user_data["action"] = "admin_bcast_msg"
            await query.edit_message_text("📢 اكتب نص الإعلان:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ إلغاء","nav|root")]]))
            return WAIT_FOLDER

        return ConversationHandler.END

    # ── Laboratuvar Programı ─────────────────────────
    if cb.startswith("lab|") and is_admin(uid) and (is_main_admin(uid) or get_admin_perm(uid, "can_countdown")):
        lab_parts = cb.split("|")
        lab_act   = lab_parts[1]

        if lab_act == "add_new":
            context.user_data["action"] = "lab_add_name"
            await query.message.reply_text(
                "🔬 Lab adını yaz (örn: Elektronik Lab, Fizik Lab):")
            return ConversationHandler.END

        if lab_act == "lab_cls":
            cls_val = lab_parts[2] if len(lab_parts) > 2 else "all"
            context.user_data["pending_lab_cls"] = "" if cls_val == "all" else cls_val
            context.user_data["action"] = "lab_add_date"
            await query.message.reply_text("📅 Tarih yaz (DD/MM/YYYY):\nÖrn: 22/04/2026")
            return WAIT_FOLDER

        if lab_act == "view":
            week  = lab_parts[2]
            sched = load_lab_schedule()
            entry = next((e for e in sched if e.get("week")==week), None)
            grp   = entry.get("group","?") if entry else "?"
            # Grup değiştir
            grps_cfg = load_class_groups()
            all_subs = []
            for cls_d in grps_cfg.values():
                for g, subs in cls_d.items():
                    all_subs.extend(subs)
            all_subs = sorted(set(all_subs))
            kb = []
            row = []
            for sub in all_subs:
                icon = "✅" if sub == grp else ""
                row.append(InlineKeyboardButton(f"{icon}{sub}", callback_data=f"lab|set|{week}|{sub}"))
                if len(row) == 4:
                    kb.append(row); row = []
            if row: kb.append(row)
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data="mgmt|lab")])
            await query.edit_message_text(f"🔬 {week}\nMevcut: {grp}\nGrup seç:", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if lab_act == "del":
            week  = lab_parts[2]
            sched = load_lab_schedule()
            sched = [e for e in sched if e.get("week") != week]
            save_lab_schedule(sched)
            await query.answer(f"✅ {week} silindi", show_alert=True)
            query.data = "mgmt|lab"; return await callback(update, context)

        if lab_act == "set":
            week  = lab_parts[2]
            group = lab_parts[3]
            lab_name = context.user_data.pop("pending_lab_name", "")
            add_lab_week(week, group, note=lab_name)
            # Onay mesajı
            from datetime import datetime as _dt2
            try:
                d2 = _dt2.strptime(week, "%Y-%m-%d")
                ARABIC_DAYS2 = {0:"الاثنين",1:"الثلاثاء",2:"الأربعاء",3:"الخميس",4:"الجمعة",5:"السبت",6:"الأحد"}
                day_ar = ARABIC_DAYS2.get(d2.weekday(),"")
                date_disp = f"{day_ar} {d2.strftime('%d/%m/%Y')}"
            except:
                date_disp = week
            confirm = (
                f"✅ تم إضافة موعد المختبر\n\n"
                f"🔬 المختبر: {lab_name or '—'}\n"
                f"📅 التاريخ: {date_disp}\n"
                f"👥 المجموعة: {group}"
            )
            try:
                await query.message.reply_text(confirm)
            except: pass
            await query.answer("✅", show_alert=False)
            query.data = "mgmt|lab"
            return await callback(update, context)

    # ── Arama sonucu dosya/klasör aç ─────────────────
    if cb.startswith("goto_folder|"):
        fp_enc = cb.split("|", 1)[1]
        new_path = fp_enc.split("~~") if fp_enc else []
        context.user_data["path"] = new_path
        context.user_data["page"] = 0
        context.user_data.pop("action", None)
        await show_folder(query, context, new_path)
        return ConversationHandler.END

    if cb.startswith("srch|"):
        parts        = cb.split("|")
        path_encoded = parts[1]
        idx          = int(parts[2])
        file_path    = path_encoded.split("~~") if path_encoded else []
        folder       = get_folder(content, file_path)
        files        = folder.get("files", [])
        if idx >= len(files):
            await query.answer(L(uid,"file_notfound"), show_alert=True); return ConversationHandler.END
        f = files[idx]
        await _send_file(query.message, f, uid)
        log_user_message(user, "file_view", f.get("caption",""))
        return ConversationHandler.END

    # ── Klasördeki dosyayı aç ────────────────────────
    if cb.startswith("getfile|"):
        idx    = int(cb.split("|")[1])
        folder = get_folder(content, path)
        files  = folder.get("files",[])
        if idx >= len(files):
            await query.answer(L(uid,"file_notfound"), show_alert=True); return ConversationHandler.END
        f = files[idx]
        await _send_file(query.message, f, uid)
        log_user_message(user, "file_view", f.get("caption",""))
        return ConversationHandler.END

    # ── Sayfa navigasyonu ─────────────────────────────
    if cb == "noop":
        return ConversationHandler.END

    # ── Onay sistemi (süper admin onaylıyor) ──────────
    if cb.startswith("newadmin|") and is_main_admin(uid):
        parts_n = cb.split("|")
        act_n   = parts_n[1]
        adm_id  = parts_n[2]

        if act_n == "cls":
            cls_val = None if parts_n[3] == "all" else parts_n[3]
            set_admin_perm(adm_id, "cls", cls_val)
            if cls_val:
                # Grup sor
                kb = [
                    [InlineKeyboardButton("Hepsi", callback_data=f"newadmin|grp|{adm_id}|all")],
                    [InlineKeyboardButton("A",     callback_data=f"newadmin|grp|{adm_id}|A"),
                     InlineKeyboardButton("B",     callback_data=f"newadmin|grp|{adm_id}|B"),
                     InlineKeyboardButton("C",     callback_data=f"newadmin|grp|{adm_id}|C")],
                    [InlineKeyboardButton("Atla",  callback_data=f"newadmin|skip|{adm_id}")],
                ]
                await query.edit_message_text(f"Sinif: {cls_val}\nHangi grubu yonetsin?", reply_markup=InlineKeyboardMarkup(kb))
            else:
                await query.edit_message_text(f"Admin {adm_id} eklendi. Kisit yok.")
            return ConversationHandler.END

        if act_n == "grp":
            grp_val = None if parts_n[3] == "all" else parts_n[3]
            set_admin_perm(adm_id, "grp", grp_val)
            if grp_val:
                # Alt grup sor
                kb = [
                    [InlineKeyboardButton("Hepsi", callback_data=f"newadmin|sub|{adm_id}|all")],
                    [InlineKeyboardButton(f"{grp_val}1", callback_data=f"newadmin|sub|{adm_id}|{grp_val}1"),
                     InlineKeyboardButton(f"{grp_val}2", callback_data=f"newadmin|sub|{adm_id}|{grp_val}2"),
                     InlineKeyboardButton(f"{grp_val}3", callback_data=f"newadmin|sub|{adm_id}|{grp_val}3")],
                    [InlineKeyboardButton("Atla", callback_data=f"newadmin|skip|{adm_id}")],
                ]
                cls_v = get_admin_cls(adm_id) or "?"
                await query.edit_message_text(
                    f"Sinif: {cls_v}  Grup: {grp_val}\nAlt grup?",
                    reply_markup=InlineKeyboardMarkup(kb))
            else:
                await query.edit_message_text(f"Admin {adm_id} ayarlandi.")
            return ConversationHandler.END

        if act_n == "sub":
            sub_val = None if parts_n[3] == "all" else parts_n[3]
            set_admin_perm(adm_id, "subgrp", sub_val)
            cls_v = get_admin_cls(adm_id) or "Hepsi"
            grp_v = get_admin_grp(adm_id) or "Hepsi"
            await query.edit_message_text(
                f"Tamamlandi!\nAdmin: {adm_id}\nSinif: {cls_v}\nGrup: {grp_v}\nAlt grup: {sub_val or 'Hepsi'}")
            return ConversationHandler.END

        if act_n == "skip":
            await query.edit_message_text(f"Admin {adm_id} eklendi. Yetkiler sonra ayarlanabilir.")
            return ConversationHandler.END

    # ── Admin Ekle: Önce sınıf/grup seç ──────────────
    if cb.startswith("newadmin_pre|") and is_main_admin(uid):
        parts_pre = cb.split("|")
        step_pre  = parts_pre[1]
        val_pre   = parts_pre[2] if len(parts_pre) > 2 else "all"

        if step_pre == "cls":
            context.user_data["pending_admin_cls"] = None if val_pre == "all" else val_pre
            if val_pre != "all":
                # Grup seç
                kb = [
                    [InlineKeyboardButton("Hepsi", callback_data="newadmin_pre|grp|all")],
                    [InlineKeyboardButton("A", callback_data="newadmin_pre|grp|A"),
                     InlineKeyboardButton("B", callback_data="newadmin_pre|grp|B"),
                     InlineKeyboardButton("C", callback_data="newadmin_pre|grp|C")],
                    [InlineKeyboardButton("İptal", callback_data="close")],
                ]
                await query.edit_message_text(
                    f"Sinif: {val_pre}\nHangi grubu yonetsin?",
                    reply_markup=InlineKeyboardMarkup(kb))
            else:
                context.user_data["pending_admin_grp"] = None
                context.user_data["action"] = "admin_add"
                await query.edit_message_text(
                    TR["admin_enter_id"],
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("İptal", callback_data="close")]]))
                return WAIT_ADMIN_ID
            return ConversationHandler.END

        if step_pre == "grp":
            context.user_data["pending_admin_grp"] = None if val_pre == "all" else val_pre
            context.user_data["action"] = "admin_add"
            cls_v = context.user_data.get("pending_admin_cls","?")
            grp_v = val_pre
            await query.edit_message_text(
                f"Sinif: {cls_v}  Grup: {grp_v}\n\nAdmin ID veya @kullanici adi yazin:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("İptal", callback_data="close")]]))
            return WAIT_ADMIN_ID
        return ConversationHandler.END

    # ── Admin Ekle — Önce Sınıf/Vardiya/Grup, Sonra ID ─
    if cb.startswith("newadmin_pre|") and is_main_admin(uid):
        pre_parts = cb.split("|")
        pre_act   = pre_parts[1]

        if pre_act == "cls":
            cls_val = pre_parts[2]  # all veya 1/2/3/4
            context.user_data["new_admin_pre_cls"] = None if cls_val == "all" else cls_val
            if cls_val == "all":
                # Direkt ID'ye geç
                context.user_data["action"] = "admin_add"
                await query.edit_message_text(
                    TR["admin_enter_id"],
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("İptal","close")]]))
                return WAIT_ADMIN_ID
            # Vardiya sor
            kb = [
                [InlineKeyboardButton("Hepsi",    callback_data=f"newadmin_pre|shift|all"),
                 InlineKeyboardButton("☀️ Sabah", callback_data=f"newadmin_pre|shift|sabahi"),
                 InlineKeyboardButton("🌙 Gece",  callback_data=f"newadmin_pre|shift|masaiy")],
                [InlineKeyboardButton("İptal", callback_data="close")],
            ]
            await query.edit_message_text(
                f"Sınıf: {cls_val}\nVardiya seç:",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if pre_act == "shift":
            shift_val = pre_parts[2]
            context.user_data["new_admin_pre_shift"] = None if shift_val == "all" else shift_val
            kb = [
                [InlineKeyboardButton("Hepsi", callback_data="newadmin_pre|grp|all"),
                 InlineKeyboardButton("A",     callback_data="newadmin_pre|grp|A"),
                 InlineKeyboardButton("B",     callback_data="newadmin_pre|grp|B"),
                 InlineKeyboardButton("C",     callback_data="newadmin_pre|grp|C")],
                [InlineKeyboardButton("İptal", callback_data="close")],
            ]
            cls_v = context.user_data.get("new_admin_pre_cls","?")
            await query.edit_message_text(
                f"Sınıf: {cls_v}  Vardiya: {shift_val}\nGrup seç:",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if pre_act == "grp":
            grp_val = pre_parts[2]
            context.user_data["new_admin_pre_grp"] = None if grp_val == "all" else grp_val
            # ID gir
            cls_v   = context.user_data.get("new_admin_pre_cls","Hepsi")
            shift_v = context.user_data.get("new_admin_pre_shift","Hepsi")
            context.user_data["action"] = "admin_add"
            await query.edit_message_text(
                f"Sınıf: {cls_v}  Vardiya: {shift_v}  Grup: {grp_val}\n\nAdmin ID veya @kullanici adı yazın:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("İptal","close")]]))
            return WAIT_ADMIN_ID

        return ConversationHandler.END

    # ── Yeni Admin Sinif/Grup Kisiti ─────────────────
    if cb.startswith("newadmin|") and is_main_admin(uid):
        parts_n = cb.split("|")
        act_n   = parts_n[1]
        adm_id  = parts_n[2]
        val     = parts_n[3] if len(parts_n) > 3 else "all"
        val_real = None if val == "all" else val

        if act_n == "cls":
            set_admin_perm(adm_id, "cls", val_real)
            if val_real:
                kb = [
                    [InlineKeyboardButton("الكل",    callback_data=f"newadmin|shift|{adm_id}|all"),
                     InlineKeyboardButton("☀️ صباحي", callback_data=f"newadmin|shift|{adm_id}|sabahi"),
                     InlineKeyboardButton("🌙 مسائي", callback_data=f"newadmin|shift|{adm_id}|masaiy")],
                    [InlineKeyboardButton("Atla", callback_data=f"newadmin|grp|{adm_id}|all")],
                ]
                await query.edit_message_text(
                    f"Sinif: {val_real}\nHangi vardiyayi yonetsin?",
                    reply_markup=InlineKeyboardMarkup(kb))
            else:
                await query.edit_message_text(f"Admin {adm_id} eklendi. Kisit yok.")
            return ConversationHandler.END

        if act_n == "shift":
            shift_val = None if val == "all" else val
            set_admin_perm(adm_id, "shift", shift_val)
            kb = [
                [InlineKeyboardButton("Hepsi", callback_data=f"newadmin|grp|{adm_id}|all")],
                [InlineKeyboardButton("A", callback_data=f"newadmin|grp|{adm_id}|A"),
                 InlineKeyboardButton("B", callback_data=f"newadmin|grp|{adm_id}|B"),
                 InlineKeyboardButton("C", callback_data=f"newadmin|grp|{adm_id}|C")],
                [InlineKeyboardButton("Atla", callback_data=f"newadmin|skip|{adm_id}")],
            ]
            cls_v = get_admin_cls(adm_id) or "?"
            await query.edit_message_text(
                f"Sinif: {cls_v}  Vardiya: {shift_val or 'Hepsi'}\nHangi grubu?",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if act_n == "grp":
            set_admin_perm(adm_id, "grp", val_real)
            if val_real:
                cls_v = get_admin_cls(adm_id) or "?"
                kb = [
                    [InlineKeyboardButton("Hepsi", callback_data=f"newadmin|sub|{adm_id}|all")],
                    [InlineKeyboardButton(f"{val_real}1", callback_data=f"newadmin|sub|{adm_id}|{val_real}1"),
                     InlineKeyboardButton(f"{val_real}2", callback_data=f"newadmin|sub|{adm_id}|{val_real}2"),
                     InlineKeyboardButton(f"{val_real}3", callback_data=f"newadmin|sub|{adm_id}|{val_real}3")],
                    [InlineKeyboardButton("Atla", callback_data=f"newadmin|skip|{adm_id}")],
                ]
                await query.edit_message_text(
                    f"Sinif: {cls_v}  Grup: {val_real}\nAlt grup?",
                    reply_markup=InlineKeyboardMarkup(kb))
            else:
                await query.edit_message_text(f"Admin {adm_id} ayarlandi.")
            return ConversationHandler.END

        if act_n == "sub":
            set_admin_perm(adm_id, "subgrp", val_real)
            cls_v = get_admin_cls(adm_id) or "Hepsi"
            grp_v = get_admin_grp(adm_id) or "Hepsi"
            log_admin_action(uid, "ADMIN_SETUP", f"ID:{adm_id} sinif:{cls_v} grp:{grp_v}")
            await query.edit_message_text(
                f"Tamamlandi!\nAdmin: {adm_id}\nSinif: {cls_v}\nGrup: {grp_v}\nAlt grup: {val_real or 'Hepsi'}")
            return ConversationHandler.END

        if act_n == "skip":
            log_admin_action(uid, "ADMIN_ADDED_NOPERM", f"ID:{adm_id}")
            await query.edit_message_text(f"Admin {adm_id} eklendi. Yetkiler: Ayarlar > Admin Yetkileri")
            return ConversationHandler.END

        return ConversationHandler.END


    # ── Onay Sistemi ─────────────────────────────────
    if cb.startswith("newadmin|") and is_main_admin(uid):
        parts_n = cb.split("|")
        act_n   = parts_n[1]
        adm_id  = parts_n[2]

        if act_n == "cls":
            cls_val = None if parts_n[3] == "all" else parts_n[3]
            set_admin_perm(adm_id, "cls", cls_val)
            if cls_val:
                # Grup sor
                kb = [
                    [InlineKeyboardButton("Hepsi", callback_data=f"newadmin|grp|{adm_id}|all")],
                    [InlineKeyboardButton("A",     callback_data=f"newadmin|grp|{adm_id}|A"),
                     InlineKeyboardButton("B",     callback_data=f"newadmin|grp|{adm_id}|B"),
                     InlineKeyboardButton("C",     callback_data=f"newadmin|grp|{adm_id}|C")],
                    [InlineKeyboardButton("Atla",  callback_data=f"newadmin|skip|{adm_id}")],
                ]
                await query.edit_message_text(f"Sinif: {cls_val}\nHangi grubu yonetsin?", reply_markup=InlineKeyboardMarkup(kb))
            else:
                await query.edit_message_text(f"Admin {adm_id} eklendi. Kisit yok.")
            return ConversationHandler.END

        if act_n == "grp":
            grp_val = None if parts_n[3] == "all" else parts_n[3]
            set_admin_perm(adm_id, "grp", grp_val)
            if grp_val:
                # Alt grup sor
                kb = [
                    [InlineKeyboardButton("Hepsi", callback_data=f"newadmin|sub|{adm_id}|all")],
                    [InlineKeyboardButton(f"{grp_val}1", callback_data=f"newadmin|sub|{adm_id}|{grp_val}1"),
                     InlineKeyboardButton(f"{grp_val}2", callback_data=f"newadmin|sub|{adm_id}|{grp_val}2"),
                     InlineKeyboardButton(f"{grp_val}3", callback_data=f"newadmin|sub|{adm_id}|{grp_val}3")],
                    [InlineKeyboardButton("Atla", callback_data=f"newadmin|skip|{adm_id}")],
                ]
                cls_v = get_admin_cls(adm_id) or "?"
                await query.edit_message_text(
                    f"Sinif: {cls_v}  Grup: {grp_val}\nAlt grup?",
                    reply_markup=InlineKeyboardMarkup(kb))
            else:
                await query.edit_message_text(f"Admin {adm_id} ayarlandi.")
            return ConversationHandler.END

        if act_n == "sub":
            sub_val = None if parts_n[3] == "all" else parts_n[3]
            set_admin_perm(adm_id, "subgrp", sub_val)
            cls_v = get_admin_cls(adm_id) or "Hepsi"
            grp_v = get_admin_grp(adm_id) or "Hepsi"
            await query.edit_message_text(
                f"Tamamlandi!\nAdmin: {adm_id}\nSinif: {cls_v}\nGrup: {grp_v}\nAlt grup: {sub_val or 'Hepsi'}")
            return ConversationHandler.END

        if act_n == "skip":
            await query.edit_message_text(f"Admin {adm_id} eklendi. Yetkiler sonra ayarlanabilir.")
            return ConversationHandler.END

    if cb.startswith("approve|") and is_main_admin(uid):
        parts   = cb.split("|")
        action  = parts[1]   # warn / block / deny
        target  = parts[2]
        req_uid = parts[3] if len(parts) > 3 else ""
        u = load_users().get(target, {})
        name = u.get("full_name") or u.get("first_name") or f"ID:{target}"

        if action == "deny":
            op = parts[4] if len(parts) > 4 else "işlem"
            try:
                await context.bot.send_message(
                    int(req_uid),
                    f"❌ Talebiniz reddedildi.\n👤 {name} ({target})\n📋 İşlem: {op}")
            except: pass
            await query.edit_message_text(f"❌ Reddedildi: {name}")
            return ConversationHandler.END

        if action == "warn":
            # Sebebi approve mesajından al — mesaj text içinde "Sebep: ..." satırı var
            msg_txt = query.message.text or ""
            reason = ""
            for line in msg_txt.split("\n"):
                if "Sebep:" in line or "سبب:" in line:
                    reason = line.split(":",1)[1].strip()
            if not reason: reason = "بقرار الإدارة"
            count = add_warn(target, reason, uid)
            # Kullanıcıya bildir
            try:
                await context.bot.send_message(
                    int(target),
                    AR["warn_msg_to_user"].format(reason=reason, count=count, max=MAX_WARNS))
            except: pass
            # Alt admin'e bildir
            if req_uid:
                try:
                    await context.bot.send_message(
                        int(req_uid),
                        f"✅ الطلب تمت الموافقة عليه\n👤 {name}\nالتحذيرات: {count}/{MAX_WARNS}")
                except: pass
            # Otomatik engel
            if count >= MAX_WARNS:
                blocked = load_blocked()
                if int(target) not in blocked:
                    blocked.append(int(target)); save_blocked(blocked)
                try:
                    await context.bot.send_message(int(target), "🚫 تم حظرك من البوت.")
                except: pass
                await query.edit_message_text(f"✅ تم التحذير والحظر التلقائي: {name} ({count}/{MAX_WARNS})")
            else:
                await query.edit_message_text(f"✅ تم إرسال التحذير: {name} ({count}/{MAX_WARNS})")
            log_admin_action(uid, "WARN_APPROVED", f"target={target} reason={reason[:30]}")
            return ConversationHandler.END

        if action == "block":
            blocked = load_blocked()
            if int(target) not in blocked:
                blocked.append(int(target))
                save_blocked(blocked)
            # Kullanıcıya bildir
            try:
                await context.bot.send_message(int(target), "🚫 تم حظرك من البوت من قِبل الإدارة.")
            except: pass
            # Talep eden admin'e bildir
            try:
                await context.bot.send_message(int(req_uid), f"✅ Onaylandı: {name} engellendi.")
            except: pass
            await query.edit_message_text(f"✅ {name} engellendi.")
            log_admin_action(uid, "BLOCK_APPROVED", f"target={target} req_by={req_uid}")
            return ConversationHandler.END

    if cb.startswith("page|"):
        page = int(cb.split("|")[1])
        context.user_data["page"] = page
        await show_folder(query, context, path)
        return ConversationHandler.END

    # ── Navigasyon ────────────────────────────────────
    if cb.startswith("nav|"):
        action = cb.split("|")[1]
        # Geri gidince o klasörün dosyalarını sil
        if action == "back" and path:
            folder_key = "|".join(path)
            folder_msgs = context.user_data.get("folder_msgs", {})
            msg_ids = folder_msgs.pop(folder_key, [])
            for mid in msg_ids:
                try:
                    await context.bot.delete_message(chat_id=query.message.chat_id, message_id=mid)
                except: pass
            path.pop()
        elif action == "root":
            # Ana sayfaya dönünce tüm klasör mesajlarını temizle
            folder_msgs = context.user_data.get("folder_msgs", {})
            for folder_key, msg_ids in list(folder_msgs.items()):
                for mid in msg_ids:
                    try:
                        await context.bot.delete_message(chat_id=query.message.chat_id, message_id=mid)
                    except: pass
            context.user_data["folder_msgs"] = {}
            path = []

        context.user_data["path"] = path
        context.user_data["page"] = 0
        context.user_data.pop("action", None)  # ai_chat ve diğer modları temizle

        # Status mesajlarını sil (dosya/klasör ekleme sayacı)
        for key in ("add_file_status_id", "add_folder_status_id"):
            mid = context.user_data.pop(key, None)
            if mid:
                try:
                    await context.bot.delete_message(chat_id=query.message.chat_id, message_id=mid)
                except: pass
        for k in ("add_file_count", "add_folder_count"):
            context.user_data.pop(k, None)
        # action'ı temizle
        context.user_data.pop("action", None)

        # Mevcut mesajı (prompt) düzenleyerek ana sayfayı göster
        await show_folder(query, context, path)
        return ConversationHandler.END

    if cb.startswith("open|"):
        path.append(cb.split("|",1)[1])
        context.user_data["path"] = path
        context.user_data["page"] = 0
        content2 = load_content()
        folder2  = get_folder(content2, path)
        files    = folder2.get("files", [])

        # Önce klasör görünümünü göster
        await show_folder(query, context, path)

        # Dosyalar varsa otomatik gönder ve mesaj ID'lerini kaydet
        if files:
            sent_ids = []
            for f in files:
                try:
                    sent_msg = await _send_file_get_msg(query.message, f, uid)
                    if sent_msg:
                        sent_ids.append(sent_msg.message_id)
                    if not is_admin(uid):
                        log_user_message(user, "file_view", f.get("caption",""))
                except Exception as e:
                    logger.error(f"Dosya gönderme hatası: {e}")
            # Gönderilen mesaj ID'lerini path ile eşleştirerek sakla
            folder_key = "|".join(path)
            folder_msgs = context.user_data.setdefault("folder_msgs", {})
            folder_msgs[folder_key] = sent_ids
        return ConversationHandler.END

    # ── Yönetim (sadece süper admin) ──────────────────
    # Alt admin — kendi sınıf kullanıcılarını görme
    if cb == "mgmt|my_users" and is_admin(uid) and not is_main_admin(uid):
        if not get_admin_perm(uid, "can_view_users"):
            await query.answer("ليس لديك صلاحية", show_alert=True); return ConversationHandler.END
        adm_cls = get_admin_cls(uid)
        adm_grp = get_admin_grp(uid)
        users_d = load_users()
        grps_d  = load_groups()
        shfts   = load_shifts()
        kb = []
        for uid_, u in users_d.items():
            if int(uid_) == ADMIN_ID: continue
            if adm_cls and get_user_class(uid_) != adm_cls: continue
            if adm_grp and not grps_d.get(uid_,"").startswith(adm_grp): continue
            name  = u.get("full_name") or u.get("first_name") or f"ID:{uid_}"
            un    = f" @{u['username']}" if u.get("username") else ""
            grp_s = grps_d.get(uid_,"?")
            sft_s = "☀️" if shfts.get(uid_,"") in ("sabahi","sabah") else ("🌙" if shfts.get(uid_,"") in ("masaiy","gece") else "")
            kb.append([InlineKeyboardButton(f"👤 {name}{un} {sft_s}[{grp_s}]", callback_data=f"user|info|{uid_}")])
        if not kb:
            await query.answer("لا يوجد طلاب في فئتك", show_alert=True)
            return ConversationHandler.END
        try:
            await query.edit_message_text(f"👥 طلاب صفي ({len(kb)}):", reply_markup=InlineKeyboardMarkup(kb[:30]))
        except:
            await query.message.reply_text(f"👥 طلاب صفي ({len(kb)}):", reply_markup=InlineKeyboardMarkup(kb[:30]))
        return ConversationHandler.END

    if cb.startswith("mgmt|") and is_main_admin(uid):
        parts  = cb.split("|")
        action = parts[1]

        if action == "lab":
            # Mevcut lab kayıtları + yeni tarih ekleme
            schedule = load_lab_schedule()
            from datetime import date as _date
            today_str = _date.today().strftime("%d/%m/%Y")
            # Gelecekteki kayıtlar
            future = get_upcoming_lab_weeks(10)
            kb = []
            for e in future:
                grp  = e.get("group","?")
                week = e.get("week","")
                note = f" — {e['note']}" if e.get("note") else ""
                kb.append([
                    InlineKeyboardButton(f"🔬 {week}: {grp}{note}", callback_data=f"lab|view|{week}"),
                    InlineKeyboardButton("🗑", callback_data=f"lab|del|{week}"),
                ])
            kb.append([InlineKeyboardButton("➕ Yeni Tarih Ekle", callback_data="lab|add_new")])
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data="nav|root")])
            await query.edit_message_text(
                f"🔬 Lab Programı\n(Bugün: {today_str})",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "admin_log_export":
            log = load_admin_log()
            if not log:
                await query.answer("Log boş", show_alert=True); return ConversationHandler.END
            import io, csv as _csv
            bio = io.StringIO()
            w = _csv.writer(bio)
            w.writerow(["Zaman", "Admin", "Islem", "Detay"])
            for entry in log:
                adm_name = load_users().get(entry.get("uid",""),{}).get("first_name", entry.get("uid","?"))
                w.writerow([entry.get("time",""), adm_name, entry.get("action",""), entry.get("detail","")])
            b = io.BytesIO(bio.getvalue().encode("utf-8-sig"))
            b.seek(0)
            try:
                await context.bot.send_document(
                    int(uid),
                    document=b,
                    filename=f"admin_log_{datetime.now(IRAQ_TZ).strftime('%Y%m%d_%H%M')}.csv",
                    caption=f"📋 Admin Log — {len(log)} kayıt")
            except Exception as e:
                await query.answer(f"Hata: {e}", show_alert=True)
            return ConversationHandler.END

        if action == "stats":
            u_d   = load_users()
            m_d   = load_messages()
            s     = load_settings()
            vc    = load_view_counts()
            ct    = load_content()
            total_msg   = sum(len(v) for v in m_d.values())
            total_views = sum((v.get("count",0) if isinstance(v,dict) else 0) for v in vc.values())

            # Aktivite türlerine göre sayım
            type_counts = {}
            for msgs in m_d.values():
                for m in msgs:
                    t = m.get("type","?")
                    type_counts[t] = type_counts.get(t, 0) + 1

            # Klasör istatistikleri
            def _folder_stats(node, depth=0):
                lines = []
                for name, sub in node.get("folders",{}).items():
                    fc   = len(sub.get("files",[]))
                    sc   = len(sub.get("folders",{}))
                    indent = "  " * depth
                    lines.append(f"{indent}📁 {name}  ({sc}📁 {fc}📎)")
                    lines.extend(_folder_stats(sub, depth+1))
                return lines

            folder_lines = _folder_stats(ct)

            # En aktif kullanıcılar
            top_users = sorted(
                [(uid_, len(msgs)) for uid_, msgs in m_d.items()],
                key=lambda x: x[1], reverse=True)[:5]

            # En çok görüntülenen dosyalar
            top_files = sorted(
                [(v.get("name","?"), v.get("count",0)) for v in vc.values() if isinstance(v,dict)],
                key=lambda x: x[1], reverse=True)[:5]

            txt = (
                f"📊 DETAYLI İSTATİSTİKLER\n{'═'*28}\n\n"
                f"👥 Toplam kullanıcı: {len(u_d)}\n"
                f"💬 Toplam kayıt: {total_msg}\n"
                f"👁 Toplam görüntüleme: {total_views}\n"
                f"🔧 Bakım: {'Açık' if s['maintenance'] else 'Kapalı'}\n\n"
                f"\n"
                f"📨 Aktivite Dağılımı:\n"
                f"  ✍️ Mesaj: {type_counts.get('msg',0)}\n"
                f"  🔘 Buton: {type_counts.get('button',0)}\n"
                f"  🔍 Arama: {type_counts.get('search',0)}\n"
                f"  👁 Dosya görüntüleme: {type_counts.get('file_view',0)}\n"
                f"  🖼 Fotoğraf: {type_counts.get('photo',0)}\n"
                f"  📄 Dosya: {type_counts.get('document',0)}\n"
                f"  🎥 Video: {type_counts.get('video',0)}\n\n"
                f"\n"
                f"📁 Klasör Yapısı:\n" +
                ("\n".join(folder_lines) if folder_lines else "  Boş") +
                f"\n\n\n"
                f"🏆 En Aktif Kullanıcılar:\n"
            )
            users_data = load_users()
            for uid_, cnt in top_users:
                u = users_data.get(uid_, {})
                name = u.get("full_name") or u.get("first_name") or f"ID:{uid_}"
                txt += f"  👤 {name[:20]} — {cnt} kayıt\n"

            if top_files:
                txt += f"\n\n🔥 En Çok Görüntülenen:\n"
                for i,(fname,cnt) in enumerate(top_files,1):
                    txt += f"  {i}. {fname[:25]} — {cnt}×\n"

            kb = [
                [InlineKeyboardButton("📊 Kullanıcı Listesi", callback_data="mgmt|users")],
                [InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")],
            ]
            await query.edit_message_text(txt[:4000], reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "users":
            # Adım 1: Sınıf seç
            kb = [[InlineKeyboardButton("📋 Tümü", callback_data="mgmt|ulist|all|all|all")]]
            for cls_id, cls_def in CLASS_DEFS.items():
                cnt = len(users_by_class(cls_id))
                kb.append([InlineKeyboardButton(
                    f"{cls_def['tr']}  ({cnt})",
                    callback_data=f"mgmt|ulist|{cls_id}|_|_")])
            kb.append([InlineKeyboardButton("🔍 Ara", callback_data="mgmt|ulist|search")])
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data="nav|root")])
            await query.edit_message_text("👥 Kullanıcılar — Sınıf Seç:", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "ulist" and len(parts) > 2 and parts[2] == "search":
            context.user_data["action"] = "user_search_admin"
            await query.edit_message_text(
                "🔍 İsim veya @username yazın:",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("◀️ İptal", callback_data="mgmt|users")]]))
            return WAIT_FOLDER

        if action == "ulist":
            # parts: mgmt|ulist|cls|sft|grp
            # _ = henüz seçilmedi
            cls_f = parts[2] if len(parts) > 2 else "all"
            sft_f = parts[3] if len(parts) > 3 else "_"
            grp_f = parts[4] if len(parts) > 4 else "_"

            u_d    = load_users()
            grps_d = load_groups()
            shfts  = load_shifts()

            # Adım 1: Sınıf seçildi, vardiya seç
            if cls_f not in ("all","search") and sft_f == "_":
                cls_name = CLASS_DEFS.get(cls_f,{}).get("tr","?")
                u_cls = [u for u in u_d if get_user_class(u)==cls_f and int(u)!=ADMIN_ID]
                s_cnt = sum(1 for u in u_cls if shfts.get(u,"") in ("sabahi","sabah"))
                g_cnt = sum(1 for u in u_cls if shfts.get(u,"") in ("masaiy","gece"))
                kb = [
                    [InlineKeyboardButton(f"☀️ Sabahci ({s_cnt})", callback_data=f"mgmt|ulist|{cls_f}|sabahi|_"),
                     InlineKeyboardButton(f"🌙 Gececi ({g_cnt})",  callback_data=f"mgmt|ulist|{cls_f}|masaiy|_")],
                    [InlineKeyboardButton(f"Tümü ({len(u_cls)})",   callback_data=f"mgmt|ulist|{cls_f}|all|_")],
                    [InlineKeyboardButton("◀️ Geri", callback_data="mgmt|users")],
                ]
                await query.edit_message_text(f"👥 {cls_name} — Vardiya Seç:", reply_markup=InlineKeyboardMarkup(kb))
                return ConversationHandler.END

            # Adım 2: Vardiya seçildi, grup seç
            if cls_f not in ("all","search") and sft_f != "_" and grp_f == "_":
                cls_name = CLASS_DEFS.get(cls_f,{}).get("tr","?")
                cls_groups = get_class_groups(cls_f)
                u_filtered = []
                for u in u_d:
                    if int(u) == ADMIN_ID: continue
                    if get_user_class(u) != cls_f: continue
                    if sft_f != "all":
                        us = shfts.get(u,"")
                        if sft_f == "sabahi" and us not in ("sabahi","sabah"): continue
                        if sft_f == "masaiy" and us not in ("masaiy","gece"): continue
                    u_filtered.append(u)
                sft_lbl = "☀️ Sabah" if sft_f=="sabahi" else ("🌙 Gece" if sft_f=="masaiy" else "Tümü")
                kb = []
                for grp in cls_groups:
                    cnt = sum(1 for u in u_filtered if grps_d.get(u,"").startswith(grp))
                    kb.append([InlineKeyboardButton(f"👥 Grup {grp} ({cnt})", callback_data=f"mgmt|ulist|{cls_f}|{sft_f}|{grp}")])
                kb.append([InlineKeyboardButton(f"Tümü ({len(u_filtered)})", callback_data=f"mgmt|ulist|{cls_f}|{sft_f}|all")])
                kb.append([InlineKeyboardButton("◀️ Geri", callback_data=f"mgmt|ulist|{cls_f}|_|_")])
                await query.edit_message_text(f"👥 {cls_name} / {sft_lbl} — Grup Seç:", reply_markup=InlineKeyboardMarkup(kb))
                return ConversationHandler.END

            # Adım 3: Liste göster
            result = []
            for uid_, u in u_d.items():
                if int(uid_) == ADMIN_ID: continue
                if cls_f not in ("all","search"):
                    if get_user_class(uid_) != cls_f: continue
                if sft_f not in ("_","all"):
                    us = shfts.get(uid_,"")
                    if sft_f == "sabahi" and us not in ("sabahi","sabah"): continue
                    if sft_f == "masaiy" and us not in ("masaiy","gece"): continue
                if grp_f not in ("_","all"):
                    if not grps_d.get(uid_,"").startswith(grp_f): continue
                result.append((uid_, u))

            if not result:
                await query.answer("Bu filtrede kullanıcı yok", show_alert=True)
                return ConversationHandler.END

            kb = []
            for uid_, u in result[:40]:
                name  = u.get("full_name") or u.get("first_name") or f"ID:{uid_}"
                un    = f" @{u['username']}" if u.get("username") else ""
                grp_s = grps_d.get(uid_,"?")
                sft_s = "☀️" if shfts.get(uid_,"") in ("sabahi","sabah") else ("🌙" if shfts.get(uid_,"") in ("masaiy","gece") else "")
                cls_s = get_user_class(uid_) or "?"
                kb.append([InlineKeyboardButton(
                    f"👤 {name}{un} {sft_s}[{cls_s}/{grp_s}]",
                    callback_data=f"user|info|{uid_}")])
            back_cb = f"mgmt|ulist|{cls_f}|{sft_f}|_" if grp_f not in ("_","all") else f"mgmt|ulist|{cls_f}|_|_" if sft_f != "_" else "mgmt|users"
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data=back_cb)])
            await query.edit_message_text(
                f"👥 {len(result)} kullanıcı:",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "add_admin":
            # Önce sınıf seç
            kb = [
                [InlineKeyboardButton("Hepsi (kisitsiz)", callback_data="newadmin_pre|cls|all")],
                [InlineKeyboardButton("1.Sinif", callback_data="newadmin_pre|cls|1"),
                 InlineKeyboardButton("2.Sinif", callback_data="newadmin_pre|cls|2"),
                 InlineKeyboardButton("3.Sinif", callback_data="newadmin_pre|cls|3"),
                 InlineKeyboardButton("4.Sinif", callback_data="newadmin_pre|cls|4")],
                [InlineKeyboardButton("İptal", callback_data="close")],
            ]
            await query.edit_message_text("Admin ekle — Önce sınıf seç:", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "del_admin":
            admins = load_admins()
            if not admins:
                await query.answer(TR["no_admins"], show_alert=True); return ConversationHandler.END
            kb = [[InlineKeyboardButton(f"🚫 {a}", callback_data=f"rem_admin|{a}")] for a in admins]
            kb.append([InlineKeyboardButton(TR["cancel"], callback_data="close")])
            await query.edit_message_text(TR["del_admin_lbl"], reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "dm_user":
            u_d = load_users(); kb = []
            for uid_, u in list(u_d.items())[-30:]:
                if int(uid_) == ADMIN_ID: continue
                name = u.get("full_name") or u.get("first_name") or f"ID:{uid_}"
                un   = f" @{u['username']}" if u.get("username") else ""
                kb.append([InlineKeyboardButton(f"👤 {name}{un}", callback_data=f"dm_pick|{uid_}")])
            kb.append([InlineKeyboardButton(TR["dm_manual"], callback_data="dm_pick|manual")])
            kb.append([InlineKeyboardButton(TR["cancel"],    callback_data="close")])
            await query.edit_message_text(TR["dm_select"], reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "broadcast":
            context.user_data["action"] = "broadcast"
            await query.edit_message_text(TR["broadcast_enter"],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_BROADCAST_MSG

        if action == "poll":
            polls = load_polls()
            kb = [[InlineKeyboardButton(TR["poll_create"], callback_data="poll|create")]]
            if polls:
                kb.append([InlineKeyboardButton(TR["poll_results"], callback_data="poll|list_results"),
                           InlineKeyboardButton(TR["poll_delete"],  callback_data="poll|list_delete")])
            kb.append([InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")])
            active = sum(1 for p in polls.values() if p.get("active"))
            txt = f"{TR['poll_panel']}\n\n📊 Toplam anket: {len(polls)}\n✅ Aktif: {active}"
            await query.edit_message_text(txt, reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

    if cb.startswith("rem_admin|") and is_main_admin(uid):
        target = int(cb.split("|")[1])
        admins = load_admins()
        if target in admins:
            admins.remove(target); save_admins(admins)
            log_admin_action(uid, "REMOVE_ADMIN", f"ID:{target} cikartildi")
        await query.answer(TR["admin_removed"].format(target), show_alert=True)
        if admins:
            kb = [[InlineKeyboardButton(f"🚫 {a}", callback_data=f"rem_admin|{a}")] for a in admins]
            kb.append([InlineKeyboardButton(TR["cancel"], callback_data="close")])
            await query.edit_message_text(TR["del_admin_lbl"], reply_markup=InlineKeyboardMarkup(kb))
        else:
            await query.edit_message_text(TR["no_admins"],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")]]))
        return ConversationHandler.END

    if cb.startswith("dm_pick|") and is_main_admin(uid):
        target = cb.split("|")[1]
        if target == "manual":
            context.user_data["action"] = "dm_manual_id"
            await query.edit_message_text(TR["dm_enter_id"],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_DM_USER_ID
        context.user_data["dm_target"] = target; context.user_data["action"] = "dm_msg"
        u    = load_users().get(target,{}); name = u.get("full_name") or u.get("first_name") or f"ID:{target}"
        await query.edit_message_text(TR["dm_enter_msg"].format(name),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
        return WAIT_DM_MSG

    if cb.startswith("dm_quick|") and is_admin(uid):
        target = cb.split("|")[1]
        context.user_data["dm_target"] = target; context.user_data["action"] = "dm_msg"
        u    = load_users().get(target,{}); name = u.get("full_name") or u.get("first_name") or f"ID:{target}"
        try:
            await query.edit_message_text(TR["dm_enter_msg"].format(name),
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
        except Exception:
            await context.bot.send_message(uid, TR["dm_enter_msg"].format(name))
        return WAIT_DM_MSG

    # ── Kullanıcı detay (süper admin) ────────────────
    if cb.startswith("user|") and is_admin(uid):
        parts  = cb.split("|")
        action = parts[1]

        # Alt admin yetki kontrolü
        if not is_main_admin(uid):
            # Profil görme yetkisi
            view_actions = {"info", "extra_cls", "toggle_xcls"}
            if action in view_actions and not get_admin_perm(uid, "can_view_users"):
                await query.answer("فعّل صلاحية رؤية المستخدمين أولاً", show_alert=True)
                return ConversationHandler.END
            # Uyarı yetkisi
            if action == "warn" and not get_admin_perm(uid, "can_warn"):
                await query.answer("ليس لديك صلاحية التحذير", show_alert=True)
                return ConversationHandler.END
            # Engel yetkisi
            if action == "block" and not get_admin_perm(uid, "can_block"):
                await query.answer("ليس لديك صلاحية الحظر", show_alert=True)
                return ConversationHandler.END
            # Sadece süper admin
            main_only = {"export", "change_cls", "change_grp", "set_cls", "set_grp",
                         "ulist", "dm"}
            if action in main_only:
                await query.answer("ليس لديك هذه الصلاحية", show_alert=True)
                return ConversationHandler.END

        if action == "info":
            target = parts[2]; u_d = load_users(); m_d = load_messages()
            u = u_d.get(target,{}); msgs = m_d.get(target,[])
            name  = u.get("full_name") or u.get("first_name") or "—"
            un    = f"@{u['username']}" if u.get("username") else "—"
            last  = u.get("last_seen","—")
            blkd  = is_blocked(target)
            total = len(msgs)
            cls   = class_label(target)
            warns_list = get_warns(target)
            note  = get_note(target)

            type_counts = {}
            for m in msgs:
                t = m.get("type","?")
                type_counts[t] = type_counts.get(t, 0) + 1

            searches = [m.get("content","") for m in msgs if m.get("type") == "search"][-5:]
            viewed   = [m.get("content","") for m in msgs if m.get("type") == "file_view"]
            from collections import Counter
            top_viewed = Counter(viewed).most_common(3)
            media_cnt = sum(1 for m in msgs if m.get("type") in ("photo","video","document","audio","voice","animation"))

            ICONS = {"msg":"✍️","photo":"🖼","video":"🎥","document":"📄",
                     "audio":"🎵","voice":"🎙","animation":"🎞","sticker":"😊",
                     "command":"⚙️","file_view":"👁","button":"🔘","search":"🔍"}

            recent = []
            for m in msgs[-10:]:
                icon = ICONS.get(m.get("type",""),"📨")
                t    = m.get("time","")[-8:]
                c    = m.get("content","")[:40]
                recent.append(f"  {icon} {t}  {c}")

            info = (
                f"╔══════════════════════════╗\n"
                f"  👤  {name}\n"
                f"  🔗  {un}\n"
                f"  🆔  {target}\n"
                f"  🎓  {TR['class_label'].format(cls)}\n"
                f"  📅  {last}\n"
                f"  🚫  Engel: {'✅' if blkd else '❌'}\n"
                f"  ⚠️   Uyarı: {len(warns_list)}/{MAX_WARNS}\n"
            )
            if note:
                info += f"  📝  Not: {note[:60]}\n"
            info += (
                f"╠══════════════════════════╣\n"
                f"  📊  Toplam: {total}\n"
                f"  ✍️   Mesaj: {type_counts.get('msg',0)}\n"
                f"  🔍  Arama: {type_counts.get('search',0)}\n"
                f"  👁   Görüntüleme: {type_counts.get('file_view',0)}\n"
                f"  🖼   Medya: {media_cnt}\n"
            )
            if searches:
                info += f"╠══════════════════════════╣\n  🔍  Son Aramalar:\n"
                for s_ in searches: info += f"     • {s_[:35]}\n"
            if top_viewed:
                info += f"  👁   Çok Açılan:\n"
                for fname, cnt in top_viewed: info += f"     • {fname[:30]} ({cnt}×)\n"
            if recent:
                info += f"╠══════════════════════════╣\n  🕐  Son Aktivite:\n"
                for line in recent[-5:]: info += f"{line}\n"
            info += "╚══════════════════════════╝"

            media_files = [m for m in msgs if m.get("file_id") and
                           m["type"] in ("photo","video","document","audio","voice","animation")]
            kb = [
                [InlineKeyboardButton(TR["unblock_btn"] if blkd else TR["block_btn"],
                                      callback_data=f"user|{'unblock' if blkd else 'block'}|{target}"),
                 InlineKeyboardButton(TR["msg_btn"], callback_data=f"dm_quick|{target}")],
                [InlineKeyboardButton(TR["warn_btn"],    callback_data=f"user|warn|{target}"),
                 InlineKeyboardButton(TR["note_btn"],    callback_data=f"user|note|{target}")],
            ]
            if warns_list:
                kb.append([InlineKeyboardButton(TR["warn_clear_btn"], callback_data=f"user|clear_warns|{target}")])
            if media_files:
                kb.insert(0,[InlineKeyboardButton(TR["send_media_btn"].format(len(media_files)),
                                                  callback_data=f"user|sendmedia|{target}")])
            if is_main_admin(uid):
                kb.append([InlineKeyboardButton("🎓 Sınıf Değiştir", callback_data=f"user|change_cls|{target}"),
                            InlineKeyboardButton("👥 Grup Değiştir",  callback_data=f"user|change_grp|{target}")])
                kb.append([InlineKeyboardButton("➕ Ek Sınav Sınıfı", callback_data=f"user|extra_cls|{target}")])
            kb.append([InlineKeyboardButton("📥 Aktivite İndir", callback_data=f"user|export|{target}")])
            kb.append([InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")])
            # Alt admin — sadece isim + kullanıcıadı + 2 buton
            if not is_main_admin(uid):
                simple_info = (
                    f"👤 {name}\n"
                    f"🔗 {un}\n"
                    f"⚠️ {len(warns_list)}/{MAX_WARNS} تحذيرات"
                    + ("\n🚫 محظور" if blkd else "")
                )
                simple_kb = []
                if get_admin_perm(uid, "can_warn") and not blkd:
                    simple_kb.append(InlineKeyboardButton("⚠️ تحذير", callback_data=f"user|warn|{target}"))
                if get_admin_perm(uid, "can_block"):
                    if blkd:
                        simple_kb.append(InlineKeyboardButton("🔓 رفع الحظر", callback_data=f"user|unblock|{target}"))
                    else:
                        simple_kb.append(InlineKeyboardButton("🚫 حظر", callback_data=f"user|block|{target}"))
                row_btns = [simple_kb] if simple_kb else []
                if warns_list and get_admin_perm(uid, "can_warn"):
                    row_btns.append([InlineKeyboardButton("🗑 مسح التحذيرات", callback_data=f"user|clear_warns|{target}")])
                row_btns.append([InlineKeyboardButton("◀️ رجوع", callback_data="mgmt|my_users")])
                await query.edit_message_text(simple_info, reply_markup=InlineKeyboardMarkup(row_btns))
                return ConversationHandler.END

            await query.edit_message_text(info[:4000], reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "export":
            target = parts[2]; m_d = load_messages(); u_d = load_users()
            msgs = m_d.get(target, [])
            u    = u_d.get(target, {})

            def clean(s):
                """Metni koru — Arapça, Türkçe, emoji dahil."""
                if not s: return "-"
                try:
                    val = str(s).strip()
                    return val if val else "-"
                except:
                    return str(s)[:500] if s else "-"
            to_ascii = clean

            name     = clean(u.get("full_name") or u.get("first_name") or "")
            username = clean(u.get("username") or "")
            cls_str  = clean(class_label(target))
            grp_str  = clean(get_user_group(target))

            TYPE_MAP = {
                "msg":       "Mesaj",
                "photo":     "Fotograf",
                "video":     "Video",
                "document":  "Dosya",
                "voice":     "Ses Mesaji",
                "animation": "GIF",
                "sticker":   "Sticker",
                "button":    "Buton",
                "search":    "Arama",
                "file_view": "Dosya Goruntuleme",
                "folder":    "Klasor Girisi",
                "command":   "Komut",
                "media":     "Medya",
                "note":      "Not",
            }

            SEP = "=" * 44
            sep = "-" * 44

            lines = [
                SEP,
                "  KULLANICI AKTiViTE RAPORU",
                SEP,
                f"  Ad       : {name}",
                f"  Kullanici: @{username}" if username != "-" else "",
                f"  ID       : {target}",
                f"  Sinif    : {cls_str}",
                f"  Grup     : {grp_str}",
                f"  Son giris: {u.get('last_seen','-')}",
                f"  Rapor    : {datetime.now(IRAQ_TZ).strftime('%Y-%m-%d %H:%M')}",
                SEP,
                "",
                "  OZET",
                sep,
                f"  Toplam aktivite  : {len(msgs)}",
            ]

            counts = {}
            for m in msgs:
                t = m.get("type", "diger")
                counts[t] = counts.get(t, 0) + 1
            for t, c in sorted(counts.items(), key=lambda x: -x[1]):
                lbl = TYPE_MAP.get(t, t)
                lines.append(f"  {lbl:<22}: {c}")

            lines += [
                "",
                sep,
                "  TUM AKTiViTE GECMiSi",
                sep,
                f"  {'Tarih-Saat':<22} {'Tur':<18} Icerik",
                sep,
            ]

            for m in msgs:
                t     = m.get("type", "")
                lbl   = TYPE_MAP.get(t, t).strip()
                zaman = m.get("time", "")
                cont  = clean(m.get("content", ""))
                fid   = m.get("file_id", "")
                if not cont or cont == "-":
                    cont = f"[{lbl}]"
                line = f"  {zaman}  [{lbl}]  {cont[:120]}"
                if fid:
                    line += f"  (file_id: {fid[:20]}...)"
                lines.append(line)

            # Hatırlatıcılar
            rems = get_user_reminders(target)
            if rems:
                lines += ["", sep, f"  HATIRLATICILARI  ({len(rems)} adet)", sep]
                import time as _t
                now_ts = _t.time()
                for i, r in enumerate(rems, 1):
                    rem_txt  = to_ascii(r.get("text", ""))
                    created  = r.get("created", "")
                    dur_min  = int((r.get("fire_ts", 0) - r.get("created_ts", 0)) / 60) if r.get("created_ts") else 0
                    left_min = max(0, int((r.get("fire_ts", 0) - now_ts) / 60))
                    if left_min >= 1440:
                        left_str = f"{left_min//1440} gun {(left_min%1440)//60} sa kaldi"
                    elif left_min >= 60:
                        left_str = f"{left_min//60} saat {left_min%60} dk kaldi"
                    else:
                        left_str = f"{left_min} dakika kaldi"
                    lines.append(f"  {i}. [{created}] {rem_txt}")
                    lines.append(f"     Kalan: {left_str}")

            # Notlar
            notes = get_user_notes(target)
            if notes:
                lines += ["", sep, f"  NOTLARI  ({len(notes)} adet)", sep]
                for i, n in enumerate(notes, 1):
                    subj  = to_ascii(n.get("subject", ""))
                    cont  = to_ascii(n.get("content", ""))
                    ntype = n.get("type", "text")
                    zaman = n.get("time", "")
                    fid_n = n.get("file_id", "")
                    lines.append(f"  {i}. [{zaman}] Ders:{subj}  Tur:{ntype}")
                    if cont and cont != "-":
                        lines.append(f"     Icerik: {cont[:200]}")
                    if fid_n:
                        lines.append(f"     Medya: {fid_n[:40]}...")

            # ── Excel dosyası oluştur ──────────────────
            import io
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                HAS_OPENPYXL = True
            except ImportError:
                HAS_OPENPYXL = False

            if not HAS_OPENPYXL:
                wb = None
            else:
                wb = openpyxl.Workbook()

            # ── Excel veya CSV ─────────────────────────
            if not HAS_OPENPYXL:
                # CSV fallback — kullanıcı bilgileri + aktiviteler
                shift_raw  = get_user_shift(target)
                shift_disp = "Sabahci" if shift_raw in ("sabahi","sabah") else ("Gececi" if shift_raw in ("masaiy","gece") else "-")
                import csv as _csv
                csv_bio = io.StringIO()
                w = _csv.writer(csv_bio)
                # Bilgi kutusu
                w.writerow(["=== KULLANICI BILGISI ==="])
                w.writerow(["Ad",          name])
                w.writerow(["Kullanici Adi", u.get("username","-") or "-"])
                w.writerow(["ID",           target])
                w.writerow(["Sinif",        cls_str or "-"])
                w.writerow(["Vardiya",      shift_disp])
                w.writerow(["Grup",         grp_str or "-"])
                w.writerow(["Kayit Tarihi", u.get("join_date", u.get("last_seen","-"))[:10] if u.get("join_date") or u.get("last_seen") else "-"])
                w.writerow(["Son Kullanim", u.get("last_seen","-")])
                w.writerow(["Toplam Aktivite", len(msgs)])
                w.writerow(["Rapor Tarihi",  datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M")])
                w.writerow([])
                w.writerow(["=== AKTIVITE GECMISI ==="])
                w.writerow(["Tarih-Saat", "Tur", "Icerik"])
                for m in msgs:
                    w.writerow([m.get("time",""), TYPE_MAP.get(m.get("type",""),"").strip(), m.get("content","")[:200]])
                b2 = io.BytesIO(csv_bio.getvalue().encode("utf-8-sig"))
                b2.seek(0)
                fn2 = f"rapor_{target}_{datetime.now(IRAQ_TZ).strftime('%Y%m%d_%H%M')}.csv"
                try:
                    await query.message.reply_document(b2, filename=fn2, caption=f"{name} | {len(msgs)} aktivite")
                except Exception as e:
                    await query.message.reply_text(f"Hata: {e}")
                return ConversationHandler.END

            # ── Sayfa 1: Özet ──────────────────────────
            ws1 = wb.active
            ws1.title = "Ozet"
            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            hdr_fill  = PatternFill("solid", fgColor="1F4E79")
            alt_fill  = PatternFill("solid", fgColor="D6E4F0")
            box_fill  = PatternFill("solid", fgColor="EBF3FB")
            box_font  = Font(bold=True, size=11, color="1F4E79")

            # ── Kullanıcı Bilgileri Kutusu ──────────────
            shift_raw = get_user_shift(target)
            shift_disp = "صباحي" if shift_raw in ("sabahi","sabah") else ("مسائي" if shift_raw in ("masaiy","gece") else "-")
            grp_disp   = grp_str or "-"
            join_date  = u.get("join_date", u.get("last_seen", "-"))[:10] if u.get("join_date") or u.get("last_seen") else "-"
            last_seen  = u.get("last_seen", "-")
            username_val = u.get("username", "-") or "-"

            box_rows = [
                ["╔══════════════════════════════╗"],
                ["  بيانات المستخدم / KULLANICI BILGISI"],
                ["╠══════════════════════════════╣"],
                ["  الاسم / Ad",          name],
                ["  اسم المستخدم",         f"@{username_val}"],
                ["  المعرف / ID",          target],
                ["  السنة / Sinif",        cls_str or "-"],
                ["  الفترة / Vardiya",     shift_disp],
                ["  المجموعة / Grup",      grp_disp],
                ["  تاريخ الانضمام",       join_date],
                ["  آخر استخدام",          last_seen],
                ["  إجمالي النشاط",        len(msgs)],
                ["  تاريخ التقرير",        datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M")],
                ["╚══════════════════════════════╝"],
            ]
            for br in box_rows:
                ws1.append(br)
                r = ws1.max_row
                for col in range(1, 3):
                    ws1.cell(r, col).fill = box_fill
                    ws1.cell(r, col).font = box_font if br[0].startswith("╠") or br[0].startswith("╔") or br[0].startswith("╚") or br[0].startswith("  بيانات") else Font(size=10)
                    ws1.cell(r, col).alignment = Alignment(horizontal="right")

            ws1.append([])
            info_rows = []  # artık kutu kullandık

            ws1.append([])
            ws1.append(["TUR", "ADET"])
            ws1["A" + str(ws1.max_row)].font = hdr_font
            ws1["A" + str(ws1.max_row)].fill = hdr_fill
            ws1["B" + str(ws1.max_row)].font = hdr_font
            ws1["B" + str(ws1.max_row)].fill = hdr_fill
            counts = {}
            for m in msgs:
                t = TYPE_MAP.get(m.get("type",""), m.get("type","?")).strip()
                counts[t] = counts.get(t, 0) + 1
            for t, cnt in sorted(counts.items(), key=lambda x: -x[1]):
                ws1.append([t, cnt])

            ws1.column_dimensions["A"].width = 22
            ws1.column_dimensions["B"].width = 40

            # ── Sayfa 2: Tüm Aktiviteler ────────────────
            ws2 = wb.create_sheet("Aktiviteler")
            headers = ["Tarih-Saat", "Tur", "Icerik"]
            ws2.append(headers)
            for col, h in enumerate(headers, 1):
                cell = ws2.cell(1, col, h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")
            for i, m in enumerate(msgs, 2):
                t    = TYPE_MAP.get(m.get("type",""), m.get("type","?")).strip()
                cont = clean(m.get("content",""))
                if not cont or cont == "-":
                    cont = f"[{t}]"
                ws2.append([m.get("time",""), t, cont[:200]])
                if i % 2 == 0:
                    for col in range(1, 4):
                        ws2.cell(i, col).fill = alt_fill
            ws2.column_dimensions["A"].width = 22
            ws2.column_dimensions["B"].width = 18
            ws2.column_dimensions["C"].width = 60

            # ── Sayfa 3: Hatırlatıcılar ──────────────────
            rems_data = get_user_reminders(target)
            if rems_data:
                ws3 = wb.create_sheet("Hatirlaticlar")
                ws3.append(["Metin", "Olusturma", "Ates Zamani", "Kalan (dk)"])
                for col in range(1, 5):
                    ws3.cell(1, col).font = hdr_font
                    ws3.cell(1, col).fill = hdr_fill
                import time as _tx
                now_ts = _tx.time()
                for r in rems_data:
                    left = max(0, int((r.get("fire_ts",0) - now_ts) / 60))
                    ws3.append([
                        clean(r.get("text","")),
                        r.get("created",""),
                        datetime.fromtimestamp(r.get("fire_ts",0)).strftime("%Y-%m-%d %H:%M:%S") if r.get("fire_ts") else "-",
                        left
                    ])
                for col in [1,2,3,4]:
                    ws3.column_dimensions[get_column_letter(col)].width = 25

            # ── Sayfa 4: Notlar ──────────────────────────
            notes_data = get_user_notes(target)
            if notes_data:
                ws4 = wb.create_sheet("Notlar")
                ws4.append(["Ders", "Tur", "Icerik", "Zaman", "MediaID"])
                for col in range(1, 5):
                    ws4.cell(1, col).font = hdr_font
                    ws4.cell(1, col).fill = hdr_fill
                for n in notes_data:
                    cont_n = clean(n.get("content",""))
                    fid_n  = n.get("file_id","")
                    ws4.append([
                        clean(n.get("subject","")),
                        n.get("type","text"),
                        cont_n[:200] if cont_n and cont_n != "-" else f"[{n.get('type','medya')}]",
                        n.get("time",""),
                        fid_n[:30] + "..." if fid_n else ""
                    ])
                for col in [1,2,3,4]:
                    ws4.column_dimensions[get_column_letter(col)].width = 25

            if not HAS_OPENPYXL:
                # openpyxl yoksa düz metin
                import csv
                bio = io.StringIO()
                writer = csv.writer(bio)
                writer.writerow(["Tarih", "Tur", "Icerik"])
                for m in msgs:
                    writer.writerow([m.get("time",""), TYPE_MAP.get(m.get("type",""),"").strip(), m.get("content","")[:200]])
                bio2 = io.BytesIO(bio.getvalue().encode("utf-8-sig"))
                bio2.seek(0)
                fname2 = f"rapor_{target}_{datetime.now(IRAQ_TZ).strftime('%Y%m%d_%H%M')}.csv"
                try:
                    await query.message.reply_document(bio2, filename=fname2, caption=f"{name} | {len(msgs)} aktivite")
                except Exception as e:
                    await query.message.reply_text(f"Hata: {e}")
                return ConversationHandler.END

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            fname = f"rapor_{target}_{datetime.now(IRAQ_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
            try:
                await query.message.reply_document(
                    bio,
                    filename=fname,
                    caption=f"{name} | ID:{target} | {len(msgs)} aktivite")
            except Exception as e:
                await query.message.reply_text(f"Hata: {e}")
            return ConversationHandler.END

        if action == "extra_cls":
            # Ek sınav takibi — kullanıcı başka sınıfın sınavlarını da görsün
            target = parts[2]
            u_e = load_users().get(target,{})
            name_e = u_e.get("full_name") or u_e.get("first_name") or target
            # Mevcut ek sınıflar
            extra_cls_data = load_json(os.path.join(BASE_DIR, "user_extra_cls.json"), {})
            current_extra = extra_cls_data.get(str(target), [])
            kb = []
            for cls_id, cls_def in CLASS_DEFS.items():
                icon = "✅" if cls_id in current_extra else "➕"
                kb.append([InlineKeyboardButton(
                    f"{icon} {cls_def['tr']}",
                    callback_data=f"user|toggle_xcls|{target}|{cls_id}")])
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data=f"user|info|{target}")])
            await query.edit_message_text(
                f"📚 Ek Sınav Sınıfları\n{name_e}\n\n✅ = ekli, ➕ = ekle/çıkar:",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "toggle_xcls":
            target = parts[2]; cls_id = parts[3]
            extra_cls_data = load_json(os.path.join(BASE_DIR, "user_extra_cls.json"), {})
            cur = extra_cls_data.get(str(target), [])
            if cls_id in cur: cur.remove(cls_id)
            else: cur.append(cls_id)
            extra_cls_data[str(target)] = cur
            save_json(os.path.join(BASE_DIR, "user_extra_cls.json"), extra_cls_data)
            await query.answer(f"✅ {'Eklendi' if cls_id in cur else 'Çıkarıldı'}", show_alert=False)
            query.data = f"user|extra_cls|{target}"; return await callback(update, context)

        if action == "change_cls":
            # Tek kullanıcı sınıf değiştir
            target = parts[2]
            u_c = load_users().get(target,{})
            name_c = u_c.get("full_name") or u_c.get("first_name") or target
            kb = [[InlineKeyboardButton(CLASS_DEFS[cls]["ar"], callback_data=f"user|set_cls|{target}|{cls}")]
                  for cls in CLASS_DEFS]
            kb.append([InlineKeyboardButton("◀️", callback_data=f"user|info|{target}")])
            await query.edit_message_text(f"🎓 Sınıf değiştir: {name_c}", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "set_cls":
            target = parts[2]; new_cls = parts[3]
            old_cls = get_user_class(target)
            set_user_class(target, new_cls)
            log_admin_action(uid, "CHANGE_CLS", f"user:{target} {old_cls}→{new_cls}")
            await query.answer(f"✅ {CLASS_DEFS[new_cls]['ar']}", show_alert=True)
            query.data = f"user|info|{target}"; return await callback(update, context)

        if action == "change_grp":
            target = parts[2]
            kb = [
                [InlineKeyboardButton("A", callback_data=f"user|set_grp|{target}|A"),
                 InlineKeyboardButton("B", callback_data=f"user|set_grp|{target}|B"),
                 InlineKeyboardButton("C", callback_data=f"user|set_grp|{target}|C")],
                [InlineKeyboardButton("A1",callback_data=f"user|set_grp|{target}|A1"),
                 InlineKeyboardButton("A2",callback_data=f"user|set_grp|{target}|A2"),
                 InlineKeyboardButton("A3",callback_data=f"user|set_grp|{target}|A3")],
                [InlineKeyboardButton("B1",callback_data=f"user|set_grp|{target}|B1"),
                 InlineKeyboardButton("B2",callback_data=f"user|set_grp|{target}|B2"),
                 InlineKeyboardButton("B3",callback_data=f"user|set_grp|{target}|B3")],
                [InlineKeyboardButton("C1",callback_data=f"user|set_grp|{target}|C1"),
                 InlineKeyboardButton("C2",callback_data=f"user|set_grp|{target}|C2"),
                 InlineKeyboardButton("C3",callback_data=f"user|set_grp|{target}|C3")],
                [InlineKeyboardButton("◀️", callback_data=f"user|info|{target}")],
            ]
            await query.edit_message_text("👥 Grup değiştir:", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "set_grp":
            target = parts[2]; new_grp = parts[3]
            old_grp = get_user_group(target)
            set_user_group(target, new_grp)
            log_admin_action(uid, "CHANGE_GRP", f"user:{target} {old_grp}→{new_grp}")
            await query.answer(f"✅ {new_grp}", show_alert=True)
            query.data = f"user|info|{target}"; return await callback(update, context)

        if action == "warn":
            target = parts[2]
            u    = load_users().get(target,{}); name = u.get("full_name") or u.get("first_name") or f"ID:{target}"
            if not is_main_admin(uid):
                if not get_admin_perm(uid, "can_warn"):
                    await query.answer("Yetkin yok", show_alert=True); return ConversationHandler.END
                if not admin_can_manage_user(uid, target):
                    await query.answer("Bu kullaniciya yetkini yok", show_alert=True); return ConversationHandler.END
                # Sebep sor, sonra onaya gönder
                context.user_data["pending_action"] = "warn"
                context.user_data["pending_target"] = target
                context.user_data["pending_tname"]  = name
                context.user_data["action"] = "pending_reason"
                await query.edit_message_text(
                    f"⚠️ Uyari sebebini yaz:\n{name}",
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("Iptal", callback_data="nav|root")
                    ]]))
                return WAIT_FOLDER
            context.user_data["warn_target"] = target
            context.user_data["action"]      = "warn_reason"
            await query.edit_message_text(
                f"{TR['warn_reason_prompt']}\n\n👤 {name}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_WARN_REASON

        if action == "note":
            target = parts[2]
            context.user_data["note_target"] = target
            context.user_data["action"]      = "user_note"
            existing = get_note(target)
            prompt   = TR["note_prompt"]
            if existing:
                prompt += f"\n\nMevcut: {existing[:80]}"
            await query.edit_message_text(prompt,
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_USER_NOTE

        if action == "clear_warns":
            target = parts[2]
            clear_warns(target)
            await query.answer(TR["warns_cleared"].format(target), show_alert=True)
            query.data = f"user|info|{target}"; return await callback(update, context)

        if action == "sendmedia":
            target = parts[2]; m_d = load_messages()
            msgs   = m_d.get(target,[])
            media  = [m for m in msgs if m.get("file_id") and
                      m["type"] in ("photo","video","document","audio","voice","animation")]
            sent = fail = 0
            for m in media[-30:]:
                ftype=m["type"]; fid=m["file_id"]; cap=m.get("content","")
                try:
                    if   ftype=="photo":     await query.message.reply_photo(fid, caption=cap)
                    elif ftype=="video":     await query.message.reply_video(fid, caption=cap)
                    elif ftype=="document":  await query.message.reply_document(fid, caption=cap)
                    elif ftype=="audio":     await query.message.reply_audio(fid, caption=cap)
                    elif ftype=="voice":     await query.message.reply_voice(fid)
                    elif ftype=="animation": await query.message.reply_animation(fid, caption=cap)
                    sent += 1
                except Exception as e: logger.warning(f"Medya: {e}")
            await query.message.reply_text(TR["media_sent"].format(sent, len(media[-30:])))
            return ConversationHandler.END

        if action == "block":
            target=parts[2]
            u = load_users().get(target,{}); name = u.get("full_name") or u.get("first_name") or f"ID:{target}"
            if not is_main_admin(uid):
                if not get_admin_perm(uid, "can_block"):
                    await query.answer("Yetkin yok", show_alert=True); return ConversationHandler.END
                if not admin_can_manage_user(uid, target):
                    await query.answer("Bu kullaniciya yetkini yok", show_alert=True); return ConversationHandler.END
                # Sebep sor, sonra onaya gönder
                context.user_data["pending_action"] = "block"
                context.user_data["pending_target"] = target
                context.user_data["pending_tname"]  = name
                context.user_data["action"] = "pending_reason"
                await query.edit_message_text(
                    f"🚫 Engelleme sebebini yaz:\n{name}",
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("Iptal", callback_data="nav|root")
                    ]]))
                return WAIT_FOLDER
            blocked=load_blocked()
            if int(target) not in blocked: blocked.append(int(target)); save_blocked(blocked)
            await query.answer(TR["blocked_ok"].format(target), show_alert=True)
            query.data=f"user|info|{target}"; return await callback(update, context)

        if action == "unblock":
            target=parts[2]; blocked=load_blocked()
            if int(target) in blocked: blocked.remove(int(target)); save_blocked(blocked)
            await query.answer(TR["unblocked_ok"].format(target), show_alert=True)
            query.data=f"user|info|{target}"; return await callback(update, context)

    # ── İçerik yönetimi ───────────────────────────────
    if cb.startswith("cnt|") and adm:
        action = cb.split("|")[1]; folder = get_folder(content, path)

        if action == "remind":
            # Öğrencilere sınav hatırlatması gönder
            idx_r  = int(cb.split("|")[2])
            adm_cls = get_admin_cls(uid) or ""
            adm_sft = get_admin_shift(uid) if hasattr(__import__("builtins"), "get_admin_shift") else ""
            cds_r  = get_countdowns(adm_cls, uid=uid)
            if idx_r >= len(cds_r):
                await query.answer("Sınav bulunamadı", show_alert=True); return ConversationHandler.END
            cd_r = cds_r[idx_r]
            name_r = cd_r.get("name","")
            date_r = cd_r.get("date","")
            cls_r  = cd_r.get("cls","") or adm_cls
            grp_r  = cd_r.get("group","")
            msg_r  = f"⏰ تذكير بالامتحان\n\n📚 {name_r}\n📅 {date_r}"
            # Hedef kullanıcılar
            targets_r = get_target_users(f"cls_{cls_r}" if cls_r else "all")
            if grp_r:
                grps_d = load_groups()
                targets_r = [u for u in targets_r if grps_d.get(u,"").startswith(grp_r)]
            sent_r = 0
            for uid_r in targets_r:
                try: await context.bot.send_message(int(uid_r), msg_r); sent_r += 1
                except: pass
            await query.answer(f"✅ {sent_r} kişiye gönderildi", show_alert=True)
            return ConversationHandler.END

        if action == "add_countdown":
            if is_main_admin(uid) or get_admin_perm(uid, "can_countdown"):
                context.user_data["action"] = "countdown_name"
                await query.edit_message_text(
                    L(uid, "countdown_prompt"),
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton(L(uid, "cancel"), callback_data="nav|root")
                    ]]))
                return WAIT_FOLDER
            else:
                await query.answer("Yetkin yok", show_alert=True)
                return ConversationHandler.END

        if action == "add_folder":
            if not get_admin_perm(uid, "can_add_folder"):
                await query.answer("Yetkin yok / ليس لديك صلاحية", show_alert=True)
                return ConversationHandler.END
            context.user_data["action"] = "add_folder"
            hint = (
                "📁 Klasör adı yaz (tek tek)\n"
                "⚡ Hızlı: Birden fazla klasör için her satıra bir isim yaz\n\n"
                "Örnek:\nرياضيات\nفيزياء\nكيمياء"
                if is_main_admin(uid) else
                "📁 اكتب اسم المجلد\n⚡ لإضافة عدة مجلدات، اكتب اسماً في كل سطر"
            )
            await query.edit_message_text(
                hint,
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")]]))
            return WAIT_FOLDER

        if action == "add_file":
            if not get_admin_perm(uid, "can_add_file"):
                await query.answer("Yetkin yok / ليس لديك صلاحية", show_alert=True)
                return ConversationHandler.END
            context.user_data["action"] = "add_file"
            existing = folder_file_names(folder)
            if existing:
                file_list = "\n".join(f"  • {n}" for n in existing[:20])
                prompt = L(uid,"add_file_prompt").format(file_list)
            else:
                prompt = L(uid,"add_file_prompt_empty")
            await query.edit_message_text(prompt,
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")]]))
            return WAIT_FILE

        # Sil/Düzenle — SADECE süper admin
        if not is_main_admin(uid):
            await query.answer("⛔", show_alert=True); return ConversationHandler.END

        if action == "del_folder":
            folds = list(folder.get("folders",{}).keys())
            if not folds:
                await query.answer(L(uid,"no_folders"), show_alert=True); return ConversationHandler.END
            # Toplu seçim için seçili klasörler
            sel = context.user_data.get("del_folder_sel", set())
            kb  = []
            for f in folds:
                icon  = "☑️" if f in sel else "⬜️"
                kb.append([InlineKeyboardButton(f"{icon} 📁 {f}", callback_data=f"dsel|folder|{f}")])
            action_row = []
            if sel:
                action_row.append(InlineKeyboardButton(f"🗑 {len(sel)} Klasörü Sil", callback_data="do|bulk_del_folder"))
            action_row.append(InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root"))
            kb.append(action_row)
            await query.edit_message_text(
                f"🗑 {'Silmek istediğin klasörleri seç:' if is_main_admin(uid) else 'اختر المجلدات للحذف:'}\n",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "del_file":
            files = folder.get("files",[])
            if not files:
                await query.answer(L(uid,"no_files"), show_alert=True); return ConversationHandler.END
            sel = context.user_data.get("del_file_sel", set())
            kb  = []
            for i, f in enumerate(files):
                cap  = f.get("caption", f.get("name","?"))[:35]
                icon = "☑️" if i in sel else "⬜️"
                kb.append([InlineKeyboardButton(f"{icon} 📎 {cap}", callback_data=f"dsel|file|{i}")])
            action_row = []
            if sel:
                action_row.append(InlineKeyboardButton(f"🗑 {len(sel)} Dosyayı Sil", callback_data="do|bulk_del_file"))
            action_row.append(InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root"))
            kb.append(action_row)
            await query.edit_message_text(
                f"🗑 {'Silmek istediğin dosyaları seç:' if is_main_admin(uid) else 'اختر الملفات للحذف:'}\n",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "rename_folder":
            folds = list(folder.get("folders",{}).keys())
            if not folds:
                await query.answer(L(uid,"no_folders"), show_alert=True); return ConversationHandler.END
            kb = [[InlineKeyboardButton(f"✏️ {f}", callback_data=f"do|pick_folder|{f}")] for f in folds]
            kb.append([InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")])
            await query.edit_message_text(L(uid,"rename_folder_select"), reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "rename_file":
            files = folder.get("files",[])
            if not files:
                await query.answer(L(uid,"no_files"), show_alert=True); return ConversationHandler.END
            kb = [[InlineKeyboardButton(f"✏️ {f.get('caption','?')}", callback_data=f"do|pick_file|{i}")]
                  for i,f in enumerate(files)]
            kb.append([InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")])
            await query.edit_message_text(L(uid,"rename_file_select"), reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

    # ── Do (sadece süper admin) ───────────────────────
    # ── Checkbox toggle (seç/kaldır) ─────────────────
    if cb.startswith("dsel|") and is_main_admin(uid):
        parts    = cb.split("|")
        sel_type = parts[1]   # "folder" veya "file"
        val      = parts[2]   # klasör adı veya dosya index

        if sel_type == "folder":
            sel = context.user_data.setdefault("del_folder_sel", set())
            if val in sel: sel.discard(val)
            else:          sel.add(val)
            context.user_data["del_folder_sel"] = sel
            # Ekranı yenile
            folder2 = get_folder(content, path)
            folds   = list(folder2.get("folders",{}).keys())
            kb      = []
            for f in folds:
                icon = "☑️" if f in sel else "⬜️"
                kb.append([InlineKeyboardButton(f"{icon} 📁 {f}", callback_data=f"dsel|folder|{f}")])
            action_row = []
            if sel:
                action_row.append(InlineKeyboardButton(f"🗑 {len(sel)} Klasörü Sil", callback_data="do|bulk_del_folder"))
            action_row.append(InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root"))
            kb.append(action_row)
            try:
                await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup(kb))
            except: pass

        elif sel_type == "file":
            idx = int(val)
            sel = context.user_data.setdefault("del_file_sel", set())
            if idx in sel: sel.discard(idx)
            else:          sel.add(idx)
            context.user_data["del_file_sel"] = sel
            folder2 = get_folder(content, path)
            files   = folder2.get("files",[])
            kb      = []
            for i, f in enumerate(files):
                cap  = f.get("caption", f.get("name","?"))[:35]
                icon = "☑️" if i in sel else "⬜️"
                kb.append([InlineKeyboardButton(f"{icon} 📎 {cap}", callback_data=f"dsel|file|{i}")])
            action_row = []
            if sel:
                action_row.append(InlineKeyboardButton(f"🗑 {len(sel)} Dosyayı Sil", callback_data="do|bulk_del_file"))
            action_row.append(InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root"))
            kb.append(action_row)
            try:
                await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup(kb))
            except: pass
        return ConversationHandler.END

    if cb.startswith("do|") and is_main_admin(uid):
        parts  = cb.split("|",2); action = parts[1]; arg = parts[2] if len(parts)>2 else ""
        folder = get_folder(content, path)

        # ── Toplu Silme ──────────────────────────────
        if action == "bulk_del_folder":
            sel   = context.user_data.pop("del_folder_sel", set())
            count = 0
            for name in sel:
                if name in folder.get("folders",{}):
                    del folder["folders"][name]; count += 1
            if count: save_content(content)
            await show_folder(query, context, path,
                note=f"✅ {count} {'klasör silindi.' if is_main_admin(uid) else 'مجلد تم حذفه.'}")
            return ConversationHandler.END

        if action == "bulk_del_file":
            sel     = context.user_data.pop("del_file_sel", set())
            files   = folder.get("files",[])
            # Büyükten küçüğe sil ki indexler bozulmasın
            for idx in sorted(sel, reverse=True):
                if 0 <= idx < len(files): files.pop(idx)
            save_content(content)
            await show_folder(query, context, path,
                note=f"✅ {len(sel)} {'dosya silindi.' if is_main_admin(uid) else 'ملف تم حذفه.'}")
            return ConversationHandler.END

        # ── Silme Onayları (tekli — eski uyumluluk) ──
        if action == "confirm_del_folder":
            kb = [
                [InlineKeyboardButton(TR["confirm_yes"], callback_data=f"do|del_folder|{arg}"),
                 InlineKeyboardButton(TR["confirm_no"],  callback_data="nav|root")]
            ]
            await query.edit_message_text(
                L(uid,"del_folder_confirm").format(arg),
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "confirm_del_file":
            idx   = int(arg)
            files = folder.get("files",[])
            if idx >= len(files):
                await query.answer(L(uid,"file_notfound"), show_alert=True); return ConversationHandler.END
            fname = files[idx].get("caption","?")
            kb = [
                [InlineKeyboardButton(TR["confirm_yes"], callback_data=f"do|del_file|{idx}"),
                 InlineKeyboardButton(TR["confirm_no"],  callback_data="nav|root")]
            ]
            await query.edit_message_text(
                L(uid,"del_file_confirm").format(fname),
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        # ── Gerçek Silme ────────────────────────────
        if action == "del_folder":
            if arg in folder.get("folders",{}): del folder["folders"][arg]; save_content(content)
            await show_folder(query, context, path, note=L(uid,"folder_deleted").format(arg))
            return ConversationHandler.END

        if action == "pick_folder":
            context.user_data["action"] = "rename_folder"; context.user_data["target"] = arg
            await query.edit_message_text(L(uid,"new_folder_name").format(arg),
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")]]))
            return WAIT_RENAME_FOLDER

        if action == "del_file":
            idx = int(arg); files = folder.get("files",[])
            if 0 <= idx < len(files):
                removed = files.pop(idx); save_content(content)
                await show_folder(query, context, path, note=L(uid,"file_deleted").format(removed.get("caption","?")))
            return ConversationHandler.END

        if action == "pick_file":
            context.user_data["action"] = "rename_file"; context.user_data["target"] = int(arg)
            await query.edit_message_text(L(uid,"new_file_name"),
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(L(uid,"cancel"), callback_data="nav|root")]]))
            return WAIT_RENAME_FILE

    # ── Ayarlar (sadece süper admin) ──────────────────
    if cb.startswith("set|") and is_main_admin(uid):
        action = cb.split("|")[1]

        if action == "name":
            context.user_data["action"] = "set_name"; s = load_settings()
            await query.edit_message_text(TR["set_name_prompt"].format(s.get("bot_name","—")),
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_BOT_NAME

        if action == "welcome":
            context.user_data["action"] = "set_welcome"
            await query.edit_message_text(TR["set_welcome_prompt"],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_WELCOME

        if action == "photo":
            context.user_data["action"] = "set_photo"
            await query.edit_message_text(TR["set_photo_prompt"],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_PHOTO

        if action == "set_description":
            context.user_data["action"] = "set_description"
            s = load_settings()
            await query.edit_message_text(
                f"📄 Bot açıklamasını yaz (BotFather'daki 'about' kısmı):\n\nMevcut: {s.get('bot_description','—')[:100]}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_BOT_NAME  # aynı state kullan

        if action == "class_names":
            names = load_class_names()
            kb = []
            for cls_id, cls_def in CLASS_DEFS.items():
                cur_name = names.get(cls_id, cls_def["ar"])
                kb.append([InlineKeyboardButton(
                    f"✏️ {cls_def['ar']} → {cur_name}",
                    callback_data=f"set|edit_clsname|{cls_id}")])
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data="nav|root")])
            await query.edit_message_text(
                "🎓 Sinif Isimleri\n\nHer sinifa ozel isim ver:",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "edit_clsname":
            cls_id = cb.split("|")[2]
            context.user_data["action"] = "set_class_name"
            context.user_data["cls_name_id"] = cls_id
            cur = load_class_names().get(cls_id, CLASS_DEFS[cls_id]["ar"])
            await query.edit_message_text(
                f"Sinif {cls_id} icin yeni isim yaz:\nSu an: {cur}",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("◀️ İptal", callback_data="set|class_names")
                ]]))
            return WAIT_FOLDER

        if action == "group_mgmt":
            # Yeni sisteme yönlendir
            query.data = "set|class_groups"; return await callback(update, context)

        if action == "reset_groups":
            save_groups({}); save_shifts({})
            await query.answer("Tum gruplar sifirlandi", show_alert=True)
            await query.delete_message()
            return ConversationHandler.END

        if action == "remind_cfg":
            s = load_settings()
            days = s.get("exam_remind_days", 1)
            lab_days = s.get("lab_remind_days", 1)
            kb = [
                [InlineKeyboardButton(f"📚 Sinav: {days} gun once", callback_data="set|remind_exam_days")],
                [InlineKeyboardButton(f"🔬 Lab: {lab_days} gun once",  callback_data="set|remind_lab_days")],
                [InlineKeyboardButton("◀️ Geri", callback_data="nav|root")],
            ]
            await query.edit_message_text(
                f"Bildirim Ayarlari\nSinav: {days} gun once\nLab: {lab_days} gun once",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "remind_exam_days":
            kb = [[
                InlineKeyboardButton("1", callback_data="set|set_exam_days|1"),
                InlineKeyboardButton("2", callback_data="set|set_exam_days|2"),
                InlineKeyboardButton("3", callback_data="set|set_exam_days|3"),
                InlineKeyboardButton("5", callback_data="set|set_exam_days|5"),
                InlineKeyboardButton("7", callback_data="set|set_exam_days|7"),
            ]]
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data="set|remind_cfg")])
            await query.edit_message_text("Kac gun once bildirim?", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "set_exam_days":
            days = int(cb.split("|")[2])
            s = load_settings(); s["exam_remind_days"] = days; save_settings(s)
            await query.answer(f"✅ {days} gun", show_alert=True)
            query.data = "set|remind_cfg"; return await callback(update, context)

        if action == "remind_lab_days":
            kb = [[
                InlineKeyboardButton("1", callback_data="set|set_lab_days|1"),
                InlineKeyboardButton("2", callback_data="set|set_lab_days|2"),
                InlineKeyboardButton("3", callback_data="set|set_lab_days|3"),
            ]]
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data="set|remind_cfg")])
            await query.edit_message_text("🔬 Lab bildirimi kac gun once?", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "set_lab_days":
            days = int(cb.split("|")[2])
            s = load_settings(); s["lab_remind_days"] = days; save_settings(s)
            await query.answer(f"✅ {days} gun", show_alert=True)
            query.data = "set|remind_cfg"; return await callback(update, context)

        if action == "anon_group":
            s = load_settings()
            cur = s.get("anon_group_id","Ayarlanmadi")
            context.user_data["action"] = "set_anon_group"
            await query.edit_message_text(
                f"Anonim Mesaj Grubu\nMevcut: {cur}\n\nGrup ID yaz (ornek: -1001234567890)\nKapatmak icin 0:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Iptal","nav|root")]]))
            return WAIT_FOLDER

        if action == "class_groups":
            cls_grps = load_class_groups()
            kb = []
            for cls_id, cls_def in CLASS_DEFS.items():
                grps_this = list(cls_grps.get(cls_id, {}).keys())
                kb.append([InlineKeyboardButton(
                    f"🎓 {cls_def['tr']}: {', '.join(grps_this)}",
                    callback_data=f"set|cgrp_cls|{cls_id}")])
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data="nav|root")])
            await query.edit_message_text("Sinif Grup Yapılandırmasi:", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "cgrp_cls":
            parts = cb.split("|")
            cls_id = parts[2]
            cls_grps = load_class_groups()
            grps_this = cls_grps.get(cls_id, {})
            cls_name = CLASS_DEFS.get(cls_id,{}).get("tr","?")
            kb = []
            for g, subs in grps_this.items():
                subs_str = "/".join(subs)
                kb.append([
                    InlineKeyboardButton(f"👥 {g} ({subs_str})", callback_data=f"set|cgrp_edit|{cls_id}|{g}"),
                    InlineKeyboardButton("🗑", callback_data=f"set|cgrp_del|{cls_id}|{g}"),
                ])
            kb.append([InlineKeyboardButton("➕ Grup Ekle", callback_data=f"set|cgrp_add|{cls_id}")])
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data="set|class_groups")])
            await query.edit_message_text(f"👥 {cls_name} Gruplari:", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "cgrp_del":
            parts = cb.split("|")
            cls_id = parts[2]; grp = parts[3]
            cls_grps = load_class_groups()
            if cls_id in cls_grps and grp in cls_grps[cls_id]:
                del cls_grps[cls_id][grp]
                save_class_groups(cls_grps)
                await query.answer(f"✅ {grp} silindi", show_alert=True)
            query.data = f"set|cgrp_cls|{cls_id}"; return await callback(update, context)

        if action == "cgrp_add":
            parts = cb.split("|")
            cls_id = parts[2]
            context.user_data["action"] = "cgrp_add_name"
            context.user_data["cgrp_add_cls"] = cls_id
            await query.edit_message_text(
                "Yeni grup adi (ornek: D):",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️","close")]]))
            return WAIT_FOLDER

        if action == "cgrp_edit":
            parts = cb.split("|")
            cls_id = parts[2]; grp = parts[3]
            cls_grps = load_class_groups()
            subs = cls_grps.get(cls_id,{}).get(grp, [])
            kb = []
            for sub in subs:
                kb.append([
                    InlineKeyboardButton(sub, callback_data=f"set|cgrp_sub_view|{cls_id}|{grp}|{sub}"),
                    InlineKeyboardButton("🗑", callback_data=f"set|cgrp_sub_del|{cls_id}|{grp}|{sub}"),
                ])
            kb.append([InlineKeyboardButton("➕ Alt Bolum Ekle", callback_data=f"set|cgrp_sub_add|{cls_id}|{grp}")])
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data=f"set|cgrp_cls|{cls_id}")])
            await query.edit_message_text(
                f"Grup {grp}: {', '.join(subs)}",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "cgrp_sub_del":
            parts = cb.split("|")
            cls_id = parts[2]; grp = parts[3]; sub = parts[4]
            cls_grps = load_class_groups()
            if cls_id in cls_grps and grp in cls_grps[cls_id] and sub in cls_grps[cls_id][grp]:
                cls_grps[cls_id][grp].remove(sub)
                save_class_groups(cls_grps)
                await query.answer(f"✅ {sub} silindi", show_alert=True)
            query.data = f"set|cgrp_edit|{cls_id}|{grp}"; return await callback(update, context)

        if action == "cgrp_sub_add":
            parts = cb.split("|")
            cls_id = parts[2]; grp = parts[3]
            context.user_data["action"] = "cgrp_sub_add_name"
            context.user_data["cgrp_cls"] = cls_id
            context.user_data["cgrp_grp"] = grp
            await query.edit_message_text(
                f"Alt bolum adi (ornek: {grp}4):",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️","close")]]))
            return WAIT_FOLDER

        if action == "admin_perms":
            # Admin listesini göster
            admins = load_admins()
            if not admins:
                await query.answer("Henüz alt admin yok.", show_alert=True)
                return ConversationHandler.END
            users_d = load_users()
            kb = []
            for adm_id in admins:
                u = users_d.get(str(adm_id), {})
                name = u.get("full_name") or u.get("first_name") or f"ID:{adm_id}"
                kb.append([InlineKeyboardButton(f"👮 {name[:25]}", callback_data=f"adminperm|show|{adm_id}")])
            kb.append([InlineKeyboardButton("◀️ Geri", callback_data="nav|root")])
            await query.edit_message_text("👮 Admin Yetki Yönetimi", reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "maint_text":
            context.user_data["action"] = "set_maint_text"
            await query.edit_message_text(
                "🔧 Bakım modu mesajını yaz (kullanıcılara gösterilecek):",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_WELCOME

        if action == "blocked":
            blocked = load_blocked()
            txt  = TR["blocked_list"].format(len(blocked)) + "\n\n"
            txt += "\n".join(f"🆔 {b}" for b in blocked) if blocked else TR["no_blocked"]
            await query.edit_message_text(txt, reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")]]))
            return ConversationHandler.END

        if action == "btn_mgmt":
            s     = load_settings()
            ub    = s.get("user_buttons", {})
            track = s.get("track_activity", True)
            btn_defs = [
                ("btn_search", "🔍 Arama Butonu"),
                ("btn_help",   "💬 Mesaj Gönder Butonu"),
                
            ]
            kb = []
            for key, label in btn_defs:
                icon = "✅" if ub.get(key, True) else "❌"
                kb.append([InlineKeyboardButton(f"{icon} {label}", callback_data=f"set|toggle_btn|{key}")])
            track_icon = "✅" if track else "❌"
            kb.append([InlineKeyboardButton(f"{track_icon} 👁 Aktivite Takibi", callback_data="set|toggle_track")])
            kb.append([InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")])
            await query.edit_message_text(
                TR["btn_mgmt_title"] + "\n\n✅ = Açık  |  ❌ = Kapalı\n"
                "Kullanıcıların göreceği butonları buradan yönet.",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "toggle_track":
            s = load_settings()
            s["track_activity"] = not s.get("track_activity", True)
            save_settings(s)
            ub    = s.get("user_buttons", {})
            track = s["track_activity"]
            btn_defs = [
                ("btn_search", "🔍 Arama Butonu"),
                ("btn_help",   "💬 Mesaj Gönder Butonu"),
                
            ]
            kb = []
            for key, label in btn_defs:
                icon = "✅" if ub.get(key, True) else "❌"
                kb.append([InlineKeyboardButton(f"{icon} {label}", callback_data=f"set|toggle_btn|{key}")])
            track_icon = "✅" if track else "❌"
            kb.append([InlineKeyboardButton(f"{track_icon} 👁 Aktivite Takibi", callback_data="set|toggle_track")])
            kb.append([InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")])
            status = "✅ AÇIK" if track else "❌ KAPALI"
            await query.edit_message_text(
                TR["btn_mgmt_title"] + f"\n\n👁 Aktivite Takibi → {status}\n\n✅ = Açık  |  ❌ = Kapalı",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action == "toggle_btn":
            key = cb.split("|")[2]
            s   = load_settings()
            ub  = s.setdefault("user_buttons", {"btn_search": True, "btn_help": True})
            ub[key] = not ub.get(key, True)
            save_settings(s)
            track    = s.get("track_activity", True)
            btn_defs = [
                ("btn_search", "🔍 Arama Butonu"),
                ("btn_help",   "💬 Mesaj Gönder Butonu"),
                
            ]
            kb = []
            for k, label in btn_defs:
                icon = "✅" if ub.get(k, True) else "❌"
                kb.append([InlineKeyboardButton(f"{icon} {label}", callback_data=f"set|toggle_btn|{k}")])
            track_icon = "✅" if track else "❌"
            kb.append([InlineKeyboardButton(f"{track_icon} 👁 Aktivite Takibi", callback_data="set|toggle_track")])
            kb.append([InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")])
            await query.edit_message_text(
                TR["btn_mgmt_title"] + "\n\n✅ = Açık  |  ❌ = Kapalı",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

    # ── AI Cevap Puanlama ────────────────────────────
    if cb.startswith("rate|") and not is_admin(uid):
        parts_r = cb.split("|")
        file_key = parts_r[1]
        stars    = int(parts_r[2])
        add_feedback(file_key, uid, stars)
        star_str = "⭐" * stars + "☆" * (5 - stars)
        await query.answer(f"{star_str} — {L(uid,'feedback_saved')}", show_alert=False)
        # Liderboard puan
        update_leaderboard(uid, stars)
        return ConversationHandler.END

    # ── Liderboard ────────────────────────────────────
    if cb == "misc|leaderboard":
        lb    = get_leaderboard(10)
        lines = [L(uid,"leaderboard_title"), ""]
        medals = ["🥇","🥈","🥉","4️⃣","5️⃣","6️⃣","7️⃣","8️⃣","9️⃣","🔟"]
        for i,(name,pts) in enumerate(lb):
            medal = medals[i] if i < len(medals) else f"{i+1}."
            lines.append(f"{medal} {name[:25]}  —  {pts} pts")
        if not lb:
            lines.append(L(uid,"leaderboard_empty"))
        await query.edit_message_text(
            "\n".join(lines),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(L(uid,"back"),  callback_data="nav|root")]]))
        return ConversationHandler.END

    # ── Sabit Mesaj Ayarla (admin) ────────────────────
    if cb == "set|pin_msg" and is_main_admin(uid):
        context.user_data["action"] = "set_pin_msg"
        pinned = get_pinned_msg()
        prompt = TR["pin_msg_prompt"]
        if pinned:
            prompt += f"\n\nMevcut: {pinned[:100]}"
        await query.edit_message_text(prompt,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
        return WAIT_WELCOME   # aynı state kullan

    # ── Duyuru Geçmişi (admin) ───────────────────────
    if cb == "bcast|history" and is_main_admin(uid):
        log_ = load_bcast_log()
        if not log_:
            await query.answer(TR["bcast_history_empty"], show_alert=True); return ConversationHandler.END
        lines = [TR["bcast_history_title"]]
        for entry in reversed(log_[-20:]):
            lines.append(
                f"\n📢 {entry.get('time','')}\n"
                f"   👥 {entry.get('count',0)} kişi\n"
                f"   📝 {entry.get('text','')[:60]}"
            )
        await query.edit_message_text(
            "\n".join(lines)[:4000],
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")]]))
        return ConversationHandler.END

    if cb == "close":
        try: await query.delete_message()
        except: pass
        return ConversationHandler.END

    # ── Notlar ───────────────────────────────────────
    if cb.startswith("notes|") and not is_admin(uid):
        parts_n = cb.split("|")
        action_n = parts_n[1]

        if action_n == "new":
            # Not alma modu — ders adı sor
            context.user_data["action"] = "note_subject_input"
            context.user_data.pop("note_subject", None)
            context.user_data.pop("note_mode", None)
            await query.edit_message_text(
                "📝 اكتب اسم الدرس (أو . للتخطي):",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("◀️ إلغاء", callback_data="nav|root")
                ]]))
            return WAIT_FOLDER

        if action_n == "pick_subject":
            # eski uyumluluk — yeni akışa yönlendir
            query.data = "notes|new"
            return await callback(update, context)

        if action_n == "sel_subj":
            subject = "|".join(parts_n[2:]) if len(parts_n) > 2 else ""
            context.user_data["note_subject"] = subject
            context.user_data["action"] = "note_taking"
            kb = [[InlineKeyboardButton("🔒 حفظ وإنهاء", callback_data="notes|done")]]
            sent = await query.edit_message_text(
                f"📖 {subject or 'ملاحظة'}\n\nاكتب ملاحظاتك:",
                reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action_n == "subject":
            subject = "|".join(parts_n[2:]) if len(parts_n) > 2 else "أخرى"
            notes = [n for n in get_user_notes(uid) if n.get("subject","أخرى") == subject]
            kb = []
            for n in notes[-10:]:
                icon = {"text":"✍️","photo":"🖼","voice":"🎙","document":"📄","video":"🎥"}.get(n["type"],"📌")
                cap  = n.get("content","")[:22] or n["type"]
                kb.append([
                    InlineKeyboardButton(f"{icon} {cap}", callback_data=f"notes|view|{n['id']}"),
                    InlineKeyboardButton("🗑", callback_data=f"notes|del|{n['id']}"),
                ])
            kb.append([InlineKeyboardButton("➕", callback_data="notes|pick_subject"),
                       InlineKeyboardButton("◀️", callback_data="nav|root")])
            await query.edit_message_text(
                f"📖 {subject} ({len(notes)})",
                reply_markup=InlineKeyboardMarkup(kb))
            return ConversationHandler.END

        if action_n == "add":
            # Sadece metin notu
            context.user_data["action"] = "note_taking"
            context.user_data.setdefault("note_subject","")
            kb = [[InlineKeyboardButton("🔒 حفظ وإنهاء", callback_data="notes|done")]]
            await query.edit_message_text(
                "✍️ اكتب ملاحظاتك:",
                reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action_n == "view":
            note_id = parts_n[2] if len(parts_n) > 2 else ""
            notes = get_user_notes(uid)
            note = next((n for n in notes if n.get("id") == note_id), None)
            if not note:
                await query.answer("❌", show_alert=True); return ConversationHandler.END
            fid = note.get("file_id")
            cap = note.get("content","") + f"\n🕐 {note.get('time','')}"
            kb = [[InlineKeyboardButton("🗑 حذف", callback_data=f"notes|del|{note_id}"),
                   InlineKeyboardButton("◀️", callback_data="nav|root")]]
            try:
                ntype = note.get("type","text")
                if ntype == "photo" and fid:
                    await query.message.reply_photo(fid, caption=cap, reply_markup=InlineKeyboardMarkup(kb))
                elif ntype == "video" and fid:
                    await query.message.reply_video(fid, caption=cap, reply_markup=InlineKeyboardMarkup(kb))
                elif ntype in ("voice","audio") and fid:
                    await query.message.reply_voice(fid, caption=cap, reply_markup=InlineKeyboardMarkup(kb))
                elif ntype == "document" and fid:
                    await query.message.reply_document(fid, caption=cap, reply_markup=InlineKeyboardMarkup(kb))
                elif fid:
                    # Bilinmeyen tip — doküman olarak gönder
                    try:
                        await query.message.reply_document(fid, caption=cap, reply_markup=InlineKeyboardMarkup(kb))
                    except:
                        await query.edit_message_text(cap, reply_markup=InlineKeyboardMarkup(kb))
                else:
                    await query.edit_message_text(cap, reply_markup=InlineKeyboardMarkup(kb))
            except: pass
            return ConversationHandler.END

        if action_n == "done":
            # Not alma modunu kapat
            subject = context.user_data.pop("note_subject","")
            context.user_data.pop("action",None)
            notes = get_user_notes(uid)
            cnt   = len(notes)
            await query.edit_message_text(f"✅ تم حفظ الملاحظات ({cnt} ملاحظة)")
            return ConversationHandler.END

        if action_n == "del":
            note_id = parts_n[2] if len(parts_n) > 2 else ""
            delete_user_note(uid, note_id)
            await query.answer("✅ حُذف", show_alert=False)
            await query.delete_message()
            return ConversationHandler.END

        return ConversationHandler.END

    if cb.startswith("cd_shift|") and is_admin(uid) and (is_main_admin(uid) or get_admin_perm(uid, "can_countdown")):
        shift_val = cb.split("|")[1]
        context.user_data["countdown_shift"] = "" if shift_val == "all" else shift_val
        admin_cls = get_admin_cls(uid)
        if admin_cls:
            context.user_data["countdown_cls"] = admin_cls
        cls_for_grp = admin_cls or context.user_data.get("countdown_cls","")
        if cls_for_grp:
            cls_grps = get_class_groups(cls_for_grp)
            grp_keys = list(cls_grps.keys())
        else:
            grp_keys = ["A","B","C"]
        kb = [[InlineKeyboardButton("الكل", callback_data="cd_grp|all")]]
        row = [InlineKeyboardButton(g, callback_data=f"cd_grp|{g}") for g in grp_keys]
        if row: kb.append(row)
        await query.edit_message_text("📅 المجموعة:", reply_markup=InlineKeyboardMarkup(kb))
        return WAIT_FOLDER

    if cb.startswith("cd_cls|") and is_admin(uid) and (is_main_admin(uid) or get_admin_perm(uid, "can_countdown")):
        cls_val = cb.split("|")[1]
        context.user_data["countdown_cls"] = "" if cls_val == "all" else cls_val
        kb = [
            [InlineKeyboardButton("الكل",   callback_data="cd_shift|all"),
             InlineKeyboardButton("صباحي", callback_data="cd_shift|sabahi"),
             InlineKeyboardButton("مسائي", callback_data="cd_shift|masaiy")],
        ]
        await query.edit_message_text("📅 الفترة الدراسية:", reply_markup=InlineKeyboardMarkup(kb))
        return WAIT_FOLDER

    if cb.startswith("cd_grp|") and is_admin(uid) and (is_main_admin(uid) or get_admin_perm(uid, "can_countdown")):
        grp_val = cb.split("|")[1]
        context.user_data["countdown_group"] = "" if grp_val == "all" else grp_val
        context.user_data["action"] = "countdown_date"
        await query.edit_message_text(
            L(uid,"countdown_date"),
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(L(uid,"back"), callback_data="nav|root")
            ]]))
        return WAIT_FOLDER

    if cb.startswith("reminder|") and not is_admin(uid):
        parts  = cb.split("|")
        action = parts[1]
        if action == "add":
            context.user_data["action"] = "reminder_text"
            await query.edit_message_text(
                L(uid,"reminder_add_prompt"),
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton(L(uid,"back"), callback_data="nav|root")
                ]]))
            return WAIT_FOLDER

        if action == "quick":
            minutes = int(parts[2])
            rem_text = context.user_data.pop("reminder_text_pending", "")
            if not rem_text:
                await query.answer("❌", show_alert=True); return ConversationHandler.END
            import time as _t
            fire_ts = _t.time() + minutes * 60
            add_reminder(uid, rem_text, fire_ts)
            if minutes < 60: label = f"{minutes} دقيقة"
            elif minutes < 1440: label = f"{minutes//60} ساعة"
            else: label = f"{minutes//1440} يوم"
            await query.edit_message_text(L(uid,"reminder_saved").format(label))
            return ConversationHandler.END

        if action == "manual_date":
            context.user_data["action"] = "reminder_day"
            await query.edit_message_text(
                "📅 كم يوم تريد؟ (0 = اليوم):",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("◀️ إلغاء", callback_data="nav|root")
                ]]))
            return WAIT_FOLDER
        if action == "pick_time":
            # Süre seçildi — callback: reminder|pick_time|DAKIKA
            minutes = int(parts[2])
            rem_text = context.user_data.pop("reminder_text_pending", context.user_data.pop("reminder_text","?"))
            import time
            fire_ts = time.time() + (minutes * 60)
            add_reminder(uid, rem_text, fire_ts)
            if minutes < 60:
                label = f"{minutes} دقيقة" if not is_main_admin(uid) else f"{minutes} dakika"
            elif minutes < 1440:
                label = f"{minutes//60} ساعة" if not is_main_admin(uid) else f"{minutes//60} saat"
            else:
                label = f"{minutes//1440} يوم" if not is_main_admin(uid) else f"{minutes//1440} gün"
            await query.edit_message_text(L(uid,"reminder_saved").format(label))
            context.user_data.pop("action", None)
            return ConversationHandler.END

        if action == "manual_date":
            context.user_data["action"] = "reminder_day"
            await query.edit_message_text(
                "📅 اليوم (1-31):" if not is_main_admin(uid) else "📅 Gün (1-31):",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️", callback_data="nav|root")]]))
            return WAIT_FOLDER
        if action == "del" and len(parts) > 2:
            idx = int(parts[2])
            delete_reminder(uid, idx)
            await query.answer(L(uid,"reminder_del"), show_alert=True)
            await query.delete_message()
        return ConversationHandler.END

    # ── Admin: Sınav/Etkinlik Ekle ───────────────────
    if cb.startswith("countdown|") and is_admin(uid):
        action = cb.split("|")[1]
        if action == "add":
            context.user_data["action"] = "countdown_name"
            await query.edit_message_text(
                L(uid,"countdown_prompt"),
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton(L(uid,"back"), callback_data="nav|root")
                ]]))
            return WAIT_FOLDER
        if action == "del" and len(cb.split("|")) > 2:
            idx = int(cb.split("|")[2])
            cd = load_countdowns()
            if 0 <= idx < len(cd): cd.pop(idx)
            save_countdowns(cd)
            await query.answer("✅ Silindi", show_alert=True)
            await query.delete_message()
        return ConversationHandler.END

    # ── Quiz cevabı ──────────────────────────────────
    if cb.startswith("quiz|ans|") and not is_admin(uid):
        parts   = cb.split("|")
        qid     = parts[2]
        ans_idx = int(parts[3])
        quizzes = load_quizzes()
        q = next((x for x in quizzes if x.get("id") == qid), None)
        if not q:
            await query.answer("❌ Sona erdi", show_alert=True)
            return ConversationHandler.END
        answered = q.setdefault("answered", {})
        if uid in answered:
            await query.answer("✅ Zaten cevapladın", show_alert=True)
            return ConversationHandler.END
        answered[uid] = ans_idx
        correct = q.get("correct", -1)
        is_correct = (ans_idx == correct)
        save_quizzes(quizzes)
        if is_correct:
            update_leaderboard(uid, 5)
            await query.answer("✅ صحيح! +5 pts 🎉", show_alert=True)
        else:
            correct_text = q["options"][correct] if 0 <= correct < len(q["options"]) else "?"
            await query.answer(f"❌ خطأ. الجواب: {correct_text}", show_alert=True)
        return ConversationHandler.END

    # ── Anket işlemleri ───────────────────────────────
    if cb.startswith("poll|") and is_admin(uid) and (is_main_admin(uid) or get_admin_perm(uid, "can_poll")):
        action = cb.split("|")[1]

        if action == "create":
            adm_cls = get_admin_cls(uid) if not is_main_admin(uid) else None
            if adm_cls:
                # Sınıf sabit — vardiya seç
                context.user_data["poll_tgt_cls"] = adm_cls
                kb = [
                    [InlineKeyboardButton("☀️ صباحي", callback_data="poll|tgt_sft|sabahi"),
                     InlineKeyboardButton("🌙 مسائي",  callback_data="poll|tgt_sft|masaiy"),
                     InlineKeyboardButton("📋 الكل",   callback_data="poll|tgt_sft|all")],
                    [InlineKeyboardButton("◀️ إلغاء",  callback_data="close")],
                ]
                await query.edit_message_text("📊 الفترة الدراسية:", reply_markup=InlineKeyboardMarkup(kb))
            else:
                # Sınıf seç
                kb = []
                for cls_id, cls_def in CLASS_DEFS.items():
                    kb.append([InlineKeyboardButton(cls_def["ar"], callback_data=f"poll|tgt_cls|{cls_id}")])
                kb.append([InlineKeyboardButton("📋 الكل", callback_data="poll|tgt_cls|all")])
                kb.append([InlineKeyboardButton("◀️ إلغاء", callback_data="close")])
                await query.edit_message_text("📊 السنة الدراسية:", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "tgt_cls":
            # Sınıf seçildi → vardiya seç
            cls_val = cb.split("|")[2] if len(cb.split("|")) > 2 else "all"
            context.user_data["poll_tgt_cls"] = "" if cls_val == "all" else cls_val
            kb = [
                [InlineKeyboardButton("☀️ صباحي", callback_data="poll|tgt_sft|sabahi"),
                 InlineKeyboardButton("🌙 مسائي",  callback_data="poll|tgt_sft|masaiy"),
                 InlineKeyboardButton("📋 الكل",   callback_data="poll|tgt_sft|all")],
                [InlineKeyboardButton("◀️ رجوع",   callback_data="poll|create")],
            ]
            await query.edit_message_text("📊 الفترة الدراسية:", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "tgt_sft":
            # Vardiya seçildi → grup seç
            sft_val = cb.split("|")[2] if len(cb.split("|")) > 2 else "all"
            context.user_data["poll_tgt_sft"] = "" if sft_val == "all" else sft_val
            tgt_cls = context.user_data.get("poll_tgt_cls","")
            adm_grp = get_admin_grp(uid) if not is_main_admin(uid) else None
            if adm_grp:
                context.user_data["poll_tgt_grp"] = adm_grp
                # Hedefi oluştur ve tip seç
                _build_poll_target(context, uid)
                kb = [
                    [InlineKeyboardButton(TR["poll_type_choice"], callback_data="poll|type|choice")],
                    [InlineKeyboardButton(TR["poll_type_open"],   callback_data="poll|type|open")],
                    [InlineKeyboardButton("◀️ إلغاء",            callback_data="close")],
                ]
                lbl = target_label(context.user_data.get("poll_target",""))
                await query.edit_message_text(f"📊 {lbl}\n{TR['poll_type_select']}", reply_markup=InlineKeyboardMarkup(kb))
            else:
                grp_keys = list(get_class_groups(tgt_cls).keys()) if tgt_cls else ["A","B","C"]
                kb = []
                row = [InlineKeyboardButton("📋 الكل", callback_data="poll|tgt_grp|all")]
                for g in grp_keys:
                    row.append(InlineKeyboardButton(g, callback_data=f"poll|tgt_grp|{g}"))
                kb.append(row)
                kb.append([InlineKeyboardButton("◀️ رجوع", callback_data="poll|create")])
                await query.edit_message_text("📊 المجموعة:", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "tgt_grp":
            grp_val = cb.split("|")[2] if len(cb.split("|")) > 2 else "all"
            context.user_data["poll_tgt_grp"] = "" if grp_val == "all" else grp_val
            _build_poll_target(context, uid)
            kb = [
                [InlineKeyboardButton(TR["poll_type_choice"], callback_data="poll|type|choice")],
                [InlineKeyboardButton(TR["poll_type_open"],   callback_data="poll|type|open")],
                [InlineKeyboardButton("◀️ إلغاء",            callback_data="close")],
            ]
            lbl = target_label(context.user_data.get("poll_target",""))
            await query.edit_message_text(f"📊 {lbl}\n{TR['poll_type_select']}", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "tgt":
            parts_tgt = cb.split("|")
            step = parts_tgt[-1]
            val  = "|".join(parts_tgt[2:-1]) if len(parts_tgt) > 3 else (parts_tgt[2] if len(parts_tgt) > 2 else "")
            if step == "done":
                context.user_data["poll_target"] = val
                lbl = target_label(val)
                kb = [
                    [InlineKeyboardButton(TR["poll_type_choice"], callback_data="poll|type|choice")],
                    [InlineKeyboardButton(TR["poll_type_open"],   callback_data="poll|type|open")],
                    [InlineKeyboardButton(TR["cancel"],           callback_data="close")],
                ]
                await query.edit_message_text(
                    f"🎯 {lbl}\n\n{TR['poll_type_select']}",
                    reply_markup=InlineKeyboardMarkup(kb))
            else:
                kb = build_target_keyboard("poll|tgt", step, val)
                kb.append([InlineKeyboardButton("◀️", callback_data="close")])
                labels = {"cls":"السنة","shift":"الفترة","grp":"المجموعة","subgrp":"الفرعية"}
                await query.edit_message_text(f"📊 {labels.get(step,step)}:", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "target":
            poll_target = cb.split("|")[2]
            context.user_data["poll_target"] = poll_target
            kb = [
                [InlineKeyboardButton(TR["poll_type_choice"], callback_data="poll|type|choice")],
                [InlineKeyboardButton(TR["poll_type_open"],   callback_data="poll|type|open")],
                [InlineKeyboardButton(TR["cancel"],           callback_data="close")],
            ]
            await query.edit_message_text(TR["poll_type_select"], reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER


        if action == "type":
            poll_type = cb.split("|")[2]  # "choice" veya "open"
            context.user_data["poll_type"]   = poll_type
            context.user_data["action"]      = "poll_question"
            await query.edit_message_text(
                TR["poll_enter_q"],
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(TR["cancel"], callback_data="close")]]))
            return WAIT_POLL_QUESTION

        if action == "list_results":
            polls = load_polls()
            if not polls:
                await query.answer(TR["poll_no_polls"], show_alert=True); return ConversationHandler.END
            kb = [[InlineKeyboardButton(f"📊 {p['question'][:40]}", callback_data=f"poll|show_result|{pid}")]
                  for pid, p in polls.items()]
            kb.append([InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")])
            await query.edit_message_text(TR["poll_select"], reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "show_result":
            poll_id = cb.split("|")[2]
            polls   = load_polls()
            poll    = polls.get(poll_id)
            if not poll:
                await query.answer("Anket bulunamadı.", show_alert=True); return ConversationHandler.END

            poll_type = poll.get("type", "choice")

            if poll_type == "open":
                # Açık uçlu — cevapları göster
                answers = poll.get("answers", {})
                lines   = [
                    f"✍️ Açık Uçlu Anket Sonuçları",
                    f"❓ {poll['question']}",
                    f"👥 Toplam Cevap: {len(answers)}",
                    f"",
                ]
                if answers:
                    for a in answers.values():
                        lines.append(
                            f"\n👤 {a['name']} {a.get('username','')}\n"
                            f"💬 {a['answer']}\n"
                            f"🕐 {a['time'][-5:]}"
                        )
                else:
                    lines.append("Henüz cevap yok.")
                kb = [[InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")]]
                await query.edit_message_text("\n".join(lines)[:4000], reply_markup=InlineKeyboardMarkup(kb))
            else:
                txt = build_poll_results_text(poll, uid, show_voters=False)
                kb  = [
                    [InlineKeyboardButton("👥 Kim Ne Seçti?", callback_data=f"poll|voter_detail|{poll_id}")],
                    [InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")],
                ]
                await query.edit_message_text(txt, reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "voter_detail":
            poll_id = cb.split("|")[2]
            polls   = load_polls()
            poll    = polls.get(poll_id)
            if not poll:
                await query.answer("Anket bulunamadı.", show_alert=True); return ConversationHandler.END
            txt = build_poll_results_text(poll, uid, show_voters=True)
            kb  = [
                [InlineKeyboardButton("💬 Yorumlar", callback_data=f"poll|comments|{poll_id}")],
                [InlineKeyboardButton("📊 Genel Sonuç", callback_data=f"poll|show_result|{poll_id}")],
                [InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")],
            ]
            await query.edit_message_text(txt[:4000], reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "comments":
            poll_id  = cb.split("|")[2]
            polls    = load_polls()
            poll     = polls.get(poll_id)
            if not poll:
                await query.answer("Anket bulunamadı.", show_alert=True); return ConversationHandler.END
            comments = poll.get("comments", {})
            if not comments:
                await query.answer(TR["poll_no_comments"], show_alert=True)
                return ConversationHandler.END
            lines = [f"💬 Yorumlar — {poll['question'][:40]}\n"]
            for c in comments.values():
                name = c.get("name","?")
                un   = f" {c['username']}" if c.get("username") else ""
                vote = c.get("vote","—")
                text_c = c.get("comment","")
                time_c = c.get("time","")[-5:]
                lines.append(
                    f"\n👤 {name}{un}\n"
                    f"🗳 {vote}\n"
                    f"💬 {text_c}\n"
                    f"🕐 {time_c}"
                )
            kb = [
                [InlineKeyboardButton("◀️ Geri", callback_data=f"poll|voter_detail|{poll_id}")],
                [InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")],
            ]
            await query.edit_message_text("\n".join(lines)[:4000], reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "list_delete":
            polls = load_polls()
            if not polls:
                await query.answer(TR["poll_no_polls"], show_alert=True); return ConversationHandler.END
            kb = [[InlineKeyboardButton(f"🗑 {p['question'][:40]}", callback_data=f"poll|confirm_delete|{pid}")]
                  for pid, p in polls.items()]
            kb.append([InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")])
            await query.edit_message_text(TR["poll_select"], reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "confirm_delete":
            poll_id = cb.split("|")[2]
            polls   = load_polls()
            if poll_id in polls:
                del polls[poll_id]
                save_polls(polls)
            await query.answer(TR["poll_deleted"], show_alert=True)
            # Panele dön
            kb = [[InlineKeyboardButton(TR["poll_create"], callback_data="poll|create")]]
            if polls:
                kb.append([InlineKeyboardButton(TR["poll_results"], callback_data="poll|list_results"),
                           InlineKeyboardButton(TR["poll_delete"],  callback_data="poll|list_delete")])
            kb.append([InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")])
            active = sum(1 for p in polls.values() if p.get("active"))
            txt = f"{TR['poll_panel']}\n\n📊 Toplam anket: {len(polls)}\n✅ Aktif: {active}"
            await query.edit_message_text(txt, reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER

        if action == "send":
            poll_id   = cb.split("|")[2]
            polls     = load_polls()
            poll      = polls.get(poll_id)
            if not poll:
                await query.answer("Anket bulunamadı.", show_alert=True); return ConversationHandler.END
            users     = load_users()
            success = fail = 0
            q         = poll["question"]
            poll_type = poll.get("type", "choice")

            # Hedef kullanıcıları belirle
            poll_target = poll.get("target", "all")
            if not poll_target or poll_target == "all":
                target_uids = [u for u in users if int(u) != ADMIN_ID and not is_blocked(u)]
            else:
                target_uids = get_target_users(poll_target)

            for uid_ in target_uids:
                if int(uid_) == ADMIN_ID: continue
                try:
                    if poll_type == "open":
                        # Açık uçlu — "Cevapla" butonu
                        kb_u = [[InlineKeyboardButton(
                            "✍️ أكتب إجابتك" if not is_main_admin(uid_) else "✍️ Cevapla",
                            callback_data=f"open_ans|{poll_id}")]]
                        await context.bot.send_message(
                            int(uid_),
                            f"✍️ سؤال مفتوح!\n\n❓ {q}",
                            reply_markup=InlineKeyboardMarkup(kb_u))
                    else:
                        opts   = poll["options"]
                        poll_kb = [[InlineKeyboardButton(f"  {o}  ", callback_data=f"vote|{poll_id}|{i}")]
                                   for i, o in enumerate(opts)]
                        await context.bot.send_message(
                            int(uid_),
                            f"📊 استطلاع جديد!\n\n❓ {q}",
                            reply_markup=InlineKeyboardMarkup(poll_kb))
                    success += 1
                except: fail += 1

            poll["active"] = True
            save_polls(polls)
            await query.edit_message_text(
                TR["poll_sent"].format(s=success, f=fail),
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Geri",   callback_data="nav|root")]]))
            return ConversationHandler.END

    # ── Açık uçlu anket cevap butonu ─────────────────
    if cb.startswith("open_ans|"):
        poll_id = cb.split("|")[1]
        polls   = load_polls()
        poll    = polls.get(poll_id)
        if not poll or not poll.get("active"):
            await query.answer(AR["poll_closed"], show_alert=True); return ConversationHandler.END
        answers = poll.get("answers", {})
        if uid in answers:
            await query.answer(AR["poll_already_voted"], show_alert=True); return ConversationHandler.END
        context.user_data["open_poll_id"] = poll_id
        context.user_data["action"]       = "open_poll_answer"
        await query.message.reply_text(
            f"✍️ {poll['question']}\n\n{'اكتب إجابتك:' if not is_main_admin(uid) else 'Cevabını yaz:'}")
        return WAIT_POLL_COMMENT

    # ── Kullanıcı oy kullanma ─────────────────────────
    if cb.startswith("vote|"):
        parts   = cb.split("|")
        poll_id = parts[1]
        opt_idx = int(parts[2])
        polls   = load_polls()
        poll    = polls.get(poll_id)
        if not poll:
            await query.answer(TR["poll_closed"] if is_main_admin(uid) else AR["poll_closed"],
                               show_alert=True); return ConversationHandler.END
        if not poll.get("active"):
            await query.answer(AR["poll_closed"], show_alert=True); return ConversationHandler.END
        votes = poll.setdefault("votes", {})
        if uid in votes:
            await query.answer(AR["poll_already_voted"], show_alert=True); return ConversationHandler.END

        # Oyu kaydet
        votes[uid] = poll["options"][opt_idx]
        save_polls(polls)
        await query.answer(AR["poll_voted"], show_alert=True)

        # Anket butonlarını güncel oy sayılarıyla güncelle
        opts   = poll["options"]
        total  = len(votes)
        counts = {o: 0 for o in opts}
        for v in votes.values():
            if v in counts: counts[v] += 1

        poll_kb = []
        for i, o in enumerate(opts):
            pct   = int(counts[o] / total * 100) if total else 0
            label = f"✅ {o}  {pct}%" if i == opt_idx else f"  {o}  {pct}%"
            poll_kb.append([InlineKeyboardButton(label, callback_data=f"vote|{poll_id}|{i}")])
        try:
            await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup(poll_kb))
        except: pass

        # Yorum iste — isteğe bağlı
        context.user_data["poll_comment_id"] = poll_id
        context.user_data["action"] = "poll_comment"
        await query.message.reply_text(L(uid, "poll_comment_prompt"))
        return WAIT_POLL_COMMENT

    return ConversationHandler.END

# ================================================================
#  METİN GİRİŞ
# ================================================================

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user   = update.effective_user; uid = str(user.id)
    text   = (update.message.text or "").strip()
    register_user(user)

    # Engellenmiş kullanıcı
    if is_blocked(uid) and not is_admin(uid):
        return ConversationHandler.END

    action = context.user_data.get("action","")

    # ── Arama — herkes ──────────────────────────────
    if action == "search":
        context.user_data.pop("action",None)
        if not is_admin(uid):
            log_user_message(user, "search", text)
        await do_search(update.message, uid, text)
        return ConversationHandler.END

    # ── Kişisel Not — Ders adı girişi ─────────────────
    if action == "note_subject_input" and not is_admin(uid):
        subject = "" if text.strip() == "." else text.strip()[:50]
        context.user_data["note_subject"] = subject
        context.user_data["action"] = "note_taking"
        lbl = subject if subject else "ملاحظة"
        kb = [[InlineKeyboardButton("🔒 حفظ وإنهاء", callback_data="notes|done")]]
        sent = await update.message.reply_text(
            f"📖 {lbl}\n\nاكتب ملاحظاتك. اضغط 🔒 عند الانتهاء:",
            reply_markup=InlineKeyboardMarkup(kb))
        context.user_data["last_inline_msg"] = sent.message_id
        return WAIT_FOLDER  # فقط نص

    # ── Kişisel Not — Not alma modu (metin)
    if action == "anon_msg" and not is_admin(uid):
        s = load_settings()
        group_id = s.get("anon_group_id")
        if not group_id:
            await update.message.reply_text(L(uid,"anon_disabled"))
            context.user_data.pop("action",None); return ConversationHandler.END
        u = load_users().get(uid,{})
        name     = u.get("full_name") or u.get("first_name") or f"ID:{uid}"
        username = f"@{u['username']}" if u.get("username") else f"ID:{uid}"
        # Gruba anonim mesaj gönder
        try:
            await context.bot.send_message(
                int(group_id),
                f"✉️ رسالة مجهولة\n\n{text}")
        except Exception as e:
            await update.message.reply_text(f"❌ Hata: {e}")
            context.user_data.pop("action",None); return ConversationHandler.END
        # Süper admin'e kimliği bildir (ayrı mesaj)
        try:
            await context.bot.send_message(
                ADMIN_ID,
                f"🔍 رسالة مجهولة — هوية المُرسِل\n\n"
                f"👤 {name}\n"
                f"📎 {username}\n"
                f"🆔 {uid}\n"
                f"📝 {text[:100]}",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("⚠️ تحذير", callback_data=f"user|warn|{uid}"),
                    InlineKeyboardButton("🚫 حظر",   callback_data=f"user|block|{uid}"),
                ]]))
        except: pass
        await update.message.reply_text(L(uid,"anon_sent"))
        log_user_message(user, "anon", text[:100])
        context.user_data.pop("action",None)
        return ConversationHandler.END

    if action == "note_taking" and not is_admin(uid):
        subject = context.user_data.get("note_subject","")
        add_user_note(uid, "text", text, subject=subject)
        await update.message.reply_text("✅ " + L(uid,"notes_saved"))
        return WAIT_FOLDER  # Sadece metin, devam et

    # ── Kişisel Not — Metin notu kaydet (eski uyumluluk)
    if action == "add_note" and not is_admin(uid):
        note_type = context.user_data.pop("note_type","text")
        subject   = context.user_data.pop("note_subject","")
        add_user_note(uid, note_type, text, subject=subject)
        await update.message.reply_text(L(uid,"notes_saved"))
        context.user_data.pop("action",None)
        return ConversationHandler.END

    # ── Hatırlatıcı — Metin ─────────────────────────
    if action == "reminder_text" and not is_admin(uid):
        context.user_data["reminder_text_pending"] = text
        context.user_data.pop("action", None)
        kb = [
            [InlineKeyboardButton("15 دقيقة", callback_data="reminder|pick_time|15"),
             InlineKeyboardButton("30 دقيقة", callback_data="reminder|pick_time|30"),
             InlineKeyboardButton("1 ساعة",   callback_data="reminder|pick_time|60")],
            [InlineKeyboardButton("2 ساعة",   callback_data="reminder|pick_time|120"),
             InlineKeyboardButton("6 ساعات",  callback_data="reminder|pick_time|360"),
             InlineKeyboardButton("12 ساعة",  callback_data="reminder|pick_time|720")],
            [InlineKeyboardButton("1 يوم",    callback_data="reminder|pick_time|1440"),
             InlineKeyboardButton("2 يوم",    callback_data="reminder|pick_time|2880"),
             InlineKeyboardButton("7 أيام",   callback_data="reminder|pick_time|10080")],
            [InlineKeyboardButton("📝 إدخال يدوي", callback_data="reminder|manual_date")],
            [InlineKeyboardButton("◀️ إلغاء", callback_data="nav|root")],
        ]
        await update.message.reply_text(
            "⏰ متى تريد التذكير؟",
            reply_markup=InlineKeyboardMarkup(kb))
        return ConversationHandler.END

    # ── Hatırlatıcı — Zaman (metin fallback, artık kullanılmıyor) ──
    if action == "reminder_day" and not is_admin(uid):
        try:
            day = int(text.strip())
            assert 0 <= day <= 365
        except:
            await update.message.reply_text("اكتب رقم الأيام (0-365)، 0 = اليوم:")
            return WAIT_FOLDER
        context.user_data["rem_day"] = day
        context.user_data["action"]  = "reminder_hour"
        await update.message.reply_text("🕐 كم ساعة؟ (0-23):")
        return WAIT_FOLDER

    if action == "reminder_hour" and not is_admin(uid):
        try:
            hour = int(text.strip())
            assert 0 <= hour <= 23
        except:
            await update.message.reply_text("اكتب رقم الساعات (0-23):")
            return WAIT_FOLDER
        context.user_data["rem_hour"] = hour
        context.user_data["action"]   = "reminder_minute"
        await update.message.reply_text("⏱ كم دقيقة؟ (0-59):")
        return WAIT_FOLDER

    if action == "reminder_minute" and not is_admin(uid):
        try:
            minute = int(text.strip())
            assert 0 <= minute <= 59
        except:
            await update.message.reply_text("اكتب رقم الدقائق (0-59):")
            return WAIT_FOLDER
        context.user_data["rem_minute"] = minute
        context.user_data["action"]     = "reminder_second"
        await update.message.reply_text("⏱ كم ثانية؟ (0-59):")
        return WAIT_FOLDER

    if action == "reminder_second" and not is_admin(uid):
        try:
            second = int(text.strip())
            assert 0 <= second <= 59
        except:
            await update.message.reply_text("اكتب رقم الثواني (0-59):")
            return WAIT_FOLDER
        day    = context.user_data.pop("rem_day",    0)
        hour   = context.user_data.pop("rem_hour",   0)
        minute = context.user_data.pop("rem_minute", 0)
        total  = (day * 86400) + (hour * 3600) + (minute * 60) + second
        if total <= 0:
            await update.message.reply_text("❌ يجب أن يكون الوقت في المستقبل. حاول مجددًا.")
            context.user_data.pop("action", None)
            return ConversationHandler.END
        import time as _t
        fire_ts  = _t.time() + total
        rem_text = context.user_data.pop("reminder_text_pending",
                   context.user_data.pop("reminder_text", "?"))
        add_reminder(uid, rem_text, fire_ts)
        parts_lbl = []
        if day:    parts_lbl.append(f"{day} يوم")
        if hour:   parts_lbl.append(f"{hour} ساعة")
        if minute: parts_lbl.append(f"{minute} دقيقة")
        if second: parts_lbl.append(f"{second} ثانية")
        label = " و ".join(parts_lbl) if parts_lbl else "قريباً"
        await update.message.reply_text(L(uid, "reminder_saved").format(label))
        context.user_data.pop("action", None)
        return ConversationHandler.END

    if action == "reminder_when" and not is_admin(uid):
        minutes = parse_time_input(text)
        if not minutes:
            await update.message.reply_text(
                "❌ " + ("Geçersiz. Örn: '2 saat', '30 dakika', '1 gün'" if is_main_admin(uid)
                         else "❌ غير صحيح. مثال: '2 ساعة' أو '30 دقيقة'"))
            return WAIT_FOLDER
        import time
        fire_ts = time.time() + (minutes * 60)
        rem_text = context.user_data.pop("reminder_text_pending", context.user_data.pop("reminder_text", text))
        add_reminder(uid, rem_text, fire_ts)
        if minutes < 60:
            label = f"{minutes} دقيقة" if not is_main_admin(uid) else f"{minutes} dakika"
        elif minutes < 1440:
            label = f"{minutes//60} ساعة" if not is_main_admin(uid) else f"{minutes//60} saat"
        else:
            label = f"{minutes//1440} يوم" if not is_main_admin(uid) else f"{minutes//1440} gün"
        await update.message.reply_text(L(uid,"reminder_saved").format(label))
        context.user_data.pop("action",None)
        return ConversationHandler.END

    # ── Anonim Soru ─────────────────────────────────

    if action == "set_class_name" and is_main_admin(uid):
        cls_id = context.user_data.pop("cls_name_id", "")
        if cls_id:
            names = load_class_names()
            names[cls_id] = text.strip()
            save_class_names(names)
            await update.message.reply_text(f"✅ Sınıf {cls_id} adı güncellendi: {text.strip()}")
        context.user_data.pop("action", None)
        return ConversationHandler.END

    if action == "admin_bcast_msg" and is_admin(uid) and not is_main_admin(uid):
        tgt_cls = context.user_data.pop("ab_cls", context.user_data.pop("admin_bcast_cls", None))
        tgt_grp = context.user_data.pop("ab_grp", context.user_data.pop("admin_bcast_grp", None))
        tgt_sft = context.user_data.pop("ab_sft", context.user_data.pop("admin_bcast_sft", ""))
        context.user_data.pop("action", None)
        u_admin  = load_users().get(uid, {})
        aname    = u_admin.get("full_name") or u_admin.get("first_name") or uid
        cls_name = get_class_display_name(tgt_cls) if tgt_cls else "الكل"
        bcast_text = (
            f"📢 إعلان رسمي\n{text}\n🕐 {datetime.now(IRAQ_TZ).strftime('%Y-%m-%d %H:%M')}"
        )
        # Hedef: sınıf + grup filtresi
        if tgt_cls:
            targets = [u for u in users_by_class(tgt_cls) if not is_blocked(u)]
        else:
            targets = [u for u in load_users() if int(u) != ADMIN_ID and not is_blocked(u)]
        if tgt_sft:
            shfts_d = load_shifts()
            targets = [u for u in targets if shfts_d.get(u,"") in (tgt_sft,"sabah" if tgt_sft=="sabahi" else "gece")]
        if tgt_grp:
            grps = load_groups()
            targets = [u for u in targets if grps.get(u,"").startswith(tgt_grp)]
        success = fail = 0
        for uid_ in targets:
            try: await context.bot.send_message(int(uid_), bcast_text); success += 1
            except: fail += 1
        try:
            await context.bot.send_message(
                ADMIN_ID,
                f"📢 إعلان\nالمسؤول: {aname} ({uid})\nالهدف: {cls_name} {tgt_grp or ''}\nالمرسل: {success}\n\n{text[:150]}")
        except: pass
        log_admin_action(uid, "BROADCAST", f"cls:{tgt_cls} grp:{tgt_grp} n:{success}")
        await update.message.reply_text(f"✅ {success} مستخدم")
        return ConversationHandler.END

    if action == "admin_msg_to_super" and is_admin(uid) and not is_main_admin(uid):
        identity = context.user_data.pop("admin_identity","")
        context.user_data.pop("action", None)
        notif = (
            f"📨 رسالة من مسؤول\n\n"
            f"{identity}\n\n"
            f"📝 {text}\n"
            f"🕐 {datetime.now(IRAQ_TZ).strftime('%H:%M')}"
        )
        kb_reply = [[InlineKeyboardButton("💬 رد", callback_data=f"dm_quick|{uid}")]]
        try:
            await context.bot.send_message(ADMIN_ID, notif, reply_markup=InlineKeyboardMarkup(kb_reply))
            await update.message.reply_text("✅ تم إرسال رسالتك للمسؤول الرئيسي.")
        except Exception as e:
            await update.message.reply_text(f"خطأ: {e}")
        return ConversationHandler.END

    if action == "pending_reason" and is_admin(uid) and not is_main_admin(uid):
        reason  = text.strip()
        pact    = context.user_data.pop("pending_action", "")
        ptgt    = context.user_data.pop("pending_target", "")
        tname   = context.user_data.pop("pending_tname",  f"ID:{ptgt}")
        context.user_data.pop("action", None)
        u_a     = load_users().get(uid, {})
        aname   = u_a.get("full_name") or u_a.get("first_name") or f"Admin:{uid}"
        action_ar = "تحذير" if pact == "warn" else "حظر"
        notif = (
            f"{'⚠️' if pact=='warn' else '🚫'} طلب {action_ar}\n"
            f"المسؤول: {aname} ({uid})\n"
            f"المستهدف: {tname} ({ptgt})\n"
            f"Sebep: {reason}\n"
            f"الوقت: {datetime.now(IRAQ_TZ).strftime('%H:%M')}"
        )
        approve_label = "✅ موافق — تحذير" if pact == "warn" else "✅ موافق — حظر"
        kb_approve = [[
            InlineKeyboardButton(approve_label, callback_data=f"approve|{pact}|{ptgt}|{uid}"),
            InlineKeyboardButton("❌ رفض",      callback_data=f"approve|deny|{ptgt}|{uid}|{pact}"),
        ]]
        try:
            await context.bot.send_message(
                ADMIN_ID, notif, reply_markup=InlineKeyboardMarkup(kb_approve))
            await update.message.reply_text("✅ تم إرسال الطلب، في انتظار موافقة المسؤول الرئيسي.")
        except Exception as e:
            await update.message.reply_text(f"خطأ: {e}")
        return ConversationHandler.END

    if action == "user_search_admin" and is_main_admin(uid):
        query_str = text.strip().lstrip("@").lower()
        users_d   = load_users()
        grps_d    = load_groups()
        results   = []
        for uid_, u in users_d.items():
            if int(uid_) == ADMIN_ID: continue
            name  = (u.get("full_name") or u.get("first_name") or "").lower()
            uname = (u.get("username") or "").lower()
            if query_str in name or query_str in uname:
                results.append((uid_, u))
        if not results:
            await update.message.reply_text(f"❌ '{text}' bulunamadı.")
            context.user_data.pop("action", None)
            return ConversationHandler.END
        kb = []
        for uid_, u in results[:30]:
            name  = u.get("full_name") or u.get("first_name") or f"ID:{uid_}"
            un    = f" @{u['username']}" if u.get("username") else ""
            cls_s = get_user_class(uid_) or "?"
            grp_s = grps_d.get(uid_,"?")
            kb.append([InlineKeyboardButton(
                f"👤 {name}{un}  [{cls_s}/{grp_s}]",
                callback_data=f"user|info|{uid_}")])
        kb.append([InlineKeyboardButton("◀️ Geri", callback_data="mgmt|users")])
        await update.message.reply_text(
            f"🔍 '{text}' — {len(results)} sonuç:",
            reply_markup=InlineKeyboardMarkup(kb))
        context.user_data.pop("action", None)
        return ConversationHandler.END

    if action == "set_anon_group" and is_main_admin(uid):
        val = text.strip()
        s = load_settings()
        if val == "0":
            s["anon_group_id"] = None
            await update.message.reply_text("✅ Anonim mesaj özelliği kapatıldı.")
        else:
            try:
                gid = int(val)
                s["anon_group_id"] = gid
                await update.message.reply_text(f"✅ Grup ID kaydedildi: {gid}")
            except:
                await update.message.reply_text("❌ Geçersiz ID. Örn: -1001234567890")
                return ConversationHandler.END
        save_settings(s)
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "cgrp_add_name" and is_main_admin(uid):
        grp_name = text.strip().upper()[:3]
        cls_id   = context.user_data.pop("cgrp_add_cls","1")
        cls_grps = load_class_groups()
        cls_grps.setdefault(cls_id, {})[grp_name] = [f"{grp_name}1",f"{grp_name}2",f"{grp_name}3"]
        save_class_groups(cls_grps)
        await update.message.reply_text(f"✅ Grup {grp_name} eklendi ({grp_name}1/{grp_name}2/{grp_name}3)")
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "cgrp_sub_add_name" and is_main_admin(uid):
        sub_name = text.strip().upper()[:4]
        cls_id   = context.user_data.pop("cgrp_cls","1")
        grp      = context.user_data.pop("cgrp_grp","A")
        cls_grps = load_class_groups()
        cls_grps.setdefault(cls_id,{}).setdefault(grp,[]).append(sub_name)
        save_class_groups(cls_grps)
        await update.message.reply_text(f"✅ {sub_name} eklendi")
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "lab_add_name" and is_admin(uid) and (is_main_admin(uid) or get_admin_perm(uid, "can_countdown")):
        context.user_data["pending_lab_name"] = text.strip()[:60]
        adm_cls = get_admin_cls(uid)
        if adm_cls:
            context.user_data["pending_lab_cls"] = adm_cls
        if not adm_cls and not is_main_admin(uid):
            # Sınıf seç
            kb = []
            for cls_id, cls_def in CLASS_DEFS.items():
                kb.append([InlineKeyboardButton(cls_def["ar"], callback_data=f"lab|lab_cls|{cls_id}")])
            kb.append([InlineKeyboardButton("الكل", callback_data="lab|lab_cls|all")])
            await update.message.reply_text("🔬 السنة الدراسية:", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER
        context.user_data["action"] = "lab_add_date"
        await update.message.reply_text("📅 Tarih yaz (DD/MM/YYYY):\nÖrn: 22/04/2026")
        return WAIT_FOLDER

    if action == "lab_add_date" and is_admin(uid) and (is_main_admin(uid) or get_admin_perm(uid, "can_countdown")):
        date_str = text.strip()
        from datetime import datetime as _dt
        try:
            d = _dt.strptime(date_str, "%d/%m/%Y")
            week_key = d.strftime("%Y-%m-%d")
        except:
            await update.message.reply_text("❌ Format hatalı. Örn: 22/04/2026")
            return ConversationHandler.END
        # Grup seç
        grps_cfg = load_class_groups()
        all_subs = []
        for cls_d in grps_cfg.values():
            for g, subs in cls_d.items():
                all_subs.extend(subs)
        all_subs = sorted(set(all_subs))
        kb = []
        row = []
        for sub in all_subs:
            row.append(InlineKeyboardButton(sub, callback_data=f"lab|set|{week_key}|{sub}"))
            if len(row) == 4:
                kb.append(row); row = []
        if row: kb.append(row)
        kb.append([InlineKeyboardButton("◀️ İptal", callback_data="mgmt|lab")])
        lab_name = context.user_data.get("pending_lab_name","")
        lab_cls  = context.user_data.get("pending_lab_cls","")
        context.user_data["pending_lab_date"] = week_key
        context.user_data.pop("action", None)
        # Sınıfa özel grupları göster
        if lab_cls:
            cls_grps = get_class_groups(lab_cls)
            all_subs_cls = []
            for g, subs in cls_grps.items():
                all_subs_cls.extend(subs)
            all_subs_cls = sorted(set(all_subs_cls))
            kb = []
            row = []
            for sub in all_subs_cls:
                row.append(InlineKeyboardButton(sub, callback_data=f"lab|set|{week_key}|{sub}"))
                if len(row) == 4: kb.append(row); row = []
            if row: kb.append(row)
        kb.append([InlineKeyboardButton("◀️ İptal", callback_data="mgmt|lab")])
        await update.message.reply_text(
            f"📅 {date_str}  🔬 {lab_name}\nHangi grup giriyor?",
            reply_markup=InlineKeyboardMarkup(kb))
        return WAIT_FOLDER

    if action == "countdown_name" and is_admin(uid):
        context.user_data["countdown_name"] = text
        context.user_data.pop("action", None)
        adm_cls = get_admin_cls(uid)
        if not adm_cls and not is_main_admin(uid):
            # Sınıf seç
            kb = []
            for cls_id, cls_def in CLASS_DEFS.items():
                kb.append([InlineKeyboardButton(cls_def["ar"], callback_data=f"cd_cls|{cls_id}")])
            kb.append([InlineKeyboardButton("الكل", callback_data="cd_cls|all")])
            await update.message.reply_text("📅 السنة الدراسية:", reply_markup=InlineKeyboardMarkup(kb))
            return WAIT_FOLDER
        if adm_cls:
            context.user_data["countdown_cls"] = adm_cls
        # Vardiya seç
        kb = [
            [InlineKeyboardButton("الكل",   callback_data="cd_shift|all"),
             InlineKeyboardButton("صباحي", callback_data="cd_shift|sabahi"),
             InlineKeyboardButton("مسائي", callback_data="cd_shift|masaiy")],
        ]
        await update.message.reply_text(
            "لمن هذا الامتحان؟" if not is_main_admin(uid) else "Bu sinav kime?",
            reply_markup=InlineKeyboardMarkup(kb))
        return ConversationHandler.END

    # ── Sınav Tarihi (Admin) ─────────────────────────
    if action == "countdown_date" and is_admin(uid):
        name = context.user_data.pop("countdown_name","?")
        # Admin sınıfına göre hedef
        admin_cls = get_admin_cls(uid) if not is_main_admin(uid) else None
        shift_val = context.user_data.pop("countdown_shift", "")
        grp_val   = context.user_data.pop("countdown_group", "")
        cls_val   = context.user_data.pop("countdown_cls", admin_cls or "")
        ok = add_countdown(name, text, cls=cls_val, shift=shift_val, group=grp_val)
        if ok:
            # Bildirimi süper admin'e gönder
            if not is_main_admin(uid):
                u_ = load_users().get(uid, {})
                adm_name = u_.get("full_name") or u_.get("first_name") or uid
                try:
                    await context.bot.send_message(
                        ADMIN_ID,
                        f"Sinav Eklendi\nAdmin: {adm_name}\nSinav: {name}\nTarih: {text}")
                except: pass
            await update.message.reply_text(L(uid,"countdown_saved").format(name))
        else:
            await update.message.reply_text("❌ Geçersiz tarih. Örn: 20/05/2026")
        context.user_data.pop("action",None)
        return ConversationHandler.END

    # ── Quiz Sorusu (Admin) ──────────────────────────
    if action == "quiz_question" and is_main_admin(uid):
        import re as _re2
        lines = text.strip().split("\n")
        # Format: Soru\nA. ...\nB. ...\nC. ...\nD. ...\nCEVAP: A
        try:
            question = lines[0].strip()
            opts = []
            correct_idx = 0
            for line in lines[1:]:
                line = line.strip()
                m = _re2.match(r'^([A-Da-d])[.)]\s*(.+)$', line)
                if m:
                    opts.append(m.group(2).strip())
                elif line.upper().startswith("CEVAP:") or line.upper().startswith("الجواب:"):
                    ans_letter = line.split(":")[-1].strip().upper()
                    correct_idx = "ABCD".index(ans_letter) if ans_letter in "ABCD" else 0
            if len(opts) < 2:
                await update.message.reply_text("❌ En az 2 seçenek gerekli.\nFormat:\nSoru\nA. Seç1\nB. Seç2\nCEVAP: A")
                return WAIT_FOLDER
            qid = f"q_{datetime.now(IRAQ_TZ).strftime('%Y%m%d%H%M%S')}"
            quizzes = load_quizzes()
            quizzes.append({
                "id": qid, "question": question, "options": opts,
                "correct": correct_idx, "active": True,
                "answered": {}, "created": datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M")
            })
            save_quizzes(quizzes)
            await update.message.reply_text(
                f"✅ Quiz oluşturuldu!\n\n"
                f"❓ {question}\n"
                f"\n" +
                "\n".join(f"{'ABCD'[i]}. {o}" for i,o in enumerate(opts)) +
                f"\n\n✅ Doğru: {'ABCD'[correct_idx]}")
        except Exception as e:
            await update.message.reply_text(f"❌ Hata: {e}")
        context.user_data.pop("action",None)
        return ConversationHandler.END

    # ── Anonim Soruya Cevap (Admin) ──────────────────

    # ── Kullanıcıdan admin'e mesaj ──────────────────
    if action == "user_msg_to_admin" and not is_admin(uid):
        log_user_message(user, "msg", text)
        users_d  = load_users()
        u        = users_d.get(uid, {})
        name     = u.get("full_name") or u.get("first_name") or f"ID:{uid}"
        un       = f" @{u['username']}" if u.get("username") else ""
        cls_id   = get_user_class(uid)
        grp_id   = get_user_group(uid)
        cls_name = get_class_display_name(cls_id) if cls_id else ""
        msg_to   = context.user_data.pop("msg_to", "main")

        notif = (
            f"💬 رسالة جديدة\n\n"
            f"👤 {name}{un}\n"
            f"🆔 {uid}\n"
            f"🎓 {cls_name}  👥 {grp_id or '-'}\n"
            f"\n"
            f"📝 {text}\n"
            f"🕐 {datetime.now(IRAQ_TZ).strftime('%H:%M')}"
        )
        kb_admin = [[InlineKeyboardButton("💬 رد", callback_data=f"dm_quick|{uid}")]]

        # Süper admin her zaman alır
        try:
            await context.bot.send_message(
                ADMIN_ID, notif, reply_markup=InlineKeyboardMarkup(kb_admin))
        except Exception as e:
            logger.error(f"Süper admin bildirimi: {e}")

        # Sınıf adminine de gönder
        if cls_id:
            perms_data = load_admin_perms()
            for adm_id_str, adm_perms in perms_data.items():
                if adm_perms.get("cls") == cls_id and adm_id_str != str(ADMIN_ID):
                    try:
                        await context.bot.send_message(
                            int(adm_id_str), notif,
                            reply_markup=InlineKeyboardMarkup(kb_admin))
                    except: pass

        # Kullanıcıya onay
        if msg_to == "class_admin":
            conf = "✅ تم إرسال رسالتك لمسؤول صفك وللمسؤول الرئيسي."
        else:
            conf = L(uid, "msg_to_admin_sent")
        await update.message.reply_text(conf)
        context.user_data.pop("action", None)
        return ConversationHandler.END

    # ── Açık uçlu anket cevabı ────────────────────────
    if action == "open_poll_answer" and not is_admin(uid):
        poll_id = context.user_data.pop("open_poll_id", None)
        context.user_data.pop("action", None)
        if poll_id and text:
            polls = load_polls()
            poll  = polls.get(poll_id)
            if poll:
                users_d = load_users()
                u       = users_d.get(uid, {})
                name    = u.get("full_name") or u.get("first_name") or f"ID:{uid}"
                un      = f"@{u['username']}" if u.get("username") else ""
                poll.setdefault("answers", {})[uid] = {
                    "name":    name,
                    "username": un,
                    "answer":  text[:500],
                    "time":    datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M"),
                }
                save_polls(polls)
                await update.message.reply_text(L(uid, "poll_comment_saved"))
        return ConversationHandler.END

    # ── Anket yorumu (kullanıcıdan) ───────────────────
    if action == "poll_comment" and not is_admin(uid):
        poll_id = context.user_data.pop("poll_comment_id", None)
        context.user_data.pop("action", None)
        if poll_id:
            polls = load_polls()
            poll  = polls.get(poll_id)
            if poll:
                # Yorumu kaydet: {uid: {"vote": "seçenek", "comment": "yorum", "time": ...}}
                comments = poll.setdefault("comments", {})
                users_d  = load_users()
                u        = users_d.get(uid, {})
                name     = u.get("full_name") or u.get("first_name") or f"ID:{uid}"
                un       = f"@{u['username']}" if u.get("username") else ""
                if text.strip() == "/skip":
                    await update.message.reply_text(L(uid, "poll_comment_skip"))
                else:
                    comments[uid] = {
                        "name":    name,
                        "username": un,
                        "vote":    poll.get("votes", {}).get(uid, "—"),
                        "comment": text[:300],
                        "time":    datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M"),
                    }
                    save_polls(polls)
                    await update.message.reply_text(L(uid, "poll_comment_saved"))
        return ConversationHandler.END

    # ── FAQ Soru Girişi (admin) ───────────────────────
    if action == "faq_q" and is_main_admin(uid):
        if not text: await update.message.reply_text(TR["faq_enter_q"]); return WAIT_FAQ_Q
        context.user_data["faq_q_text"] = text
        context.user_data["action"]     = "faq_a"
        await update.message.reply_text(TR["faq_enter_a"])
        return WAIT_FAQ_A

    if action == "faq_a" and is_main_admin(uid):
        q    = context.user_data.pop("faq_q_text", "")
        kws  = [kw.strip() for kw in q.split(",") if kw.strip()]
        if not kws: kws = [q]
        faqs = load_faq()
        faqs.append({"keywords": kws, "answer": text})
        save_faq(faqs)
        await update.message.reply_text(TR["faq_saved"])
        log_admin_action(uid, "FAQ_ADD", kws[0][:40])
        context.user_data.pop("action", None)
        return ConversationHandler.END

    # ── Otomatik Kural Girişi (admin) ────────────────
    if action == "rule_kw" and is_main_admin(uid):
        if not text: await update.message.reply_text(TR["rule_enter_kw"]); return WAIT_RULE_KW
        context.user_data["rule_kws"] = [kw.strip() for kw in text.split(",") if kw.strip()]
        context.user_data["action"]   = "rule_resp"
        await update.message.reply_text(TR["rule_enter_resp"])
        return WAIT_RULE_RESP

    if action == "rule_resp" and is_main_admin(uid):
        kws   = context.user_data.pop("rule_kws", [])
        rules = load_auto_rules()
        rules.append({"keywords": kws, "response": text})
        save_auto_rules(rules)
        await update.message.reply_text(TR["rule_saved"])
        log_admin_action(uid, "RULE_ADD", ", ".join(kws[:3]))
        context.user_data.pop("action", None)
        return ConversationHandler.END

    # ── Uyarı Sebebi (admin) ─────────────────────────
    if action == "warn_reason" and is_main_admin(uid):
        target   = context.user_data.pop("warn_target","")
        req_admin = context.user_data.pop("warn_req_admin","")
        if target:
            count = add_warn(target, text, uid)
            await update.message.reply_text(TR["user_warned"].format(target, count, MAX_WARNS))
            log_admin_action(uid, "WARN", f"→{target}: {text[:40]}")
            # Kullanıcıya bildir
            try:
                await context.bot.send_message(
                    int(target),
                    AR["warn_msg_to_user"].format(reason=text, count=count, max=MAX_WARNS))
            except: pass
            # Talep eden alt admin'e bildir
            if req_admin:
                try:
                    u = load_users().get(target,{})
                    name = u.get("full_name") or u.get("first_name") or f"ID:{target}"
                    await context.bot.send_message(
                        int(req_admin),
                        f"✅ Uyarı onaylandı ve gönderildi.\n"
                        f"👤 {name} ({target})\n"
                        f"📋 Sebep: {text[:80]}\n"
                        f"⚠️ Uyarı sayısı: {count}/{MAX_WARNS}")
                except: pass
            # Otomatik engel
            if count >= MAX_WARNS:
                blocked = load_blocked()
                if int(target) not in blocked:
                    blocked.append(int(target)); save_blocked(blocked)
                await update.message.reply_text(TR["user_auto_blocked"].format(target, MAX_WARNS))
                log_admin_action(uid, "AUTO_BAN", f"→{target}")
        context.user_data.pop("action", None)
        return ConversationHandler.END

    # ── Kullanıcı Notu (admin) ───────────────────────
    if action == "user_note" and is_main_admin(uid):
        target = context.user_data.pop("note_target","")
        if target:
            set_note(target, text)
            msg = TR["note_saved"] if text.strip() else TR["note_cleared"]
            await update.message.reply_text(msg)
            log_admin_action(uid, "NOTE", f"→{target}")
        context.user_data.pop("action", None)
        return ConversationHandler.END

    # ── Klasör Açıklaması (admin) ────────────────────
    if action == "folder_desc" and is_main_admin(uid):
        path = context.user_data.get("path", [])
        set_folder_desc(path, text)
        msg = L(uid,"folder_desc_saved") if text.strip() else L(uid,"folder_desc_cleared")
        await update.message.reply_text(msg)
        log_admin_action(uid, "FOLDER_DESC", "›".join(path))
        context.user_data.pop("action", None)
        return ConversationHandler.END

    # ── Whitelist Kullanıcı Ekle (admin) ─────────────
    if action == "secret_add_id" and is_main_admin(uid):
        try:
            new_id = int(text.strip())
            wl     = load_whitelist()
            users_list = wl.setdefault("users", [])
            if new_id not in users_list:
                users_list.append(new_id); save_whitelist(wl)
            await update.message.reply_text(TR["secret_added"].format(new_id))
            log_admin_action(uid, "WHITELIST_ADD", str(new_id))
        except:
            await update.message.reply_text(TR["invalid_id"])
        context.user_data.pop("action", None)
        return ConversationHandler.END

    # ── Hedefli Broadcast (admin) ────────────────────
    if action == "broadcast_targeted" and is_main_admin(uid):
        targets = context.user_data.pop("broadcast_targets", [])
        success = fail = 0
        for uid_ in targets:
            try:
                bcast_msg = (
                    f"╔══════════════════════╗\n"
                    f"  📢  إعلان رسمي\n"
                    f"╠══════════════════════╣\n\n"
                    f"{text}\n\n"
                    f"╚══════════════════════╝\n"
                    f"  🕐  {datetime.now(IRAQ_TZ).strftime('%Y-%m-%d %H:%M')}"
                )
                await context.bot.send_message(int(uid_), bcast_msg); success += 1
            except: fail += 1
        await update.message.reply_text(TR["broadcast_done"].format(success, fail))
        log_admin_action(uid, "BROADCAST_TARGETED", f"{success} gönderildi")
        # Duyuru log kaydet
        bcast_log = load_bcast_log()
        bcast_log.append({
            "time":  datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M"),
            "text":  text[:100],
            "count": success,
            "type":  "targeted"
        })
        save_bcast_log(bcast_log)
        context.user_data.pop("action", None)
        return ConversationHandler.END

    # ── Anket sorusu ──────────────────────────────────
    if action == "poll_question" and is_admin(uid):
        if not text:
            await update.message.reply_text(TR["poll_enter_q"]); return WAIT_POLL_QUESTION
        context.user_data["poll_question"] = text
        poll_type = context.user_data.get("poll_type", "choice")

        if poll_type == "open":
            # Açık uçlu — seçenek istemiyoruz, direkt kaydet
            q       = text
            poll_id = make_poll_id()
            polls   = load_polls()
            polls[poll_id] = {
                "question": q, "type": "open",
                "options": [], "votes": {}, "answers": {}, "active": False,
                "target": context.user_data.get("poll_target","all")
            }
            save_polls(polls)
            kb = [
                [InlineKeyboardButton(TR["poll_send_btn"],   callback_data=f"poll|send|{poll_id}")],
                [InlineKeyboardButton(TR["poll_cancel_btn"], callback_data="close")],
            ]
            await update.message.reply_text(
                f"✍️ Açık Uçlu Anket Önizleme:\n\n❓ {q}\n\nKullanıcılar yazılı cevap verecek.\n\nGönderilsin mi?",
                reply_markup=InlineKeyboardMarkup(kb))
            context.user_data.pop("action", None)
            context.user_data.pop("poll_question", None)
            context.user_data.pop("poll_type", None)
            return ConversationHandler.END
        else:
            # Çoktan seçmeli — seçenek iste
            context.user_data["action"] = "poll_options"
            await update.message.reply_text(TR["poll_enter_opts"])
            return WAIT_POLL_OPTIONS

    # ── Anket seçenekleri ─────────────────────────────
    if action == "poll_options" and is_admin(uid):
        opts = [o.strip() for o in text.split("\n") if o.strip()][:6]
        if len(opts) < 2:
            await update.message.reply_text("❌ En az 2 seçenek girin!"); return WAIT_POLL_OPTIONS
        q       = context.user_data.get("poll_question", "?")
        poll_id = make_poll_id()
        # Anketi kaydet (henüz gönderilmedi)
        polls = load_polls()
        polls[poll_id] = {"question": q, "options": opts, "votes": {}, "active": False, "target": context.user_data.get("poll_target","all")}
        save_polls(polls)
        opts_text = "\n".join(f"  {i+1}. {o}" for i, o in enumerate(opts))
        kb = [
            [InlineKeyboardButton(TR["poll_send_btn"],   callback_data=f"poll|send|{poll_id}")],
            [InlineKeyboardButton(TR["poll_cancel_btn"], callback_data="close")],
        ]
        await update.message.reply_text(
            TR["poll_preview"].format(q=q, opts=opts_text),
            reply_markup=InlineKeyboardMarkup(kb))
        context.user_data.pop("action", None)
        context.user_data.pop("poll_question", None)
        return ConversationHandler.END

    # ── Admin değilse: AI motorunu çalıştır ─────────────────
    if not is_admin(uid):
        if not text: return ConversationHandler.END

        # Spam koruması
        if not check_rate_limit(uid):
            await update.message.reply_text(L(uid, "spam_warning"))
            return ConversationHandler.END

        # Sabit mesaj (ilk kez)
        if context.user_data.get("shown_pinned") is None:
            pinned = get_pinned_msg()
            if pinned:
                context.user_data["shown_pinned"] = True
                await update.message.reply_text(L(uid, "pinned_msg_label").format(pinned))

        # Mesajı kaydet
        log_user_message(user, "msg", text)
        return ConversationHandler.END

    # ── Admin işlemleri ──────────────────────────────
    content = load_content(); path = context.user_data.get("path",[]); folder = get_folder(content,path)

    if action == "admin_add" and is_main_admin(uid):
        raw = text.strip()
        new_id = None
        # ID mi yoksa @kullanıcı adı mı?
        if raw.isdigit() or (raw.startswith("-") and raw[1:].isdigit()):
            new_id = int(raw)
        else:
            # @username ile ara
            username_search = raw.lstrip("@").lower()
            users_d = load_users()
            for uid_, u in users_d.items():
                if u.get("username","").lower() == username_search:
                    new_id = int(uid_)
                    break
            if not new_id:
                await update.message.reply_text(
                    f"❌ @{username_search} bulunamadı. ID veya @kullanıcı adı girin:")
                return WAIT_ADMIN_ID
        admins = load_admins()
        if new_id in admins or new_id == ADMIN_ID:
            await update.message.reply_text(TR["admin_exists"])
            context.user_data.pop("action",None); return ConversationHandler.END
        admins.append(new_id); save_admins(admins)
        # Önceden seçilen sınıf/grup varsa uygula
        pre_cls = context.user_data.pop("pending_admin_cls", None)
        pre_grp = context.user_data.pop("pending_admin_grp", None)
        if pre_cls: set_admin_perm(str(new_id), "cls", pre_cls)
        if pre_grp: set_admin_perm(str(new_id), "grp", pre_grp)
        log_admin_action(uid, "ADD_ADMIN", f"ID:{new_id} cls:{pre_cls} grp:{pre_grp}")
        found_user = load_users().get(str(new_id), {})
        display = found_user.get("full_name") or f"@{found_user.get('username','')}" or str(new_id)
        cls_lbl = f"Sinif: {pre_cls}" if pre_cls else "Kisitsiz"
        grp_lbl = f"  Grup: {pre_grp}" if pre_grp else ""
        await update.message.reply_text(
            f"✅ Admin eklendi: {display} ({new_id})\n{cls_lbl}{grp_lbl}\n\n"
            f"Yetkiler: Ayarlar → Admin Yetkileri")
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "dm_manual_id" and is_main_admin(uid):
        try: target_id = str(int(text))
        except: await update.message.reply_text(TR["invalid_id"]); return WAIT_DM_USER_ID
        context.user_data["dm_target"] = target_id; context.user_data["action"] = "dm_msg"
        await update.message.reply_text(TR["dm_enter_msg"].format(target_id))
        return WAIT_DM_MSG

    if action == "dm_msg" and is_main_admin(uid):
        target = context.user_data.get("dm_target")
        if not target: await update.message.reply_text(TR["dm_no_target"]); return ConversationHandler.END
        dm_formatted = (
            f"💌  رسالة شخصية\n"
            f"\n"
            f"{text}\n"
            f"\n"
            f"🕐 {datetime.now(IRAQ_TZ).strftime('%Y-%m-%d %H:%M')}"
        )
        try:
            await context.bot.send_message(int(target), dm_formatted)
            await update.message.reply_text(TR["dm_sent"].format(target))
        except Exception as e:
            await update.message.reply_text(TR["dm_fail"].format(e))
        context.user_data.pop("action",None); context.user_data.pop("dm_target",None)
        return ConversationHandler.END

    if action == "broadcast" and is_main_admin(uid):
        users = load_users(); success = fail = 0
        for uid_,u in users.items():
            if int(uid_) == ADMIN_ID or is_blocked(uid_): continue
            try:
                bcast_msg = (
                    f"╔══════════════════════╗\n"
                    f"  📢  إعلان رسمي\n"
                    f"╠══════════════════════╣\n\n"
                    f"{text}\n\n"
                    f"╚══════════════════════╝\n"
                    f"  🕐  {datetime.now(IRAQ_TZ).strftime('%Y-%m-%d %H:%M')}"
                )
                await context.bot.send_message(int(uid_), bcast_msg); success += 1
            except: fail += 1
        await update.message.reply_text(TR["broadcast_done"].format(success,fail))
        # Log kaydet
        bcast_log = load_bcast_log()
        bcast_log.append({
            "time":  datetime.now(IRAQ_TZ).strftime("%Y-%m-%d %H:%M"),
            "text":  text[:100],
            "count": success,
            "type":  "all"
        })
        save_bcast_log(bcast_log)
        log_admin_action(uid, "BROADCAST_ALL", f"{success} gönderildi")
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "set_name" and is_main_admin(uid):
        if not text: await update.message.reply_text(TR["folder_empty"]); return WAIT_BOT_NAME
        s = load_settings(); s["bot_name"] = text; save_settings(s)
        await update.message.reply_text(TR["set_name_ok"].format(text))
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "set_description" and is_main_admin(uid):
        if not text: await update.message.reply_text(TR["folder_empty"]); return WAIT_BOT_NAME
        s = load_settings(); s["bot_description"] = text; save_settings(s)
        await update.message.reply_text(f"✅ Açıklama güncellendi: {text[:80]}")
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "set_welcome" and is_main_admin(uid):
        s = load_settings(); s["welcome_msg"] = text; save_settings(s)
        await update.message.reply_text(TR["set_welcome_ok"])
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "set_maint_text" and is_main_admin(uid):
        s = load_settings(); s["maintenance_text"] = text; save_settings(s)
        await update.message.reply_text(f"✅ Bakım mesajı güncellendi.")
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "set_pin_msg" and is_main_admin(uid):
        set_pinned_msg(text)
        msg_txt = TR["pin_msg_saved"] if text.strip() else "✅ Sabit mesaj silindi."
        await update.message.reply_text(msg_txt)
        log_admin_action(uid, "PIN_MSG", text[:50])
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "add_folder":
        if not text: await update.message.reply_text(L(uid,"folder_empty")); return WAIT_FOLDER
        # Çoklu satır — her satır ayrı klasör
        names = [n.strip() for n in text.split("\n") if n.strip()]
        added_names = []
        for name in names:
            if name and name not in folder.get("folders",{}):
                folder.setdefault("folders",{})[name] = {"folders":{},"files":[]}
                added_names.append(name)
        if added_names:
            save_content(content)
        if not added_names:
            await update.message.reply_text(L(uid,"folder_exists").format(text)); return WAIT_FOLDER
        text = ", ".join(added_names)  # sayaç mesajı için
        count     = context.user_data.get("add_folder_count", 0) + 1
        context.user_data["add_folder_count"] = count
        status_id = context.user_data.get("add_folder_status_id")
        kb = [[InlineKeyboardButton(L(uid,"close"), callback_data="nav|root")]]
        status_text = (
            f"{'✅ Klasörler ekleniyor...' if is_main_admin(uid) else '✅ جارٍ إضافة المجلدات...'}\n"
            f"\n"
            f"📁 {count} {'klasör eklendi' if is_main_admin(uid) else 'مجلد تمت إضافته'}\n\n"
            f"{'Devam yaz, bitince Kapat.' if is_main_admin(uid) else 'اكتب المزيد أو اضغط إغلاق.'}"
        )
        if status_id:
            try:
                await context.bot.edit_message_text(
                    status_text, chat_id=update.effective_chat.id,
                    message_id=status_id, reply_markup=InlineKeyboardMarkup(kb))
            except:
                sent = await update.message.reply_text(status_text, reply_markup=InlineKeyboardMarkup(kb))
                context.user_data["add_folder_status_id"] = sent.message_id
        else:
            sent = await update.message.reply_text(status_text, reply_markup=InlineKeyboardMarkup(kb))
            context.user_data["add_folder_status_id"] = sent.message_id
        return WAIT_FOLDER

    if action == "rename_folder" and is_main_admin(uid):
        old = context.user_data.get("target","")
        if not text: await update.message.reply_text(L(uid,"folder_empty")); return WAIT_RENAME_FOLDER
        if text in folder.get("folders",{}):
            await update.message.reply_text(L(uid,"folder_exists").format(text)); return WAIT_RENAME_FOLDER
        if old in folder.get("folders",{}):
            folder["folders"][text] = folder["folders"].pop(old); save_content(content)
        await update.message.reply_text(L(uid,"folder_renamed").format(old,text))
        context.user_data.pop("action",None); context.user_data.pop("target",None)
        return ConversationHandler.END

    if action == "rename_file" and is_main_admin(uid):
        idx = context.user_data.get("target",0); files = folder.get("files",[])
        if not text: await update.message.reply_text(L(uid,"folder_empty")); return WAIT_RENAME_FILE
        if 0 <= idx < len(files):
            files[idx]["caption"] = text; files[idx]["name"] = text; save_content(content)
        await update.message.reply_text(L(uid,"file_renamed").format(text))
        context.user_data.pop("action",None); context.user_data.pop("target",None)
        return ConversationHandler.END

    if action == "add_file":
        if text.startswith("http://") or text.startswith("https://"):
            fobj    = {"type":"link","url":text,"caption":text,"name":text}
        else:
            fobj    = {"type":"text","caption":text,"name":text}
        folder.setdefault("files",[]).append(fobj); save_content(content)
        count     = context.user_data.get("add_file_count", 0) + 1
        context.user_data["add_file_count"] = count
        status_id = context.user_data.get("add_file_status_id")
        kb = [[InlineKeyboardButton(L(uid,"close"), callback_data="nav|root")]]
        status_text = (
            f"{'✅ Dosyalar ekleniyor...' if is_main_admin(uid) else '✅ جارٍ إضافة الملفات...'}\n"
            f"\n"
            f"📎 {count} {'dosya/link eklendi' if is_main_admin(uid) else 'ملف/رابط تمت إضافته'}\n\n"
            f"{'Devam gönder, bitince Kapat.' if is_main_admin(uid) else 'أرسل المزيد أو اضغط إغلاق.'}"
        )
        if status_id:
            try:
                await context.bot.edit_message_text(
                    status_text, chat_id=update.effective_chat.id,
                    message_id=status_id, reply_markup=InlineKeyboardMarkup(kb))
            except:
                sent = await update.message.reply_text(status_text, reply_markup=InlineKeyboardMarkup(kb))
                context.user_data["add_file_status_id"] = sent.message_id
        else:
            sent = await update.message.reply_text(status_text, reply_markup=InlineKeyboardMarkup(kb))
            context.user_data["add_file_status_id"] = sent.message_id
        return WAIT_FILE

    return ConversationHandler.END

# ================================================================
#  MEDYA GİRİŞ
# ================================================================

async def handle_media(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user; uid = str(user.id); msg = update.message
    register_user(user)

    if not is_main_admin(uid):
        if is_blocked(uid): return ConversationHandler.END
        s = load_settings()
        if s["maintenance"] and not is_admin(uid):
            await msg.reply_text(s.get("maintenance_text","🔧")); return ConversationHandler.END
        if   msg.photo:
            fid = msg.photo[-1].file_id; cap = msg.caption or "(photo)"
            log_user_message(user,"photo", cap, file_id=fid)
        elif msg.video:
            fid = msg.video.file_id; cap = msg.caption or msg.video.file_name or "(video)"
            log_user_message(user,"video", cap, file_id=fid)
        elif msg.document:
            fid = msg.document.file_id; cap = msg.document.file_name or "(doc)"
            log_user_message(user,"document", cap, file_id=fid)
        elif msg.audio:
            fid = msg.audio.file_id; cap = msg.audio.file_name or "(audio)"
            log_user_message(user,"audio", cap, file_id=fid)
        elif msg.voice:
            fid = msg.voice.file_id
            log_user_message(user,"voice","(voice)", file_id=fid)
        elif msg.animation:
            fid = msg.animation.file_id
            log_user_message(user,"animation","(gif)", file_id=fid)
        elif msg.sticker:
            fid = msg.sticker.file_id; emoji = msg.sticker.emoji or "(sticker)"
            log_user_message(user,"sticker", emoji, file_id=fid)
        return ConversationHandler.END

    action = context.user_data.get("action","")

    if action == "broadcast" and is_main_admin(uid):
        users = load_users(); success = fail = 0
        cap_raw  = msg.caption or ""
        cap_text = (
            f"📢  إعلان رسمي\n"
            f"\n"
            f"{cap_raw}\n"
            f"\n"
            f"🕐 {datetime.now(IRAQ_TZ).strftime('%Y-%m-%d %H:%M')}"
        ) if cap_raw else None
        for uid_,u in users.items():
            if int(uid_) == ADMIN_ID or is_blocked(uid_): continue
            try:
                if   msg.photo:     await context.bot.send_photo(int(uid_),    msg.photo[-1].file_id, caption=cap_text)
                elif msg.video:     await context.bot.send_video(int(uid_),    msg.video.file_id,     caption=cap_text)
                elif msg.document:  await context.bot.send_document(int(uid_), msg.document.file_id,  caption=cap_text)
                elif msg.audio:     await context.bot.send_audio(int(uid_),    msg.audio.file_id,     caption=cap_text)
                elif msg.voice:     await context.bot.send_voice(int(uid_),    msg.voice.file_id)
                elif msg.animation: await context.bot.send_animation(int(uid_),msg.animation.file_id, caption=cap_text)
                success += 1
            except: fail += 1
        await msg.reply_text(TR["broadcast_done"].format(success,fail))
        context.user_data.pop("action",None); return ConversationHandler.END

    if action == "dm_msg" and is_main_admin(uid):
        target = context.user_data.get("dm_target"); cap_raw = msg.caption or ""
        if not target: await msg.reply_text(TR["dm_no_target"]); return ConversationHandler.END
        cap_text = (
            f"💌  رسالة شخصية\n"
            f"\n"
            f"{cap_raw}\n"
            f"\n"
            f"🕐 {datetime.now(IRAQ_TZ).strftime('%Y-%m-%d %H:%M')}"
        ) if cap_raw else None
        try:
            if   msg.photo:     await context.bot.send_photo(int(target),    msg.photo[-1].file_id, caption=cap_text)
            elif msg.video:     await context.bot.send_video(int(target),    msg.video.file_id,     caption=cap_text)
            elif msg.document:  await context.bot.send_document(int(target), msg.document.file_id,  caption=cap_text)
            elif msg.audio:     await context.bot.send_audio(int(target),    msg.audio.file_id,     caption=cap_text)
            elif msg.voice:     await context.bot.send_voice(int(target),    msg.voice.file_id)
            elif msg.animation: await context.bot.send_animation(int(target),msg.animation.file_id, caption=cap_text)
            await msg.reply_text(TR["dm_sent"].format(target))
        except Exception as e:
            await msg.reply_text(TR["dm_fail"].format(e))
        context.user_data.pop("action",None); context.user_data.pop("dm_target",None)
        return ConversationHandler.END

    if action in ("add_note","note_taking") and not is_admin(uid):
        subject = context.user_data.get("note_subject","")
        fid = None; cap = msg.caption or ""; note_type = "media"
        if msg.photo:
            fid = msg.photo[-1].file_id; note_type = "photo"
        elif msg.video:
            fid = msg.video.file_id; note_type = "video"
        elif msg.voice:
            fid = msg.voice.file_id; note_type = "voice"
        elif msg.audio:
            fid = msg.audio.file_id; note_type = "audio"
        elif msg.document:
            fid = msg.document.file_id; cap = cap or msg.document.file_name or ""; note_type = "document"
        elif msg.video_note:
            fid = msg.video_note.file_id; note_type = "video"
        add_user_note(uid, note_type, cap, file_id=fid, subject=subject)
        await msg.reply_text("✅ " + L(uid,"notes_saved"))
        if action == "note_taking":
            return WAIT_FILE  # Not alma modu devam
        context.user_data.pop("action",None); context.user_data.pop("note_subject",None)
        return ConversationHandler.END

    if action == "set_photo" and is_main_admin(uid):
        if not msg.photo:
            await msg.reply_text(TR["set_photo_fail"]); return WAIT_PHOTO
        fid = msg.photo[-1].file_id
        s = load_settings(); s["bot_photo_id"] = fid; save_settings(s)
        await msg.reply_text(TR["set_photo_ok"])
        context.user_data.pop("action",None); return ConversationHandler.END

    if action != "add_file": return ConversationHandler.END

    content = load_content(); path = context.user_data.get("path",[]); folder = get_folder(content,path)
    now  = datetime.now(IRAQ_TZ).strftime("%Y%m%d%H%M%S")
    cap  = msg.caption or ""
    fobj = None

    # ── Medya grubu (aynı anda birden fazla dosya) ───
    media_group_id = msg.media_group_id
    if media_group_id:
        # Grubun dosyalarını geçici olarak topla
        group_key = f"mg_{media_group_id}"
        if group_key not in context.user_data:
            context.user_data[group_key] = []
            # 2 saniye sonra hepsini kaydet
            async def _flush_group(gid, gkey, p, c):
                import asyncio
                await asyncio.sleep(2)
                items   = context.user_data.pop(gkey, [])
                if not items: return
                ct      = load_content()
                fld     = get_folder(ct, p)
                added   = 0
                for item in items:
                    fld.setdefault("files", []).append(item)
                    added += 1
                save_content(ct)
                kb = [[InlineKeyboardButton(L(str(user.id),"close"), callback_data="nav|root")]]
                try:
                    await context.bot.send_message(
                        user.id,
                        f"✅ {added} dosya eklendi!\n\n📎 " +
                        ("Daha fazla dosya gönderebilirsiniz." if is_main_admin(str(user.id)) else
                         "يمكنك إرسال المزيد من الملفات."),
                        reply_markup=InlineKeyboardMarkup(kb))
                except Exception as e:
                    logger.error(f"Grup flush hatası: {e}")
            import asyncio
            asyncio.create_task(_flush_group(media_group_id, group_key, path, content))

    try:
        if msg.photo:
            fid = msg.photo[-1].file_id; local = await download_file(context,fid,f"photo_{now}.jpg")
            fobj = {"type":"photo","file_id":fid,"caption":cap or f"photo_{now}","name":f"photo_{now}.jpg",
                    **({"local_path":local} if local else {})}
        elif msg.video:
            fname = msg.video.file_name or f"video_{now}.mp4"; fid = msg.video.file_id
            local = await download_file(context,fid,fname)
            fobj = {"type":"video","file_id":fid,"caption":cap or fname,"name":fname,
                    **({"local_path":local} if local else {})}
        elif msg.animation:
            fid = msg.animation.file_id; local = await download_file(context,fid,f"gif_{now}.gif")
            fobj = {"type":"animation","file_id":fid,"caption":cap or "GIF","name":f"gif_{now}.gif",
                    **({"local_path":local} if local else {})}
        elif msg.document:
            fname = msg.document.file_name or f"doc_{now}"; fid = msg.document.file_id
            local = await download_file(context,fid,fname)
            fobj = {"type":"document","file_id":fid,"caption":cap or fname,"name":fname,
                    **({"local_path":local} if local else {})}
        elif msg.audio:
            fname = msg.audio.file_name or f"audio_{now}.mp3"; fid = msg.audio.file_id
            local = await download_file(context,fid,fname)
            fobj = {"type":"audio","file_id":fid,"caption":cap or msg.audio.title or fname,"name":fname,
                    **({"local_path":local} if local else {})}
        elif msg.voice:
            fid = msg.voice.file_id; local = await download_file(context,fid,f"voice_{now}.ogg")
            fobj = {"type":"voice","file_id":fid,"caption":cap or "voice","name":f"voice_{now}.ogg",
                    **({"local_path":local} if local else {})}
    except Exception as e:
        logger.error(f"handle_media hata: {e}")

    if fobj:
        # Medya grubuysa listeye ekle, flush_group kaydedecek
        if media_group_id:
            context.user_data[f"mg_{media_group_id}"].append(fobj)
            return WAIT_FILE

        # Tekli dosya — direkt kaydet
        folder.setdefault("files",[]).append(fobj); save_content(content)

        # Sayacı güncelle veya yeni mesaj gönder
        status_id = context.user_data.get("add_file_status_id")
        count     = context.user_data.get("add_file_count", 0) + 1
        context.user_data["add_file_count"] = count
        kb = [[InlineKeyboardButton(L(uid,"close"), callback_data="nav|root")]]
        status_text = (
            f"{'✅ Dosyalar ekleniyor...' if is_main_admin(uid) else '✅ جارٍ إضافة الملفات...'}\n"
            f"\n"
            f"📎 {count} {'dosya eklendi' if is_main_admin(uid) else 'ملف تمت إضافته'}\n\n"
            f"{'📤 Devam gönder, bitince Kapat.' if is_main_admin(uid) else '📤 أرسل المزيد أو اضغط إغلاق.'}"
        )
        if status_id:
            try:
                await context.bot.edit_message_text(
                    status_text, chat_id=msg.chat_id,
                    message_id=status_id, reply_markup=InlineKeyboardMarkup(kb))
            except:
                sent = await msg.reply_text(status_text, reply_markup=InlineKeyboardMarkup(kb))
                context.user_data["add_file_status_id"] = sent.message_id
        else:
            sent = await msg.reply_text(status_text, reply_markup=InlineKeyboardMarkup(kb))
            context.user_data["add_file_status_id"] = sent.message_id
        return WAIT_FILE
    else:
        await msg.reply_text(L(uid,"unsupported"))
    return ConversationHandler.END

# ================================================================
#  GENEL MESAJ HANDLER — fallback log + güvenlik
# ================================================================

async def handle_any_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Fallback handler — sadece loglama ve güvenlik için."""
    if not update.message: return
    user = update.effective_user; uid = str(user.id); msg = update.message
    text = (msg.text or "").strip()

    if text in ALL_BTNS: return
    register_user(user)
    if is_blocked(uid): return

    s = load_settings()
    if s["maintenance"] and not is_admin(uid):
        await msg.reply_text(s.get("maintenance_text","🔧")); return

    # Sadece log — AI handle_text'te işlendi
    if not is_admin(uid) and text:
        log_user_message(user, "msg", text)

# ================================================================
#  MAIN
# ================================================================

def main():
    # Railway health check
    import threading
    from http.server import HTTPServer, BaseHTTPRequestHandler
    class _H(BaseHTTPRequestHandler):
        def do_GET(self):
            self.send_response(200); self.end_headers(); self.wfile.write(b"OK")
        def log_message(self, *a): pass
    _port = int(os.environ.get("PORT", 8080))
    threading.Thread(target=lambda: HTTPServer(("0.0.0.0", _port), _H).serve_forever(), daemon=True).start()
    logger.info(f"Health check port: {_port}")

    app = Application.builder().token(TOKEN).build()

    media_f = (filters.PHOTO | filters.VIDEO | filters.Document.ALL |
               filters.AUDIO | filters.VOICE | filters.ANIMATION | filters.Sticker.ALL)
    text_f  = filters.TEXT & ~filters.COMMAND

    import re
    try:
        escaped     = [re.escape(b) for b in ALL_BTNS]
        reply_btn_f = filters.Regex(f"^({'|'.join(escaped)})$")
    except Exception as e:
        logger.warning(f"reply_btn_f hatası: {e}")
        reply_btn_f = filters.Regex("^NOMATCH$")  # fallback

    conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(callback),
        ],
        states={
            WAIT_FOLDER:        [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_FILE:          [MessageHandler(text_f & ~reply_btn_f, handle_text),
                                 MessageHandler(media_f, handle_media), CallbackQueryHandler(callback)],
            WAIT_ADMIN_ID:      [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_RENAME_FOLDER: [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_RENAME_FILE:   [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_BOT_NAME:      [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_WELCOME:       [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_PHOTO:         [MessageHandler(media_f, handle_media), CallbackQueryHandler(callback)],
            WAIT_DM_USER_ID:    [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_DM_MSG:        [MessageHandler(text_f & ~reply_btn_f, handle_text),
                                 MessageHandler(media_f, handle_media), CallbackQueryHandler(callback)],
            WAIT_BROADCAST_MSG: [MessageHandler(text_f & ~reply_btn_f, handle_text),
                                 MessageHandler(media_f, handle_media), CallbackQueryHandler(callback)],
            WAIT_SEARCH:        [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_USER_MSG:      [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_POLL_QUESTION: [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_POLL_OPTIONS:  [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_POLL_COMMENT:  [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_WARN_REASON:   [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_USER_NOTE:     [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_FOLDER_DESC:   [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_FAQ_Q:         [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_FAQ_A:         [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_RULE_KW:       [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_RULE_RESP:     [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_SCHEDULE_HOURS:[MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_SCHEDULE_MSG:  [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_TAG:           [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
            WAIT_SECRET_ID:     [MessageHandler(text_f & ~reply_btn_f, handle_text), CallbackQueryHandler(callback)],
        },
        fallbacks=[
            CommandHandler("start", start),
            CommandHandler("menu",  menu_command),
            MessageHandler(reply_btn_f, handle_reply_buttons),
            MessageHandler(media_f, handle_media),
            CallbackQueryHandler(callback),
        ],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("menu",  menu_command))
    # ÖNEMLİ: Reply butonları conv'dan ÖNCE kayıtlı olmalı
    app.add_handler(MessageHandler(reply_btn_f, handle_reply_buttons))
    app.add_handler(conv)
    app.add_handler(MessageHandler(text_f & ~reply_btn_f, handle_text))
    app.add_handler(MessageHandler(media_f, handle_media))
    app.add_handler(MessageHandler(filters.ALL & ~filters.COMMAND, handle_any_message))

    print("✅ Bot başlatıldı!")
    print(f"👑 Süper Admin  : {ADMIN_ID}")
    print(f"💾 Veri dizini  : {BASE_DIR}")
    print("")
    print("📌 Yeni özellikler:")
    print("  🆕 Yeni kullanıcı bildirimi")
    print("  🗑 Silme onay ekranı")
    print("  📊 Dosya görüntüleme sayacı")
    print("  🔢 Klasör eleman sayısı")
    if "/data" in BASE_DIR:
        print("✅ Railway Volume aktif")
    else:
        print("⚠️  Volume YOK → Railway: Volumes → Mount /data")

    # ── Hatırlatıcı Kontrol Job'u ─────────────────
    async def _check_reminders(ctx):
        import time
        now  = time.time()
        rems = load_reminders()
        fired = []
        remaining = []
        for r in rems:
            if r.get("fire_ts", 0) <= now:
                fired.append(r)
            else:
                remaining.append(r)
        if fired:
            save_reminders(remaining)
            for r in fired:
                try:
                    uid_ = r.get("uid","")
                    lang = AR if not is_main_admin(uid_) else TR
                    msg  = lang["reminder_fired"].format(r.get("text",""))
                    await ctx.bot.send_message(int(uid_), msg)
                except Exception as e:
                    logger.warning(f"Hatırlatıcı gönderilemedi: {e}")

    async def _check_exam_reminders(ctx):
        """Sınavdan N gün önce kullanıcılara bildirim gönder."""
        from datetime import datetime as _dt, timedelta
        s = load_settings()
        remind_days = int(s.get("exam_remind_days", 1))
        target_date = (_dt.now() + timedelta(days=remind_days)).strftime("%d/%m/%Y")
        cds = load_countdowns()
        changed = False
        for cd in cds:
            if cd.get("date","") == target_date and not cd.get("notified"):
                name_cd = cd.get("name","Sinav")
                cls_cd  = cd.get("cls","")
                shft_cd = cd.get("shift","")
                tgt_val = f"cls_{cls_cd}" if cls_cd else "all"
                targets = get_target_users(tgt_val)
                if shft_cd:
                    shfts = load_shifts()
                    targets = [u for u in targets if shfts.get(u,"") in (shft_cd,"sabah" if shft_cd=="sabahi" else "gece")]
                day_lbl = "غداً" if remind_days == 1 else f"بعد {remind_days} أيام"
                msg = f"⏰ تذكير بالامتحان\n\n📚 {name_cd}\n📅 {day_lbl} — {target_date}"
                for uid_ in targets:
                    try: await ctx.bot.send_message(int(uid_), msg)
                    except: pass
                # Süper admin'e de bildir
                try:
                    await ctx.bot.send_message(ADMIN_ID,
                        f"📋 Sinav bildirimi gönderildi\nSinav: {name_cd}\nTarih: {target_date}\nKisi: {len(targets)}")
                except: pass
                cd["notified"] = True
                changed = True
        if changed: save_countdowns(cds)

    async def _check_lab_reminders(ctx):
        """Laboratuvar gününden N gün önce grubu bildir."""
        from datetime import datetime as _dt, timedelta, date as _date
        s_r = load_settings()
        lab_days = int(s_r.get("lab_remind_days", 1))
        tomorrow = (_dt.now() + timedelta(days=lab_days)).strftime("%Y-%m-%d")
        schedule = load_lab_schedule()
        for entry in schedule:
            if entry.get("week") == tomorrow and not entry.get("notified"):
                grp    = entry.get("group","?")
                note   = entry.get("note","")
                msg = (f"🔬 تذكير المختبر\n\n"
                       f"📅 غداً يدخل: {grp}\n"
                       + (f"📝 {note}" if note else ""))
                # Bu gruba ait kullanıcılara gönder
                grps_d = load_groups()
                users_d = load_users()
                for uid_, u in users_d.items():
                    if is_blocked(uid_): continue
                    if grps_d.get(uid_,"") == grp:
                        try: await ctx.bot.send_message(int(uid_), msg)
                        except: pass
                # Süper admin + adminlere bildir
                try: await ctx.bot.send_message(ADMIN_ID, f"🔬 Lab yarın: {grp}\n{msg}")
                except: pass
                admins_list = load_admins()
                for adm in admins_list:
                    try: await ctx.bot.send_message(adm, f"🔬 Lab yarın: {grp}")
                    except: pass
                entry["notified"] = True
        save_lab_schedule(schedule)

    job_queue = app.job_queue
    if job_queue:
        job_queue.run_repeating(_check_reminders, interval=30, first=5)
        job_queue.run_repeating(_check_exam_reminders, interval=3600, first=30)
        job_queue.run_repeating(_check_lab_reminders, interval=3600, first=60)
        print("✅ Hatırlatıcı job başlatıldı (60sn aralık)")

    # Mevcut webhook varsa sil — run_polling başlamadan önce
    app.run_polling(
        drop_pending_updates=True,
        allowed_updates=["message","callback_query","my_chat_member"],
    )

if __name__ == "__main__":
    main()
